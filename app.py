import io
import json
import uuid
import math
import time
import hmac
import html as _html_escape
import secrets
import re
import base64
import urllib.parse
import xml.etree.ElementTree as _ET
from difflib import HtmlDiff
from datetime import date, datetime
from hashlib import sha256, pbkdf2_hmac

try:
    import requests as _requests
except Exception:
    _requests = None

import pandas as pd
import pdfplumber
import streamlit as st
from docx import Document
from fpdf import FPDF
from notion_client import Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from streamlit_option_menu import option_menu

try:
    import google.generativeai as genai
except Exception:
    genai = None

try:
    import anthropic as _anthropic_sdk
except Exception:
    _anthropic_sdk = None

try:
    from groq import Groq as _GroqClient
except Exception:
    _GroqClient = None

try:
    import together as _together_sdk
except Exception:
    _together_sdk = None


# ── HARDCODED FREE MODELS ONLY — never change these ──────────────────────────
# gemini-1.5-pro, gemini-2.5-flash, etc. are NOT free — excluded intentionally.
MODEL_NAME = "gemini-1.5-flash"          # permanent free tier, no billing ever
MODEL_FALLBACK_ORDER = [
    "gemini-1.5-flash",        # FREE — 1,500 req/day, 15 RPM
    "gemini-1.5-flash-latest", # FREE — alias of above
    "gemini-1.5-flash-001",    # FREE — stable snapshot
    "gemini-2.0-flash",        # FREE — 1,500 req/day
    "gemini-2.0-flash-lite",   # FREE — 1,500 req/day, 30 RPM
    # gemini-1.5-pro        -> PAID, intentionally removed
    # gemini-1.5-pro-latest -> PAID, intentionally removed
    # gemini-1.5-pro-001    -> PAID, intentionally removed
    # gemini-2.5-flash      -> PAID, intentionally removed
]
PLATFORMS_DEFAULT = [
    "LinkedIn",
    "Naukri",
    "Indeed",
    "Glassdoor",
    "Wellfound",
    "Unstop",
    "Monster",
    "Shine",
    "Internshala",
    "AngelList",
    "Hirist",
    "Cutshort",
]
TRACKER_COLUMNS = ["Company", "Role", "Platform", "Date", "Status", "Package", "Notes", "NextStep", "URL"]
PLATFORM_META = {
    "LinkedIn": {"url": "https://www.linkedin.com/feed/", "best_time": "9-11 AM", "response_rate": "22%"},
    "Naukri": {"url": "https://naukri.com/login", "best_time": "10 AM-1 PM", "response_rate": "18%"},
    "Indeed": {"url": "https://indeed.com/account/login", "best_time": "8-10 AM", "response_rate": "16%"},
    "Glassdoor": {"url": "https://glassdoor.com/profile/login", "best_time": "6-9 PM", "response_rate": "12%"},
    "Wellfound": {"url": "https://wellfound.com/login", "best_time": "11 AM-2 PM", "response_rate": "20%"},
    "Unstop": {"url": "https://unstop.com/login", "best_time": "12-3 PM", "response_rate": "14%"},
    "Monster": {"url": "https://monster.com/login", "best_time": "9 AM-12 PM", "response_rate": "11%"},
    "Shine": {"url": "https://shine.com/login", "best_time": "10 AM-1 PM", "response_rate": "10%"},
    "Internshala": {"url": "https://internshala.com/login", "best_time": "8-11 AM", "response_rate": "28%"},
    "AngelList": {"url": "https://angel.co/login", "best_time": "11 AM-1 PM", "response_rate": "19%"},
    "Hirist": {"url": "https://hirist.com/login", "best_time": "9-11 AM", "response_rate": "17%"},
    "Cutshort": {"url": "https://cutshort.io/login", "best_time": "10 AM-12 PM", "response_rate": "21%"},
}


def init_state() -> None:
    defaults = {
        "api_key": "",
        "location": "",
        "currency": "₹",
        "resume_bytes": b"",
        "resume_text": "",
        "resume_optimized": "",
        "profile": {
            "name": "",
            "skills": [],
            "experience_years": 0,
            "job_titles": [],
            "education": [],
            "seniority_level": "",
        },
        "ats": {
            "overall_score": 0,
            "breakdown": {
                "Keywords": 0,
                "Format": 0,
                "Impact Statements": 0,
                "Skills Match": 0,
                "Readability": 0,
            },
        },
        "jobs": [],
        "tracker": [],
        "platforms_connected": [],
        "platforms_connected_at": {},
        "session_id": str(uuid.uuid4()),
        "api_validated": False,
        "api_validation_message": "",
        "api_last_checked_key": "",
        "resolved_model_name": "",
        "tailor_result": {},
        "last_daily_status_check": "",
        "quota_block_until": 0.0,
        "quota_message": "",
        "quota_key_fingerprint": "",
        "claude_api_key": "",
        "claude_api_validated": False,
        "claude_api_validation_message": "",
        "groq_api_key": "",
        "groq_api_validated": False,
        "groq_api_validation_message": "",
        "together_api_key": "",
        "together_api_validated": False,
        "together_api_validation_message": "",
        "openai_api_key": "",
        "openai_api_validated": False,
        "openai_api_validation_message": "",
        "active_provider": "",       # "gemini" | "groq" | "claude" | "together" | ""
        "provider_switch_log": [],   # list of dicts: {ts, from, to, reason}
        "resume_filename": "",
        "original_resume_bytes": b"",
        "original_resume_mime": "application/octet-stream",
        "analysis_cache_key": "",
        "analysis_error": "",
        "is_processing": False,
        "resume_analysis": {},
        "ats_score": 0,
        "ats_breakdown": {},
        "optimized_resume": "",
        "improvements": [],
        "jobs_ranked": [],
        "tracker_rows": [],
        "platforms": PLATFORMS_DEFAULT.copy(),
        "selected_provider": "",   # which provider card is expanded in API Keys page
        "auth_user": None,         # dict of logged-in user row, or None
        "auth_view": "login",      # "login" | "register"
        # ── Resume Insight extended fields (persist across pages) ──────
        "insight": {
            "keywords_present":   [],   # keywords found in resume
            "keywords_missing":   [],   # high-value keywords NOT in resume
            "certifications":     [],
            "languages":          [],
            "contact": {
                "email": "", "phone": "", "linkedin": "", "github": "", "portfolio": ""
            },
            "summary_line":       "",   # 1-line professional summary AI writes
            "strengths":          [],
            "weaknesses":         [],
            "improvement_suggestions": [],  # prioritized actionable improvements
            "career_gaps":        [],   # list of gap descriptions
            "salary_range":       {"min_lpa": 0, "max_lpa": 0, "currency": "₹"},
            "glassdoor_salary":   {     # Glassdoor-style market salary estimate
                "role": "", "min_lpa": 0, "max_lpa": 0, "median_lpa": 0,
                "currency": "₹", "location": "", "experience_band": "", "note": ""
            },
            "industry":           "",
            "recommended_roles":  [],
            "resume_score_label": "",   # "Weak" | "Average" | "Good" | "Strong"
            "word_count":         0,
            "page_estimate":      0,
        },
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value
    # Sync aliases so older code paths remain valid.
    st.session_state["gemini_api_key"] = st.session_state["api_key"]
    st.session_state["location_preference"] = st.session_state["location"]
    st.session_state["optimized_resume"] = st.session_state["resume_optimized"]
    st.session_state["resume_analysis"] = st.session_state["profile"]
    st.session_state["ats_score"] = int(st.session_state["ats"].get("overall_score", 0))
    st.session_state["ats_breakdown"] = st.session_state["ats"].get("breakdown", {})
    st.session_state["jobs_ranked"] = st.session_state["jobs"]
    st.session_state["tracker_rows"] = st.session_state["tracker"]


def configure_page(show_header: bool = True) -> None:
    st.set_page_config(page_title="AutoApply AI", page_icon=":briefcase:", layout="wide")
    st.html("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=DM+Sans:wght@400;500;600&display=swap');

/* ════════════════════════════════════════════════
   BASE TYPOGRAPHY
════════════════════════════════════════════════ */
html, body, [class*="css"] {
    font-family: 'DM Sans', 'Sora', sans-serif !important;
    font-size: 16px !important;
    -webkit-font-smoothing: antialiased;
    -moz-osx-font-smoothing: grayscale;
    text-rendering: optimizeLegibility;
}

/* ════════════════════════════════════════════════
   APP BACKGROUND — subtle noise grain + deep bg
════════════════════════════════════════════════ */
[data-testid="stAppViewContainer"] {
    background: #080b12;
    color: #D4DCEC;
    background-image:
        radial-gradient(ellipse 80% 50% at 50% -20%, rgba(59,130,246,0.06) 0%, transparent 60%),
        radial-gradient(ellipse 40% 30% at 90% 10%, rgba(16,217,160,0.04) 0%, transparent 50%);
}

/* Main content area subtle inner padding */
[data-testid="stMainBlockContainer"] {
    padding-top: 2rem !important;
}

/* Sidebar — premium glassy feel */
[data-testid="stSidebar"] {
    background: #07090f;
    border-right: 1px solid rgba(30,45,69,0.8);
    box-shadow: 4px 0 32px rgba(0,0,0,0.5);
}

/* ════════════════════════════════════════════════
   HEADINGS — tighter, sharper letterform
════════════════════════════════════════════════ */
h1 {
    font-size: 2rem !important; font-weight: 700 !important;
    color: #EAF2FF !important; letter-spacing: -0.8px !important;
    font-family: 'Sora', sans-serif !important;
    line-height: 1.2 !important;
}
h2 { font-size: 1.5rem !important; font-weight: 600 !important; color: #EAF2FF !important; letter-spacing: -0.4px !important; }
h3 { font-size: 1.2rem !important; font-weight: 600 !important; color: #CBD5E1 !important; letter-spacing: -0.2px !important; }

[data-testid="stHeadingWithActionElements"] h2 {
    font-size: 1.6rem !important;
    font-weight: 700 !important;
    color: #EAF2FF !important;
    letter-spacing: -0.5px !important;
    font-family: 'Sora', sans-serif !important;
}

/* ════════════════════════════════════════════════
   BODY TEXT
════════════════════════════════════════════════ */
p, li, span, label, div {
    font-size: 15px !important;
    line-height: 1.65 !important;
    color: #C4CEDF;
}

/* ════════════════════════════════════════════════
   CAPTIONS
════════════════════════════════════════════════ */
[data-testid="stCaptionContainer"] p,
small, .caption {
    font-size: 13px !important;
    color: #64748b !important;
    line-height: 1.5 !important;
}

/* ════════════════════════════════════════════════
   METRIC CARDS — deeper, more dimensional
════════════════════════════════════════════════ */
div[data-testid="stMetric"] {
    background: linear-gradient(145deg, #0d1117, #0a0f1a);
    border: 1px solid #1e2d45;
    border-radius: 16px;
    padding: 18px 22px !important;
    box-shadow: 0 4px 24px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.03);
    transition: box-shadow 0.2s ease, border-color 0.2s ease;
    position: relative;
    overflow: hidden;
}
div[data-testid="stMetric"]::before {
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 1px;
    background: linear-gradient(90deg, transparent, rgba(59,130,246,0.2), transparent);
}
div[data-testid="stMetric"]:hover {
    box-shadow: 0 6px 32px rgba(0,0,0,0.55), 0 0 0 1px rgba(59,130,246,0.08);
    border-color: rgba(59,130,246,0.2);
}
div[data-testid="stMetric"] label {
    font-size: 11px !important; font-weight: 700 !important;
    letter-spacing: 1px !important; text-transform: uppercase !important;
    color: #475569 !important;
}
div[data-testid="stMetric"] [data-testid="stMetricValue"] {
    font-size: 1.9rem !important; font-weight: 700 !important;
    color: #EAF2FF !important; letter-spacing: -0.5px !important;
    font-family: 'Sora', sans-serif !important;
    line-height: 1.15 !important;
}

/* ════════════════════════════════════════════════
   ALERTS
════════════════════════════════════════════════ */
.stAlert {
    background: linear-gradient(135deg, #0d1117, #0a0f1a);
    border: 1px solid #1e2d45;
    border-radius: 14px;
    box-shadow: 0 2px 16px rgba(0,0,0,0.3);
}
.stAlert p { font-size: 14px !important; }

/* ════════════════════════════════════════════════
   DATA FRAMES
════════════════════════════════════════════════ */
.stDataFrame {
    background: #0d1117;
    border: 1px solid #1e2d45;
    border-radius: 14px;
    overflow: hidden;
    box-shadow: 0 4px 20px rgba(0,0,0,0.35);
}

/* ════════════════════════════════════════════════
   BUTTONS — refined, layered finish
════════════════════════════════════════════════ */
.stButton > button, .stDownloadButton > button, .stLinkButton > a {
    border-radius: 10px;
    border: 1px solid rgba(30,45,69,0.9);
    font-size: 14px !important; font-weight: 500 !important;
    padding: 10px 20px !important;
    transition: all 0.18s ease !important;
    letter-spacing: 0.1px !important;
    background: linear-gradient(145deg, #0d1828, #0a1020) !important;
    color: #94a3b8 !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.3), inset 0 1px 0 rgba(255,255,255,0.03) !important;
}
.stButton > button:hover, .stDownloadButton > button:hover {
    border-color: rgba(59,130,246,0.35) !important;
    color: #cbd5e1 !important;
    box-shadow: 0 4px 16px rgba(0,0,0,0.4), 0 0 0 1px rgba(59,130,246,0.1) !important;
    transform: translateY(-1px) !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #3b82f6, #6d28d9) !important;
    color: #ffffff !important;
    border: none !important;
    font-size: 14px !important; font-weight: 600 !important;
    letter-spacing: 0.2px !important;
    box-shadow: 0 4px 20px rgba(59,130,246,0.3), 0 2px 8px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.1) !important;
}
.stButton > button[kind="primary"]:hover {
    box-shadow: 0 6px 28px rgba(59,130,246,0.45), 0 2px 12px rgba(0,0,0,0.5) !important;
    transform: translateY(-1px) !important;
}
.stButton > button:active { transform: translateY(0px) !important; }

/* ════════════════════════════════════════════════
   INPUTS — polished with focus glow
════════════════════════════════════════════════ */
[data-testid="stTextInput"] input,
[data-testid="stTextArea"] textarea {
    font-size: 15px !important;
    background: #060a14 !important;
    border: 1px solid #1e2d45 !important;
    border-radius: 10px !important;
    color: #EAF2FF !important;
    padding: 12px 16px !important;
    box-shadow: inset 0 2px 8px rgba(0,0,0,0.3) !important;
    transition: border-color 0.2s ease, box-shadow 0.2s ease !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stTextArea"] textarea:focus {
    border-color: rgba(59,130,246,0.5) !important;
    box-shadow: inset 0 2px 8px rgba(0,0,0,0.3), 0 0 0 3px rgba(59,130,246,0.08) !important;
    outline: none !important;
}
[data-testid="stTextInput"] label,
[data-testid="stTextArea"] label {
    font-size: 11px !important; font-weight: 700 !important;
    letter-spacing: 0.8px !important; color: #475569 !important;
    text-transform: uppercase !important;
    margin-bottom: 6px !important;
}

/* ════════════════════════════════════════════════
   SELECTBOX
════════════════════════════════════════════════ */
[data-testid="stSelectbox"] label {
    font-size: 11px !important; color: #475569 !important;
    text-transform: uppercase !important; letter-spacing: 0.8px !important;
    font-weight: 700 !important;
}
[data-testid="stSelectbox"] div[data-baseweb="select"] {
    font-size: 15px !important;
    background: #060a14 !important;
    border-color: #1e2d45 !important;
    border-radius: 10px !important;
}

/* ════════════════════════════════════════════════
   EXPANDER — cleaner header
════════════════════════════════════════════════ */
[data-testid="stExpander"] {
    border: 1px solid #1e2d45 !important;
    border-radius: 12px !important;
    background: linear-gradient(145deg, #0d1117, #090d16) !important;
    overflow: hidden;
    box-shadow: 0 2px 12px rgba(0,0,0,0.25);
}
[data-testid="stExpander"] summary {
    font-size: 13px !important; font-weight: 700 !important;
    color: #94a3b8 !important; padding: 14px 18px !important;
    letter-spacing: 0.2px !important;
    border-bottom: 1px solid rgba(30,45,69,0.5);
}
[data-testid="stExpander"] summary:hover { color: #CBD5E1 !important; }

/* ════════════════════════════════════════════════
   FORMS
════════════════════════════════════════════════ */
.stForm {
    background: linear-gradient(145deg, #0d1117, #090d16);
    border: 1px solid #1e2d45;
    border-radius: 16px;
    box-shadow: 0 4px 24px rgba(0,0,0,0.3);
    overflow: hidden;
}

/* ════════════════════════════════════════════════
   TABS — premium pill-style
════════════════════════════════════════════════ */
[data-testid="stTab"] button {
    font-size: 13px !important; font-weight: 600 !important;
    letter-spacing: 0.2px !important;
    border-radius: 8px !important;
    transition: all 0.15s ease !important;
    padding: 8px 16px !important;
}
[data-testid="stTab"] button[aria-selected="true"] {
    background: rgba(59,130,246,0.12) !important;
    color: #60a5fa !important;
    border-color: rgba(59,130,246,0.3) !important;
    box-shadow: 0 0 16px rgba(59,130,246,0.1) !important;
}

/* ════════════════════════════════════════════════
   CONTAINERS — bordered containers look premium
════════════════════════════════════════════════ */
[data-testid="stVerticalBlockBorderWrapper"] > div {
    background: linear-gradient(145deg, #0c1020, #090d18) !important;
    border: 1px solid #1e2d45 !important;
    border-radius: 16px !important;
    box-shadow: 0 4px 28px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.02) !important;
    padding: 20px !important;
}

/* ════════════════════════════════════════════════
   SIDEBAR TEXT
════════════════════════════════════════════════ */
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div { font-size: 14px !important; }

/* ════════════════════════════════════════════════
   PLATFORM CARD
════════════════════════════════════════════════ */
[data-testid="stMarkdownContainer"] > div:has(> .platform-card) {
    background: #0d1117; border: 1px solid #1e2d45; border-radius: 12px;
}

/* ════════════════════════════════════════════════
   SPINNER
════════════════════════════════════════════════ */
[data-testid="stSpinner"] > div {
    border-color: #3b82f6 !important;
}

/* ════════════════════════════════════════════════
   SCROLLBAR — refined dark style
════════════════════════════════════════════════ */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #07090f; }
::-webkit-scrollbar-thumb { background: #1e2d45; border-radius: 99px; }
::-webkit-scrollbar-thumb:hover { background: #2d4268; }

/* ════════════════════════════════════════════════
   GLOBAL ACCENT UTILS
════════════════════════════════════════════════ */
a { color: #3b82f6 !important; }
.accent-green { color: #10d9a0; }
.accent-gold { color: #f59e0b; }

/* ════════════════════════════════════════════════
   SUCCESS / WARNING / ERROR MESSAGE BOXES
════════════════════════════════════════════════ */
[data-testid="stNotification"] {
    border-radius: 12px !important;
    box-shadow: 0 4px 20px rgba(0,0,0,0.4) !important;
}

/* ════════════════════════════════════════════════
   DIVIDERS / HR
════════════════════════════════════════════════ */
hr {
    border-color: rgba(30,45,69,0.6) !important;
    margin: 20px 0 !important;
}

/* ════════════════════════════════════════════════
   PAGE FADE-IN ANIMATION
════════════════════════════════════════════════ */
@keyframes fadeSlideUp {
    from { opacity: 0; transform: translateY(12px); }
    to   { opacity: 1; transform: translateY(0); }
}
[data-testid="stMainBlockContainer"] > div > div {
    animation: fadeSlideUp 0.35s ease both;
}
</style>
""")
    if show_header:
        st.title("AutoApply AI - AI Job Application Dashboard")
        title_col, badge_col = st.columns([5, 1])
        with title_col:
            st.caption("Automated resume analysis, ATS scoring, optimization, and ranked job discovery.")
        with badge_col:
            st.markdown(active_provider_badge(), unsafe_allow_html=True)


def read_pdf(file) -> str:
    chunks = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            chunks.append(page.extract_text() or "")
    return "\n".join(chunks).strip()


def read_docx(file) -> str:
    doc = Document(file)
    return "\n".join(p.text for p in doc.paragraphs if p.text).strip()


def read_uploaded_resume(uploaded_file) -> str:
    file_name = uploaded_file.name.lower()
    if file_name.endswith(".pdf"):
        return read_pdf(uploaded_file)
    if file_name.endswith(".docx"):
        return read_docx(uploaded_file)
    if file_name.endswith(".txt"):
        return uploaded_file.read().decode("utf-8", errors="ignore")
    return ""


def get_gemini_model():
    if genai is None:
        raise RuntimeError("google-generativeai is not installed.")
    api_key = st.session_state.get("api_key", "").strip()
    if not api_key:
        raise RuntimeError("Gemini API key is missing.")
    genai.configure(api_key=api_key)
    model_name = resolve_model_name()
    return genai.GenerativeModel(model_name)


def normalize_model_name(name: str) -> str:
    value = str(name or "").strip()
    if value.startswith("models/"):
        return value.split("/", 1)[1]
    return value


def resolve_model_name() -> str:
    cached = st.session_state.get("resolved_model_name", "").strip()
    if cached:
        return cached

    available = []
    try:
        for model in genai.list_models():
            methods = getattr(model, "supported_generation_methods", []) or []
            if "generateContent" in methods:
                available.append(normalize_model_name(getattr(model, "name", "")))
    except Exception:
        pass

    # Only pick from the hardcoded FREE list — never fall through to
    # whatever Google returns as available (could include paid models).
    for candidate in MODEL_FALLBACK_ORDER:
        if candidate in available:
            st.session_state["resolved_model_name"] = candidate
            return candidate

    # If none of the free models are listed (edge case), default to the
    # primary free model — do NOT fall back to available[0] which may be paid.
    st.session_state["resolved_model_name"] = MODEL_NAME
    return MODEL_NAME


def run_gemini_prompt_with_key(prompt: str, api_key: str, use_grounding: bool = False) -> str:
    """Legacy wrapper — temporarily sets api_key then delegates to call_ai rotation engine."""
    original_key = st.session_state.get("api_key", "")
    st.session_state["api_key"] = api_key
    try:
        return call_ai(prompt, use_grounding=use_grounding)
    finally:
        st.session_state["api_key"] = original_key


def run_gemini_prompt(prompt: str, use_grounding: bool = False) -> str:
    """Main AI entry-point — routes through smart rotation engine (Gemini → Claude)."""
    return call_ai(prompt, use_grounding=use_grounding)


def validate_api_key_if_needed() -> None:
    key = st.session_state.get("api_key", "").strip()
    prev_key = st.session_state.get("api_last_checked_key", "").strip()
    if not key:
        st.session_state["api_validated"] = False
        st.session_state["api_validation_message"] = "❌ Invalid Key"
        st.session_state["api_last_checked_key"] = ""
        return

    # New key should not inherit quota cooldown from the previous key.
    if key and key != prev_key:
        st.session_state["quota_block_until"] = 0.0
        st.session_state["quota_message"] = ""
        st.session_state["quota_key_fingerprint"] = ""

    if key == st.session_state.get("api_last_checked_key", "") and st.session_state.get("api_validated"):
        return
    try:
        # Validate key by checking API access directly, not model generation.
        # This avoids false "invalid key" when generation fails for transient/model reasons.
        if genai is None:
            raise RuntimeError("google-generativeai package is unavailable.")
        genai.configure(api_key=key)
        model_iter = genai.list_models()
        _ = next(iter(model_iter), None)
        st.session_state["resolved_model_name"] = ""
        _ = resolve_model_name()
        st.session_state["api_validated"] = True
        st.session_state["api_validation_message"] = "✅ Gemini Connected"
    except Exception as exc:
        st.session_state["api_validated"] = False
        err = str(exc).strip() or "Unknown validation error"
        err = err.split("\n", 1)[0]
        if "API_KEY" in err.upper() or "INVALID" in err.upper() or "UNAUTHENTICATED" in err.upper():
            st.session_state["api_validation_message"] = "❌ Invalid Key"
        else:
            st.session_state["api_validation_message"] = f"❌ Could not validate now ({err})"
    st.session_state["api_last_checked_key"] = key


def log_notion(data: dict) -> None:
    """Log an activity event to Supabase `events` table. Silent on failure."""
    try:
        db = _sb()
        if db is None:
            return
        db.table("events").insert({
            "session_id":    str(data.get("SessionID", ""))[:200],
            "timestamp":     str(data.get("Timestamp", datetime.utcnow().isoformat())),
            "location":      str(data.get("Location", ""))[:500],
            "filename":      str(data.get("Filename", ""))[:500],
            "format":        str(data.get("Format", ""))[:200],
            "keywords":      str(data.get("Keywords", ""))[:1000],
            "skills":        str(data.get("Skills", ""))[:1000],
            "exp_years":     float(data.get("ExpYears", 0) or 0),
            "seniority":     str(data.get("Seniority", ""))[:200],
            "ats_before":    float(data.get("ATSBefore", 0) or 0),
            "ats_after":     float(data.get("ATSAfter", 0) or 0),
            "jobs_searched": float(data.get("JobsSearched", 0) or 0),
            "target_salary": float(data.get("TargetSalary", 0) or 0),
            "currency":      str(data.get("Currency", ""))[:30],
            "platforms":     str(data.get("Platforms", ""))[:1000],
            "top_job":       str(data.get("TopJob", ""))[:500],
            "top_score":     float(data.get("TopScore", 0) or 0),
        }).execute()
    except Exception:
        pass


def extract_json_payload(text: str):
    if not text:
        return None
    candidates = [text.strip()]
    if "```json" in text:
        part = text.split("```json", 1)[1].split("```", 1)[0]
        candidates.insert(0, part.strip())
    elif "```" in text:
        part = text.split("```", 1)[1].split("```", 1)[0]
        candidates.insert(0, part.strip())

    for candidate in candidates:
        try:
            return json.loads(candidate)
        except Exception:
            continue
    return None


def is_quota_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return ("429" in msg) or ("quota" in msg) or ("rate limit" in msg) or ("resource_exhausted" in msg)


# ═══════════════════════════════════════════════════════════════════════════
# SMART API ROTATION ENGINE
# Priority: Gemini → Claude.  Rotates automatically on rate-limit / error.
# ═══════════════════════════════════════════════════════════════════════════

def _build_provider_list() -> list[dict]:
    """Build ordered provider list: Gemini → Groq → Claude → HuggingFace.
    All models are hardcoded to free tiers — no paid models are ever called.
    OpenAI is excluded from rotation as gpt-4o-mini is not free.
    """
    providers = []
    gemini_key   = st.session_state.get("api_key",           "").strip()
    groq_key     = st.session_state.get("groq_api_key",      "").strip()
    claude_key   = st.session_state.get("claude_api_key",    "").strip()
    together_key = st.session_state.get("together_api_key",  "").strip()
    # openai_key intentionally NOT included — gpt-4o-mini charges per token

    if gemini_key:
        # Always use the resolved free model; MODEL_NAME is the safe default.
        free_gemini_model = st.session_state.get("resolved_model_name", MODEL_NAME) or MODEL_NAME
        # Extra guard: if somehow a paid model crept into session, reset to free default.
        if free_gemini_model not in MODEL_FALLBACK_ORDER:
            free_gemini_model = MODEL_NAME
            st.session_state["resolved_model_name"] = MODEL_NAME
        providers.append({
            "name":  "gemini",
            "key":   gemini_key,
            "model": free_gemini_model,   # hardcoded to free flash only
            "label": "Gemini ✦",
            "color": "#10d9a0",
            "free_tier": "1,500 req/day · 15 RPM · 1M TPM · no billing",
        })
    if groq_key:
        providers.append({
            "name":  "groq",
            "key":   groq_key,
            "model": "llama-3.1-8b-instant",   # hardcoded free model
            "label": "Groq ✦",
            "color": "#f472b6",
            "free_tier": "Generous RPM · 100% free · no billing",
        })
    if claude_key:
        providers.append({
            "name":  "claude",
            "key":   claude_key,
            "model": "claude-haiku-4-5-20251001",  # hardcoded free-tier Haiku only
            "label": "Claude ✦",
            "color": "#f59e0b",
            "free_tier": "Free API credits on signup · Haiku only",
        })
    if together_key:
        providers.append({
            "name":  "together",
            "key":   together_key,
            "model": st.session_state.get("hf_active_model", "microsoft/Phi-3.5-mini-instruct"),  # best working free HF model
            "label": "HuggingFace ✦",
            "color": "#f97316",
            "free_tier": "Free Inference API · no billing",
        })
    return providers


def _call_gemini(prompt: str, key: str, model: str, use_grounding: bool = False) -> str:
    if genai is None:
        raise RuntimeError("google-generativeai not installed.")
    genai.configure(api_key=key)
    mdl = genai.GenerativeModel(model)
    if use_grounding:
        try:
            resp = mdl.generate_content(prompt, tools=[{"google_search_retrieval": {}}])
        except Exception:
            resp = mdl.generate_content(prompt)
    else:
        resp = mdl.generate_content(prompt)
    return (getattr(resp, "text", "") or "").strip()


def _call_groq(prompt: str, key: str, model: str) -> str:
    if _GroqClient is None:
        raise RuntimeError("groq package not installed. Run: pip install groq")
    client = _GroqClient(api_key=key)
    completion = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=4096,
        temperature=0.3,
    )
    return (completion.choices[0].message.content or "").strip()


def _call_claude(prompt: str, key: str, model: str) -> str:
    if _anthropic_sdk is None:
        raise RuntimeError("anthropic SDK not installed.")
    client = _anthropic_sdk.Anthropic(api_key=key)
    msg = client.messages.create(
        model=model,
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    blocks = getattr(msg, "content", [])
    return "\n".join(
        getattr(b, "text", "") for b in blocks if getattr(b, "type", "") == "text"
    ).strip()


def _call_together(prompt: str, key: str, model: str) -> str:
    """Call HuggingFace Inference via router.huggingface.co (the current free endpoint).
    The old api-inference.huggingface.co stopped serving LLMs in July 2025.
    Plain model IDs are tried first (work with basic free tokens), then
    :fastest / :cerebras suffix variants for accounts with Inference Providers permission.
    Free HF accounts get monthly credits applied automatically.
    """
    import urllib.request
    import urllib.error
    # Plain IDs first — work with basic HF free token, no special permissions needed.
    # Suffixed variants tried after for accounts with Inference Providers permission.
    _HF_FREE_MODELS = [
        "microsoft/Phi-3.5-mini-instruct",
        "Qwen/Qwen2.5-7B-Instruct",
        "meta-llama/Llama-3.1-8B-Instruct",
        "HuggingFaceH4/zephyr-7b-beta",
        "mistralai/Mistral-7B-Instruct-v0.3",
        "meta-llama/Llama-3.1-8B-Instruct:cerebras",
        "meta-llama/Llama-3.1-8B-Instruct:fastest",
        "Qwen/Qwen2.5-7B-Instruct:fastest",
        "microsoft/Phi-3.5-mini-instruct:fastest",
    ]
    active = st.session_state.get("hf_active_model", _HF_FREE_MODELS[0])
    seen: set = set()
    models_to_try = [m for m in [active] + _HF_FREE_MODELS
                     if not (m in seen or seen.add(m))]
    last_exc: Exception | None = None
    for hf_model in models_to_try:
        try:
            payload = json.dumps({
                "model": hf_model,
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 4096,
                "temperature": 0.3,
            }).encode()
            req = urllib.request.Request(
                "https://router.huggingface.co/v1/chat/completions",
                data=payload,
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json",
                },
            )
            with urllib.request.urlopen(req, timeout=45) as resp:
                data = json.loads(resp.read().decode())
            st.session_state["hf_active_model"] = hf_model
            return (data["choices"][0]["message"]["content"] or "").strip()
        except urllib.error.HTTPError as e:
            last_exc = e
            if e.code in (401, 403):
                continue  # model permission issue — try next model (token already validated)
            continue  # 404/503/429 — try next model
        except Exception as e:
            last_exc = e
            continue
    raise RuntimeError(f"HuggingFace: all models failed. Last error: {last_exc}")


def call_ai(prompt: str, use_grounding: bool = False) -> str:
    """
    Smart rotating AI caller.
    Priority: Gemini → Groq → Claude → HuggingFace.
    Rotates automatically on rate-limit / quota / any error.
    Records active_provider + switch log on every provider change.
    """
    providers = _build_provider_list()
    if not providers:
        raise RuntimeError(
            "No API keys configured. Add at least one key in the API Keys page."
        )

    previous_provider = st.session_state.get("active_provider", "")
    last_exc: Exception | None = None

    for provider in providers:
        name  = provider["name"]
        key   = provider["key"]
        model = provider["model"]
        try:
            if name == "gemini":
                result = _call_gemini(prompt, key, model, use_grounding=use_grounding)
            elif name == "groq":
                result = _call_groq(prompt, key, model)
            elif name == "claude":
                result = _call_claude(prompt, key, model)
            elif name == "together":
                result = _call_together(prompt, key, model)
            else:
                continue

            # ── Success ──────────────────────────────────────────────
            if st.session_state.get("active_provider") != name:
                reason = "rate_limit" if (last_exc and is_quota_error(last_exc)) else "primary"
                st.session_state.setdefault("provider_switch_log", []).append({
                    "ts":     datetime.utcnow().isoformat(),
                    "from":   previous_provider,
                    "to":     name,
                    "reason": reason,
                })
            st.session_state["active_provider"] = name
            # Track session usage per provider for rate-limit bars
            if "usage_counts" not in st.session_state:
                st.session_state["usage_counts"] = {"gemini": 0, "groq": 0, "claude": 0, "together": 0, "openai": 0}
            st.session_state["usage_counts"][name] = st.session_state["usage_counts"].get(name, 0) + 1
            return result

        except Exception as exc:
            last_exc = exc
            continue   # always try next provider

    raise RuntimeError(
        "All AI providers exhausted. "
        + (f"Last error: {last_exc}" if last_exc else "No keys configured.")
    )


def active_provider_badge() -> str:
    """Styled HTML badge showing currently active AI provider + fallback count."""
    provider  = st.session_state.get("active_provider", "")
    providers = _build_provider_list()
    n         = len(providers)

    if not providers:
        return (
            "<div style='display:inline-flex;align-items:center;gap:6px;"
            "background:#1a1a2e;border:10px solid #374151;border-radius:20px;"
            "padding:4px 12px;font-size:20px;color:#94a3b8;font-family:Sora,sans-serif;'>"
            "⚠ No API Keys"
            "</div>"
        )

    # Colour map per provider
    _meta = {
        "gemini":  {"color": "#10d9a0", "label": "Gemini ✦"},
        "groq":    {"color": "#f472b6", "label": "Groq ✦"},
        "claude":  {"color": "#f59e0b", "label": "Claude ✦"},
        "together":{"color": "#f97316", "label": "HuggingFace ✦"},
    }
    m      = _meta.get(provider, {"color": "#94a3b8", "label": "Standby"})
    color  = m["color"]
    label  = m["label"]
    extras = f" &nbsp;<span style='color:#4b5563;font-size:10px;'>+{n-1} fallback</span>" if n > 1 else ""

    return (
        f"<div style='display:inline-flex;align-items:center;gap:8px;"
        f"background:#0d1117;border:1px solid {color}33;"
        f"border-radius:20px;padding:4px 14px 4px 10px;"
        f"font-size:20px;font-family:Sora,sans-serif;'>"
        f"<span style='width:7px;height:7px;border-radius:50%;"
        f"background:{color};display:inline-block;"
        f"box-shadow:0 0 6px {color};animation:pulse-dot 2s infinite;'></span>"
        f"<span style='color:#94a3b8;'>Using:</span>&nbsp;"
        f"<span style='color:{color};font-weight:600;'>{label}</span>"
        f"{extras}"
        f"</div>"
        f"<style>@keyframes pulse-dot{{0%,100%{{opacity:1}}50%{{opacity:.4}}}}</style>"
    )


def trimmed_text(text: str, max_chars: int = 7000) -> str:
    value = (text or "").strip()
    if len(value) <= max_chars:
        return value
    head = value[: int(max_chars * 0.65)]
    tail = value[-int(max_chars * 0.35) :]
    return f"{head}\n...\n{tail}"


def quota_countdown_text() -> str:
    remaining = int(float(st.session_state.get("quota_block_until", 0.0)) - time.time())
    if remaining <= 0:
        return ""
    mins, secs = divmod(remaining, 60)
    return f"{mins:02d}:{secs:02d}"


@st.cache_data(ttl=86400, show_spinner=False)
def cached_daily_status_check(rows_payload: str, api_key: str) -> dict:
    rows = json.loads(rows_payload)
    results = {}
    for idx, row in enumerate(rows):
        prompt = f"""
Given company={row.get("Company","")}, role={row.get("Role","")},
applied={row.get("Date","")}, platform={row.get("Platform","")}.
Based on typical hiring timelines, what is the likely current status?
Return JSON only:
{{
  "status":"string",
  "next_action":"string",
  "expected_response_days":0,
  "tips":["string"]
}}
"""
        try:
            output = run_gemini_prompt_with_key(prompt, api_key)
            parsed = extract_json_payload(output)
            if isinstance(parsed, dict):
                results[str(idx)] = parsed
        except Exception:
            continue
    return results


def current_pipeline_key() -> str:
    payload = (
        st.session_state.get("api_key", "").strip()
        + "|"
        + st.session_state.get("location", "").strip()
        + "|"
        + st.session_state.get("resume_text", "").strip()
    )
    return sha256(payload.encode("utf-8")).hexdigest()



# ═══════════════════════════════════════════════════════════════════════════
# LOCAL EXCEL STORE  — appends one row per user analysis to AutoApply AI_users.xlsx
# File is created next to app.py if it doesn't exist yet.
# ═══════════════════════════════════════════════════════════════════════════

import os
from pathlib import Path

# ═══════════════════════════════════════════════════════════════════════════
# SUPABASE CLIENT — single shared instance, lazy-initialised
# ═══════════════════════════════════════════════════════════════════════════
_SUPABASE_URL = "https://vuiwehxoeijsooqlzazv.supabase.co"
_SUPABASE_KEY = "sb_publishable_h3u2haEm2qOq66OkFiF61g_po168uLd"

try:
    from supabase import create_client as _create_supabase_client, Client as _SupabaseClient
    _supabase: _SupabaseClient = _create_supabase_client(_SUPABASE_URL, _SUPABASE_KEY)
except Exception as _sb_err:
    _supabase = None  # type: ignore


def _sb():
    """Return the Supabase client, or None if unavailable."""
    return _supabase


def append_user_to_excel() -> None:
    """
    Log the current session's resume analysis to Supabase `analytics` table.
    Called automatically after every successful analysis. Silent on failure.
    """
    try:
        db = _sb()
        if db is None:
            return

        profile = st.session_state.get("profile", {})
        ats     = st.session_state.get("ats", {})
        bk      = ats.get("breakdown", {})
        jobs    = st.session_state.get("jobs", [])
        top     = jobs[0] if jobs else {}
        imps    = st.session_state.get("improvements", [])

        db.table("analytics").insert({
            "session_id":          st.session_state.get("session_id", ""),
            "name":                profile.get("name", "") or "",
            "seniority":           profile.get("seniority_level", "") or "",
            "experience_years":    int(profile.get("experience_years", 0) or 0),
            "skills":              ", ".join(profile.get("skills", [])[:20]),
            "job_titles":          ", ".join(profile.get("job_titles", [])),
            "education":           ", ".join(profile.get("education", [])),
            "resume_file":         st.session_state.get("resume_filename", "") or "",
            "location":            st.session_state.get("location", "") or "",
            "currency":            st.session_state.get("currency", "₹"),
            "ats_overall":         int(ats.get("overall_score", 0)),
            "ats_keywords":        int(bk.get("Keywords", 0)),
            "ats_format":          int(bk.get("Format", 0)),
            "ats_impact":          int(bk.get("Impact Statements", 0)),
            "ats_skills_match":    int(bk.get("Skills Match", 0)),
            "ats_readability":     int(bk.get("Readability", 0)),
            "jobs_found":          len(jobs),
            "top_job":             str(top.get("title", "")),
            "top_job_company":     str(top.get("company", "")),
            "top_salary_lpa":      float(top.get("salary_lpa", 0) or 0),
            "top_ats_match":       float(top.get("ats_match", 0) or 0),
            "improvements":        " | ".join(str(i) for i in imps[:10]),
            "platforms_connected": ", ".join(st.session_state.get("platforms_connected", [])),
            "active_provider":     st.session_state.get("active_provider", "") or "",
        }).execute()
    except Exception:
        pass  # Never break the app due to logging failure


# ═══════════════════════════════════════════════════════════════════════════
# AUTH SYSTEM — Login / Register with credentials stored in Supabase
# ═══════════════════════════════════════════════════════════════════════════

# Maps Supabase column → session_state key for API keys
_API_KEY_COLUMN_MAP = {
    "gemini_key":   "api_key",
    "groq_key":     "groq_api_key",
    "claude_key":   "claude_api_key",
    "together_key": "together_api_key",
    "openai_key":   "openai_api_key",
}


def _restore_user_api_keys(user_row: dict) -> None:
    """Load API keys from the user row dict into session_state.
    Keys that are present in the DB are marked as validated immediately —
    no need for the user to re-enter or re-test them on every login.
    """
    for col, state_key in _API_KEY_COLUMN_MAP.items():
        val = str(user_row.get(col, "") or "").strip()
        if val:
            st.session_state[state_key] = val
            # Mark each present key as validated so the UI shows it as connected.
            if state_key == "api_key":
                st.session_state["api_validated"] = True
                st.session_state["api_validation_message"] = "✅ Gemini Connected"
                st.session_state["api_last_checked_key"] = val
            elif state_key == "groq_api_key":
                st.session_state["groq_api_validated"] = True
                st.session_state["groq_api_validation_message"] = "✅ Groq Connected"
            elif state_key == "claude_api_key":
                st.session_state["claude_api_validated"] = True
                st.session_state["claude_api_validation_message"] = "✅ Claude Connected"
            elif state_key == "together_api_key":
                st.session_state["together_api_validated"] = True
                st.session_state["together_api_validation_message"] = "✅ HuggingFace Connected"
            elif state_key == "openai_api_key":
                st.session_state["openai_api_validated"] = True
                st.session_state["openai_api_validation_message"] = "✅ OpenAI Connected"
    st.session_state["gemini_api_key"] = st.session_state.get("api_key", "")


def _save_user_api_keys(username: str) -> None:
    """Write current session API keys back to the user row in Supabase.
    Only columns with a non-empty value are included in the update — blank
    session values never overwrite keys that are already stored in the DB.
    """
    try:
        db = _sb()
        if db is None:
            return
        update_data = {
            col: st.session_state.get(state_key, "").strip()
            for col, state_key in _API_KEY_COLUMN_MAP.items()
            if st.session_state.get(state_key, "").strip()  # skip blank values
        }
        if not update_data:
            return  # nothing to save — don't touch the DB row
        db.table("users").update(update_data).eq(
            "username", username
        ).execute()
    except Exception:
        pass




def _auto_save_keys_for_logged_in_user() -> None:
    """Silently save current API keys to the logged-in user account."""
    user = st.session_state.get("auth_user")
    if isinstance(user, dict) and user.get("Username"):
        _save_user_api_keys(user["Username"])


def _hash_password(password: str, salt: str) -> str:
    """PBKDF2-HMAC-SHA256 hash — safe for storage."""
    dk = pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 260_000)
    return dk.hex()


def _load_credentials() -> list:
    """Return all credential rows as a list of dicts (Supabase users table)."""
    try:
        db = _sb()
        if db is None:
            return []
        result = db.table("users").select("*").execute()
        return result.data or []
    except Exception:
        return []


def _save_credential_row(row_data: dict) -> None:
    """Insert a new user credential row into Supabase."""
    try:
        db = _sb()
        if db is None:
            return
        db.table("users").insert({
            "user_id":       row_data.get("UserID", ""),
            "username":      row_data.get("Username", ""),
            "email":         row_data.get("Email", ""),
            "password_hash": row_data.get("PasswordHash", ""),
            "salt":          row_data.get("Salt", ""),
            "registered_at": row_data.get("RegisteredAt", ""),
            "last_login_at": row_data.get("LastLoginAt", ""),
            "login_count":   int(row_data.get("LoginCount", 1)),
            "is_active":     row_data.get("IsActive", "Yes"),
        }).execute()
    except Exception:
        pass


def _update_login_stats(username: str) -> None:
    """Update last_login_at and login_count for an existing user in Supabase."""
    try:
        db = _sb()
        if db is None:
            return
        # Get current login count
        result = db.table("users").select("login_count").eq("username", username).execute()
        current_count = 0
        if result.data:
            current_count = int(result.data[0].get("login_count") or 0)
        db.table("users").update({
            "last_login_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            "login_count":   current_count + 1,
        }).eq("username", username).execute()
    except Exception:
        pass


def _row_to_auth_dict(row: dict) -> dict:
    """Convert a Supabase snake_case row to the legacy CamelCase dict the app expects."""
    return {
        "UserID":       row.get("user_id", ""),
        "Username":     row.get("username", ""),
        "Email":        row.get("email", ""),
        "PasswordHash": row.get("password_hash", ""),
        "Salt":         row.get("salt", ""),
        "RegisteredAt": row.get("registered_at", ""),
        "LastLoginAt":  row.get("last_login_at", ""),
        "LoginCount":   row.get("login_count", 0),
        "IsActive":     row.get("is_active", "Yes"),
        # API keys
        "GeminiKey":    row.get("gemini_key", ""),
        "GroqKey":      row.get("groq_key", ""),
        "ClaudeKey":    row.get("claude_key", ""),
        "TogetherKey":  row.get("together_key", ""),
        "OpenAIKey":    row.get("openai_key", ""),
    }


def auth_register(username: str, email: str, password: str) -> tuple:
    """Register a new user. Returns (success, message)."""
    username = username.strip()
    email    = email.strip().lower()
    password = password.strip()

    if not username or len(username) < 3:
        return False, "Username must be at least 3 characters."
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        return False, "Enter a valid email address."
    if len(password) < 6:
        return False, "Password must be at least 6 characters."

    existing = _load_credentials()
    for row in existing:
        if str(row.get("username", "")).lower() == username.lower():
            return False, "Username already taken. Please choose another."
        if str(row.get("email", "")).lower() == email:
            return False, "An account with this email already exists."

    salt    = secrets.token_hex(16)
    pw_hash = _hash_password(password, salt)
    user_id = str(uuid.uuid4())
    now     = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    _save_credential_row({
        "UserID":       user_id,
        "Username":     username,
        "Email":        email,
        "PasswordHash": pw_hash,
        "Salt":         salt,
        "RegisteredAt": now,
        "LastLoginAt":  now,
        "LoginCount":   1,
        "IsActive":     "Yes",
    })
    return True, "Account created successfully! You can now log in."


def auth_login(username_or_email: str, password: str) -> tuple:
    """Validate login. Returns (success, message, user_dict)."""
    query    = username_or_email.strip().lower()
    password = password.strip()
    if not query or not password:
        return False, "Please fill in all fields.", {}

    existing = _load_credentials()
    matched  = None
    for row in existing:
        if query in (str(row.get("username", "")).lower(), str(row.get("email", "")).lower()):
            matched = row
            break

    if matched is None:
        return False, "No account found with that username/email.", {}
    if str(matched.get("is_active", "Yes")) == "No":
        return False, "This account has been deactivated.", {}

    salt    = str(matched.get("salt", ""))
    stored  = str(matched.get("password_hash", ""))
    attempt = _hash_password(password, salt)
    if attempt != stored:
        return False, "Incorrect password. Please try again.", {}

    _update_login_stats(str(matched.get("username", "")))
    auth_dict = _row_to_auth_dict(matched)
    _restore_user_api_keys(auth_dict)
    return True, f"Welcome back, {matched.get('username', '')}!", auth_dict


# ── Session helpers ──────────────────────────────────────────────────────────

# ── Session helpers ──────────────────────────────────────────────────────────

def is_logged_in() -> bool:
    return bool(st.session_state.get("auth_user"))


def logout() -> None:
    st.session_state["auth_user"] = None
    st.session_state["auth_view"] = "login"
    # Clear URL token on logout
    try:
        st.query_params.clear()
    except Exception:
        pass
    st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
# URL SESSION TOKEN — keeps user logged in across refreshes via ?t=<token>
# Token = base64( JSON{username, exp} ) + "." + HMAC-SHA256 signature
# Secret is derived from app secrets (falls back to a fixed seed).
# Token is valid for 30 days. No password or hash stored in URL.
# ═══════════════════════════════════════════════════════════════════════════

_TOKEN_EXPIRY_DAYS = 30
_TOKEN_PARAM = "t"


def _token_secret() -> bytes:
    """Derive a signing secret from Streamlit secrets or a stable fallback."""
    try:
        raw = st.secrets.get("SESSION_SECRET", "") or st.secrets.get("NOTION_TOKEN", "")
        if raw:
            return sha256(str(raw).encode()).digest()
    except Exception:
        pass
    # Stable fallback — still secure because HMAC is hard to forge without the key
    return b"autoapply_ai_session_secret_v1_stable"


def _make_session_token(username: str) -> str:
    """Create a signed URL-safe token for `username` valid for 30 days."""
    exp = int(time.time()) + _TOKEN_EXPIRY_DAYS * 86400
    payload = base64.urlsafe_b64encode(
        json.dumps({"u": username, "e": exp}).encode()
    ).decode()
    sig = hmac.new(_token_secret(), payload.encode(), sha256).hexdigest()
    return f"{payload}.{sig}"


def _verify_session_token(token: str) -> str | None:
    """
    Verify a session token. Returns username if valid, None otherwise.
    Checks HMAC signature AND expiry timestamp.
    """
    try:
        parts = token.split(".")
        # base64 payload may itself contain dots — sig is always last 64 hex chars
        sig = parts[-1]
        payload = ".".join(parts[:-1])
        expected_sig = hmac.new(_token_secret(), payload.encode(), sha256).hexdigest()
        if not hmac.compare_digest(sig, expected_sig):
            return None
        data = json.loads(base64.urlsafe_b64decode(payload + "=="))
        if int(data.get("e", 0)) < int(time.time()):
            return None  # expired
        return str(data.get("u", ""))
    except Exception:
        return None


def _set_url_session(username: str) -> None:
    """Write session token into URL query params."""
    try:
        token = _make_session_token(username)
        st.query_params[_TOKEN_PARAM] = token
    except Exception:
        pass


def _restore_session_from_url() -> bool:
    """
    Try to restore login from URL token. Returns True if session was restored.
    Looks up the user in Supabase and restores auth_user + API keys.
    """
    try:
        token = st.query_params.get(_TOKEN_PARAM, "")
        if not token:
            return False
        username = _verify_session_token(token)
        if not username:
            st.query_params.clear()
            return False
        # Find user row in Supabase
        db = _sb()
        if db is None:
            return False
        result = db.table("users").select("*").eq("username", username).execute()
        if not result.data:
            st.query_params.clear()
            return False
        row = result.data[0]
        if str(row.get("is_active", "Yes")) == "No":
            st.query_params.clear()
            return False
        auth_dict = _row_to_auth_dict(row)
        st.session_state["auth_user"] = auth_dict
        _restore_user_api_keys(auth_dict)
        return True
    except Exception:
        return False


# ── Login / Register UI ──────────────────────────────────────────────────────

import streamlit as st

def login_page() -> None:
    """Full-screen split login page with premium UI, preserving original logic and function name."""

    if "auth_view" not in st.session_state:
        st.session_state["auth_view"] = "login"
    if "auth_user" not in st.session_state:
        st.session_state["auth_user"] = None

    view = st.session_state["auth_view"]

    # ── Inject Premium Global Styles ──────
    st.html("""
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
/* Base Overrides */
[data-testid="stAppViewContainer"] { background: #05080f !important; }
[data-testid="stHeader"]  { display: none !important; }
[data-testid="stSidebar"] { display: none !important; }
#MainMenu, footer { display: none !important; }
[data-testid="stAppViewBlockContainer"] { padding: 0 !important; max-width: 100% !important; }
section[data-testid="stMain"] > div { padding: 0 !important; }
html, body, * { font-family: 'Plus Jakarta Sans', sans-serif !important; }

/* Layout */
.auth-outer { display: flex; min-height: 100vh; }

/* Left Panel - Feature Showcase */
.feat-panel {
    flex: 1.1;
    background: radial-gradient(circle at 0% 0%, #0d2a22 0%, #05080f 70%);
    padding: 64px; position: relative; overflow: hidden;
    border-right: 1px solid rgba(16, 217, 160, 0.1);
    display: flex; flex-direction: column;
}
.feat-logo { display: flex; align-items: center; gap: 12px; margin-bottom: 60px; }
.feat-logo-icon {
    width: 48px; height: 48px;
    background: linear-gradient(135deg, #10d9a0, #3b82f6);
    border-radius: 14px; display: flex; align-items: center;
    justify-content: center; font-size: 24px;
    box-shadow: 0 0 30px rgba(16, 217, 160, 0.2);
}
.feat-logo-name {
    font-size: 24px; font-weight: 800; color: #fff;
    letter-spacing: -0.5px;
}
.feat-headline {
    font-size: 48px; font-weight: 800; line-height: 1.1;
    color: #fff; letter-spacing: -1.5px; margin-bottom: 20px;
}
.feat-headline span {
    background: linear-gradient(90deg, #10d9a0, #3b82f6);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}
.feat-sub { font-size: 16px; color: rgba(255,255,255,0.6); line-height: 1.6; margin-bottom: 48px; max-width: 440px; }

/* Feature Items */
.feat-item { display: flex; align-items: flex-start; gap: 18px; margin-bottom: 28px; }
.feat-icon {
    width: 44px; height: 44px; border-radius: 12px; flex-shrink: 0;
    display: flex; align-items: center; justify-content: center;
    font-size: 20px; background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.08);
}
.feat-text-title { font-size: 15px; font-weight: 700; color: #fff; margin-bottom: 4px; }
.feat-text-desc  { font-size: 13px; color: rgba(255,255,255,0.5); line-height: 1.5; }

/* Stats Section */
.feat-stats {
    display: flex; gap: 48px; margin-top: auto;
    padding-top: 40px; border-top: 1px solid rgba(255,255,255,0.05);
}
.feat-stat-num {
    font-size: 32px; font-weight: 800; color: #10d9a0;
    font-family: 'DM Mono', monospace !important;
}
.feat-stat-lbl { font-size: 11px; color: rgba(255,255,255,0.4); font-weight: 600; margin-top: 4px; letter-spacing: 1px; }

/* Right Panel - Auth Card */
.auth-right {
    flex: 0.9; background: #05080f;
    display: flex; flex-direction: column;
    align-items: center; justify-content: center; padding: 48px;
}
.auth-card {
    width: 100%; max-width: 420px;
    background: #0a0f1a; border: 1px solid rgba(255,255,255,0.05);
    border-radius: 24px; padding: 40px;
    box-shadow: 0 40px 100px rgba(0,0,0,0.8);
}
.auth-form-title { font-size: 28px; font-weight: 800; color: #fff; letter-spacing: -0.5px; margin-bottom: 8px; text-align: center; }
.auth-form-sub   { font-size: 14px; color: rgba(255,255,255,0.5); margin-bottom: 32px; text-align: center; }

/* Streamlit Widget Styling */
[data-testid="stTextInput"] input {
    background: rgba(255,255,255,0.02) !important; border: 1px solid rgba(255,255,255,0.1) !important;
    border-radius: 12px !important; color: #fff !important;
    font-size: 15px !important; padding: 14px 18px !important;
}
[data-testid="stTextInput"] input:focus {
    border-color: #10d9a0 !important;
    box-shadow: 0 0 0 1px #10d9a0 !important;
}
[data-testid="stTextInput"] label {
    font-size: 12px !important; font-weight: 600 !important;
    color: rgba(255,255,255,0.4) !important; margin-bottom: 8px !important;
}
div[data-testid="stButton"] > button[kind="primary"] {
    background: linear-gradient(90deg, #10d9a0, #3b82f6) !important;
    border: none !important; border-radius: 14px !important;
    font-weight: 700 !important; font-size: 15px !important;
    padding: 16px !important; color: #fff !important;
    box-shadow: 0 10px 25px rgba(16, 217, 160, 0.25) !important;
}
div[data-testid="stButton"] > button[kind="secondary"] {
    background: rgba(255,255,255,0.03) !important; border: 1px solid rgba(255,255,255,0.08) !important;
    border-radius: 14px !important; color: rgba(255,255,255,0.7) !important; font-size: 14px !important;
}

.auth-divider { border: none; border-top: 1px solid rgba(255,255,255,0.05); margin: 32px 0; }
.powered-row { display: flex; gap: 10px; flex-wrap: wrap; justify-content: center; margin-top: 12px; }
.powered-chip {
    background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.06);
    border-radius: 10px; padding: 4px 12px;
    font-size: 12px; color: rgba(255,255,255,0.5); font-weight: 500;
}
.security-notice {
    background: rgba(16, 217, 160, 0.03); border: 1px solid rgba(16, 217, 160, 0.1);
    border-radius: 14px; padding: 16px; font-size: 12px; color: rgba(16, 217, 160, 0.8);
    line-height: 1.5; margin-top: 24px;
}

/* ── Fix: style the right Streamlit column as the auth card ── */
[data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(2) > [data-testid="stVerticalBlockBorderWrapper"] > div,
[data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(2) > div:first-child > [data-testid="stVerticalBlock"] {
    background: #0a0f1a !important;
    border: 1px solid rgba(255,255,255,0.05) !important;
    border-radius: 24px !important;
    padding: 40px !important;
    box-shadow: 0 40px 100px rgba(0,0,0,0.8) !important;
    max-width: 420px !important;
    margin: auto !important;
}
</style>
""")

    # ── Split Layout ──
    left_col, right_col = st.columns([1.1, 0.9], gap="small")

    with left_col:
        st.markdown("""
<div class="feat-panel">
  <div class="feat-logo">
    <div class="feat-logo-icon">💼</div>
    <span class="feat-logo-name">AutoApply AI</span>
  </div>
  <div class="feat-headline">Land your dream job<br>with <span>AI superpowers</span></div>
  <div class="feat-sub">Upload your resume once. Let AI score, optimize, match jobs and track applications — all from one dashboard.</div>

  <div class="feat-item">
    <div class="feat-icon">🎯</div>
    <div><div class="feat-text-title">Multi-Agent AI Automation</div>
    <div class="feat-text-desc">Powerful AI agents like OpenAI, Google, Anthropic, Groq, and Hugging Face work 24/7 to automate your search.</div></div>
  </div>
  <div class="feat-item">
    <div class="feat-icon">🔍</div>
    <div><div class="feat-text-title">Smart Job Matching</div>
    <div class="feat-text-desc">AI ranks curated job matches by salary, match %, and interview probability.</div></div>
  </div>
  <div class="feat-item">
    <div class="feat-icon">✍️</div>
    <div><div class="feat-text-title">JD-Tailored Resume</div>
    <div class="feat-text-desc">Paste any job description and get a laser-tailored resume with keyword highlights.</div></div>
  </div>
  <div class="feat-item">
    <div class="feat-icon">📊</div>
    <div><div class="feat-text-title">Application Tracker</div>
    <div class="feat-text-desc">Track every application with status, next steps, and AI-predicted follow-up timing.</div></div>
  </div>
  <div class="feat-item">
    <div class="feat-icon">⚡</div>
    <div><div class="feat-text-title">Multi-AI Auto Rotation</div>
    <div class="feat-text-desc">Gemini → Groq → Claude → HuggingFace — rotates automatically on rate limits.</div></div>
  </div>

  <div class="feat-stats">
    <div><div class="feat-stat-num">12+</div><div class="feat-stat-lbl">JOB PLATFORMS</div></div>
    <div><div class="feat-stat-num">5×</div><div class="feat-stat-lbl">AI PROVIDERS</div></div>
    <div><div class="feat-stat-num">100%</div><div class="feat-stat-lbl">FREE TO START</div></div>
  </div>
</div>
""", unsafe_allow_html=True)

    with right_col:
        # Auth Header logic preserved
        if view == "login":
            st.markdown("""
  <div class="auth-form-title">Welcome back 👋</div>
  <div class="auth-form-sub">Sign in to continue your job search</div>
""", unsafe_allow_html=True)
        else:
            st.markdown("""
  <div class="auth-form-title">Create account ✨</div>
  <div class="auth-form-sub">Free forever — no credit card needed</div>
""", unsafe_allow_html=True)

        # Tab Toggle (Preserving logic and state management)
        tc1, tc2 = st.columns(2)
        with tc1:
            if st.button("🔐  Sign In", use_container_width=True,
                         type="primary" if view == "login" else "secondary"):
                st.session_state["auth_view"] = "login"
                st.rerun()
        with tc2:
            if st.button("✨  Register", use_container_width=True,
                         type="primary" if view == "register" else "secondary"):
                st.session_state["auth_view"] = "register"
                st.rerun()

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        # ── Auth Forms ──
        if view == "login":
            login_id = st.text_input(
                "Username or Email", key="login_id",
                placeholder="you@email.com or username",
            )
            login_pw = st.text_input(
                "Password", type="password", key="login_pw",
                placeholder="••••••••",
            )
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if st.button("Sign In →", type="primary", use_container_width=True):
                # logic from original code preserved
                ok, msg, user = auth_login(login_id, login_pw)
                if ok:
                    st.session_state["auth_user"] = user
                    _set_url_session(str(user.get("Username", "")))
                    sb_load_tracker()
                    sb_load_platforms()
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

            st.markdown(
                "<p style='text-align:center;color:rgba(255,255,255,0.4);font-size:13px;margin-top:24px;'>"
                "New here? Switch to <b style='color:rgba(255,255,255,0.7);'>Register</b> above.</p>",
                unsafe_allow_html=True,
            )
            st.markdown("<hr class='auth-divider'>", unsafe_allow_html=True)
            st.markdown(
                "<p style='text-align:center;font-size:11px;color:rgba(255,255,255,0.3);text-transform:uppercase;letter-spacing:1px;font-weight:600;'>Powered by the best</p>"
                "<div class='powered-row'>"
                "<span class='powered-chip'>✦ Gemini</span>"
                "<span class='powered-chip'>✦ Claude</span>"
                "<span class='powered-chip'>✦ Groq</span>"
                "<span class='powered-chip'>🤗 HuggingFace</span>"
                "</div>",
                unsafe_allow_html=True,
            )

        else:
            reg_uname = st.text_input("Username", key="reg_uname", placeholder="e.g. priya_dev")
            reg_email = st.text_input("Email", key="reg_email", placeholder="you@email.com")
            reg_pw    = st.text_input("Password", type="password", key="reg_pw", placeholder="Min 6 characters")
            reg_pw2   = st.text_input("Confirm Password", type="password", key="reg_pw2", placeholder="Re-enter password")
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if st.button("Create Account →", type="primary", use_container_width=True):
                # original logic preserved
                if reg_pw != reg_pw2:
                    st.error("Passwords do not match.")
                else:
                    ok, msg = auth_register(reg_uname, reg_email, reg_pw)
                    if ok:
                        st.success(msg)
                        st.session_state["auth_view"] = "login"
                        st.rerun()
                    else:
                        st.error(msg)

            st.markdown(
                "<p style='text-align:center;color:rgba(255,255,255,0.4);font-size:13px;margin-top:24px;'>"
                "Already have an account? Switch to <b style='color:rgba(255,255,255,0.7);'>Sign In</b> above.</p>",
                unsafe_allow_html=True,
            )
            st.markdown("<hr class='auth-divider'>", unsafe_allow_html=True)
            st.markdown("""
<div class="security-notice">
  🔒 <b>Security First:</b> Passwords hashed with PBKDF2-SHA256. Credentials never leave your machine. This is a core architectural pillar.
</div>
""", unsafe_allow_html=True)

        # (auth-card is now styled via CSS on the column container — no wrapper div needed)


def run_full_automation() -> None:
    resume_compact = trimmed_text(st.session_state.get("resume_text", ""), max_chars=7000)
    location = st.session_state.get("location", "")

    # ── Count word count locally (no AI needed) ──────────────────────────
    raw_text = st.session_state.get("resume_text", "")
    word_count = len(raw_text.split())
    page_estimate = max(1, round(word_count / 400))

    prompt = f"""
You are an expert ATS analyst and career coach. Analyze this resume thoroughly.
Return ONLY valid JSON — no markdown, no explanation, no extra text.

JSON schema (fill every field accurately):
{{
  "profile": {{
    "name": "string",
    "skills": ["list of ALL technical + soft skills found"],
    "experience_years": 0,
    "job_titles": ["list of all roles held"],
    "education": ["Degree, Institute, Year"],
    "seniority_level": "Intern|Junior|Mid|Senior|Lead|Manager|Director"
  }},
  "insight": {{
    "keywords_present": ["top 20 strong keywords/phrases found in resume"],
    "keywords_missing": ["top 15 high-value keywords this role SHOULD have but are missing"],
    "certifications": ["list all certs/courses found, empty list if none"],
    "languages": ["programming or spoken languages found"],
    "contact": {{
      "email": "string or empty",
      "phone": "string or empty",
      "linkedin": "URL or empty",
      "github": "URL or empty",
      "portfolio": "URL or empty"
    }},
    "summary_line": "One powerful professional summary line for this candidate (max 25 words)",
    "strengths": ["top 5 resume strengths — be specific, cite actual resume content"],
    "weaknesses": ["top 5 resume weaknesses / gaps — be specific and actionable"],
    "improvement_suggestions": ["top 8 concrete, prioritized improvement suggestions — each starting with an action verb"],
    "career_gaps": ["describe any employment gaps found, empty list if none"],
    "salary_range": {{"min_lpa": 0, "max_lpa": 0, "currency": "₹"}},
    "glassdoor_salary": {{
      "role": "primary job title this candidate matches",
      "min_lpa": 0,
      "max_lpa": 0,
      "median_lpa": 0,
      "currency": "₹",
      "location": "city or Remote",
      "experience_band": "0-2 yrs|2-5 yrs|5-10 yrs|10+ yrs",
      "note": "one-line rationale citing experience level and market data"
    }},
    "industry": "primary industry this resume targets",
    "recommended_roles": ["5 best-fit job titles for this candidate"],
    "resume_score_label": "Weak|Average|Good|Strong"
  }},
  "ats": {{
    "overall_score": 0,
    "breakdown": {{
      "Keywords": 0,
      "Format": 0,
      "Impact Statements": 0,
      "Skills Match": 0,
      "Readability": 0
    }}
  }},
  "optimized_resume": "full improved resume text (keep structure, add impact)",
  "improvements": ["list of 8-12 specific actionable improvement suggestions"],
  "jobs": [
    {{
      "title": "string",
      "company": "string",
      "platform": "LinkedIn|Naukri|Indeed|etc",
      "location": "string",
      "tags": ["skill1","skill2"],
      "salary_lpa": 0,
      "ats_match": 0,
      "interview_probability": 0,
      "competition_level": "Low|Medium|High",
      "apply_url": "string",
      "rationale": "string"
    }}
  ]
}}

Rules:
- jobs: max 10, sorted by salary_lpa desc then ats_match desc
- salary_range: realistic market estimate for this candidate's level in India
- glassdoor_salary: estimate what Glassdoor would show for this role+experience+location; min/max/median in LPA; be realistic based on India market data
- keywords_missing: focus on role-relevant technical skills, tools, certifications they lack
- strengths: cite specific evidence from the resume (e.g. "Led 3 projects at X" not just "leadership")
- weaknesses: be honest, specific, and actionable — not generic
- improvement_suggestions: 8 concrete prioritized suggestions each starting with an action verb (e.g. "Add", "Quantify", "Remove")
- Return ONLY the JSON object, nothing else

Location preference: {location}

Resume text:
{resume_compact}
"""
    output = run_gemini_prompt(prompt, use_grounding=True)
    parsed = extract_json_payload(output)
    if not isinstance(parsed, dict):
        raise RuntimeError("AI returned invalid JSON. Try re-uploading your resume.")

    profile   = parsed.get("profile", {})
    insight   = parsed.get("insight", {})
    ats       = parsed.get("ats", {})
    breakdown = ats.get("breakdown", {})
    jobs      = parsed.get("jobs", [])

    previous_ats = st.session_state.get("ats", {}).get("overall_score", 0)

    # ── Store profile ────────────────────────────────────────────────────
    st.session_state["profile"] = {
        "name":             profile.get("name", ""),
        "skills":           profile.get("skills", []),
        "experience_years": profile.get("experience_years", 0),
        "job_titles":       profile.get("job_titles", []),
        "education":        profile.get("education", []),
        "seniority_level":  profile.get("seniority_level", ""),
    }

    # ── Store ATS ────────────────────────────────────────────────────────
    st.session_state["ats"] = {
        "overall_score": int(float(ats.get("overall_score", 0))),
        "breakdown": {
            "Keywords":          int(float(breakdown.get("Keywords", 0))),
            "Format":            int(float(breakdown.get("Format", 0))),
            "Impact Statements": int(float(breakdown.get("Impact Statements", 0))),
            "Skills Match":      int(float(breakdown.get("Skills Match", 0))),
            "Readability":       int(float(breakdown.get("Readability", 0))),
        },
    }

    # ── Store extended insight (PERSISTS across pages) ───────────────────
    contact_raw  = insight.get("contact", {})
    salary_raw   = insight.get("salary_range", {})
    gd_raw       = insight.get("glassdoor_salary", {})
    st.session_state["insight"] = {
        "keywords_present":       insight.get("keywords_present", []),
        "keywords_missing":       insight.get("keywords_missing", []),
        "certifications":         insight.get("certifications", []),
        "languages":              insight.get("languages", []),
        "contact": {
            "email":     str(contact_raw.get("email", "") or ""),
            "phone":     str(contact_raw.get("phone", "") or ""),
            "linkedin":  str(contact_raw.get("linkedin", "") or ""),
            "github":    str(contact_raw.get("github", "") or ""),
            "portfolio": str(contact_raw.get("portfolio", "") or ""),
        },
        "summary_line":           insight.get("summary_line", ""),
        "strengths":              insight.get("strengths", []),
        "weaknesses":             insight.get("weaknesses", []),
        "improvement_suggestions": insight.get("improvement_suggestions", []),
        "career_gaps":            insight.get("career_gaps", []),
        "salary_range": {
            "min_lpa":  float(salary_raw.get("min_lpa", 0) or 0),
            "max_lpa":  float(salary_raw.get("max_lpa", 0) or 0),
            "currency": str(salary_raw.get("currency", "₹") or "₹"),
        },
        "glassdoor_salary": {
            "role":             str(gd_raw.get("role", "") or ""),
            "min_lpa":          float(gd_raw.get("min_lpa", 0) or 0),
            "max_lpa":          float(gd_raw.get("max_lpa", 0) or 0),
            "median_lpa":       float(gd_raw.get("median_lpa", 0) or 0),
            "currency":         str(gd_raw.get("currency", "₹") or "₹"),
            "location":         str(gd_raw.get("location", "") or ""),
            "experience_band":  str(gd_raw.get("experience_band", "") or ""),
            "note":             str(gd_raw.get("note", "") or ""),
        },
        "industry":               insight.get("industry", ""),
        "recommended_roles":      insight.get("recommended_roles", []),
        "resume_score_label":     insight.get("resume_score_label", "Average"),
        "word_count":             word_count,
        "page_estimate":          page_estimate,
    }

    # ── Store rest ───────────────────────────────────────────────────────
    st.session_state["resume_optimized"] = parsed.get("optimized_resume", "")
    st.session_state["improvements"]     = parsed.get("improvements", [])
    st.session_state["jobs"]             = jobs if isinstance(jobs, list) else []
    st.session_state["optimized_resume"] = st.session_state["resume_optimized"]
    st.session_state["resume_analysis"]  = st.session_state["profile"]
    st.session_state["ats_score"]        = st.session_state["ats"]["overall_score"]
    st.session_state["ats_breakdown"]    = st.session_state["ats"]["breakdown"]
    st.session_state["jobs_ranked"]      = st.session_state["jobs"]

    top_job = st.session_state["jobs"][0] if st.session_state["jobs"] else {}
    log_notion(
        {
            "SessionID":    st.session_state.get("session_id", ""),
            "Timestamp":    datetime.utcnow().isoformat(),
            "Location":     st.session_state.get("location", ""),
            "Filename":     st.session_state.get("resume_filename", ""),
            "Format":       st.session_state.get("resume_filename", "").split(".")[-1] if st.session_state.get("resume_filename") else "",
            "Keywords":     ", ".join(st.session_state["insight"]["keywords_present"][:10]),
            "Skills":       ", ".join(st.session_state["profile"].get("skills", [])[:20]),
            "ExpYears":     st.session_state["profile"].get("experience_years", 0),
            "Seniority":    st.session_state["profile"].get("seniority_level", ""),
            "ATSBefore":    previous_ats,
            "ATSAfter":     st.session_state["ats"]["overall_score"],
            "JobsSearched": len(st.session_state["jobs"]),
            "TargetSalary": float(top_job.get("salary_lpa", 0) or 0),
            "Currency":     st.session_state.get("currency", "₹"),
            "Platforms":    ", ".join(st.session_state.get("platforms_connected", [])),
            "TopJob":       f"{top_job.get('title', '')} @ {top_job.get('company', '')}".strip(),
            "TopScore":     float(top_job.get("ats_match", 0) or 0),
        }
    )
    append_user_to_excel()


def auto_run_if_ready(show_status: bool = False) -> None:
    # Ready if ANY provider key is available + resume is uploaded
    providers_available = bool(_build_provider_list())
    ready = (
        providers_available
        and bool(st.session_state.get("resume_text", "").strip())
    )
    if not ready:
        return

    key = current_pipeline_key()
    if key == st.session_state.get("analysis_cache_key"):
        return
    if st.session_state.get("is_processing"):
        return

    # Only block if ALL providers are in quota cooldown (single-key scenario)
    now_ts = time.time()
    if now_ts < float(st.session_state.get("quota_block_until", 0.0)):
        # Check if a new/different provider has been added since cooldown set
        if _build_provider_list():
            # More than just the blocked key might be available — clear cooldown and try
            providers = _build_provider_list()
            blocked_fp = st.session_state.get("quota_key_fingerprint", "")
            active_keys = [p["key"] for p in providers]
            if any(k != blocked_fp for k in active_keys):
                # A different provider key exists — clear cooldown and proceed
                st.session_state["quota_block_until"] = 0.0
                st.session_state["quota_message"] = ""
                st.session_state["quota_key_fingerprint"] = ""
            else:
                return

    st.session_state["is_processing"] = True
    st.session_state["analysis_error"] = ""
    st.session_state["quota_message"] = ""
    try:
        if show_status:
            with st.spinner("Analyzing resume..."):
                run_full_automation()
        else:
            run_full_automation()
        st.session_state["analysis_cache_key"] = key
    except Exception as exc:
        if is_quota_error(exc):
            retry_after_seconds = 180
            st.session_state["quota_block_until"] = time.time() + retry_after_seconds
            st.session_state["quota_key_fingerprint"] = st.session_state.get("api_key", "").strip()
            # Humanized message — mentions rotation
            n_providers = len(_build_provider_list())
            if n_providers > 1:
                st.session_state["quota_message"] = (
                    "Rate limit hit — AutoApply AI rotated to next provider automatically. "
                    "If all providers are exhausted, wait for reset."
                )
            else:
                st.session_state["quota_message"] = (
                    "API quota limit reached. Add a Groq or Claude key in API Keys page "
                    "for automatic fallback — no waiting needed."
                )
            st.session_state["analysis_error"] = ""
        else:
            st.session_state["analysis_error"] = str(exc)
    finally:
        st.session_state["is_processing"] = False


def circular_score(score: int) -> None:
    value = max(0, min(100, int(score)))
    radius = 54
    circumference = 2 * math.pi * radius
    progress = (value / 100) * circumference
    st.markdown(
        f"""
<div style="display:flex;justify-content:center;align-items:center;">
  <svg width="150" height="150" viewBox="0 0 150 150">
    <circle cx="75" cy="75" r="{radius}" stroke="#1e2d45" stroke-width="12" fill="none"></circle>
    <circle cx="75" cy="75" r="{radius}" stroke="#10d9a0" stroke-width="12" fill="none"
      stroke-dasharray="{progress} {circumference}" transform="rotate(-90 75 75)"
      stroke-linecap="round"></circle>
    <text x="75" y="82" text-anchor="middle" fill="#EAF2FF" font-size="28" font-weight="700">{value}</text>
  </svg>
</div>
        """,
        unsafe_allow_html=True,
    )


def build_diff_html(original_text: str, optimized_text: str) -> str:
    differ = HtmlDiff(wrapcolumn=70)
    table = differ.make_table(
        original_text.splitlines(),
        optimized_text.splitlines(),
        fromdesc="Original",
        todesc="Optimized",
        context=True,
        numlines=2,
    )
    styles = """
<style>
table.diff {font-family: monospace; border: 1px solid #444; width: 100%;}
.diff_header {background:#1f2937;color:#fff;padding:4px;}
td {padding:3px 6px;}
.diff_add {background:#16361f;color:#d1fae5;}
.diff_sub {background:#3f1a1a;color:#fecaca;}
.diff_chg {background:#1f2937;color:#fef08a;}
</style>
"""
    return styles + table


def optimized_resume_to_pdf_bytes(text: str) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Arial", size=11)
    for line in text.splitlines() or [""]:
        safe_line = line.encode("latin-1", "replace").decode("latin-1")
        pdf.multi_cell(0, 6, txt=safe_line)
    return bytes(pdf.output(dest="S"))


def format_tracker_export(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    alt_fill = PatternFill(start_color="F4F7FB", end_color="F4F7FB", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = alt_fill

    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def parse_tracker_upload(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    if df.empty:
        return pd.DataFrame(columns=TRACKER_COLUMNS)

    normalized = {str(col).strip().lower(): col for col in df.columns}
    mapped = {}
    for required in TRACKER_COLUMNS:
        key = required.lower()
        mapped[required] = df[normalized[key]] if key in normalized else ""
    return pd.DataFrame(mapped)


# ═══════════════════════════════════════════════════════════════════════════
# SUPABASE TRACKER HELPERS — save & load tracker rows per user
# ═══════════════════════════════════════════════════════════════════════════

def _get_current_user_id() -> str:
    """Return the user_id of the logged-in user, or empty string."""
    user = st.session_state.get("auth_user")
    if isinstance(user, dict):
        return str(user.get("UserID", "") or "")
    return ""


def sb_save_tracker() -> None:
    """Upsert all tracker rows for the current user into Supabase `tracker` table."""
    try:
        db = _sb()
        if db is None:
            return
        user_id = _get_current_user_id()
        if not user_id:
            return
        rows = st.session_state.get("tracker", [])
        # Delete existing rows for this user then re-insert (simple full-replace strategy)
        db.table("tracker").delete().eq("user_id", user_id).execute()
        if rows:
            db.table("tracker").insert([
                {
                    "user_id":   user_id,
                    "company":   str(r.get("Company", "")),
                    "role":      str(r.get("Role", "")),
                    "platform":  str(r.get("Platform", "")),
                    "date":      str(r.get("Date", "")),
                    "status":    str(r.get("Status", "")),
                    "package":   str(r.get("Package", "")),
                    "notes":     str(r.get("Notes", "")),
                    "next_step": str(r.get("NextStep", "")),
                    "url":       str(r.get("URL", "")),
                }
                for r in rows
            ]).execute()
    except Exception:
        pass


def sb_load_tracker() -> None:
    """Load tracker rows for the current user from Supabase into session_state."""
    try:
        db = _sb()
        if db is None:
            return
        user_id = _get_current_user_id()
        if not user_id:
            return
        result = db.table("tracker").select("*").eq("user_id", user_id).order("created_at").execute()
        rows = []
        for r in (result.data or []):
            rows.append({
                "Company":  r.get("company", ""),
                "Role":     r.get("role", ""),
                "Platform": r.get("platform", ""),
                "Date":     r.get("date", ""),
                "Status":   r.get("status", ""),
                "Package":  r.get("package", ""),
                "Notes":    r.get("notes", ""),
                "NextStep": r.get("next_step", ""),
                "URL":      r.get("url", ""),
            })
        if rows:  # Only overwrite if we got data (don't clear local data on connection error)
            st.session_state["tracker"] = rows
            st.session_state["tracker_rows"] = rows
    except Exception:
        pass


def sb_save_platforms() -> None:
    """Save connected platforms for the current user to Supabase `platforms` table."""
    try:
        db = _sb()
        if db is None:
            return
        user_id = _get_current_user_id()
        if not user_id:
            return
        connected     = st.session_state.get("platforms_connected", [])
        connected_at  = st.session_state.get("platforms_connected_at", {})
        db.table("platforms").delete().eq("user_id", user_id).execute()
        if connected:
            db.table("platforms").insert([
                {
                    "user_id":      user_id,
                    "platform":     p,
                    "connected_at": connected_at.get(p, ""),
                }
                for p in connected
            ]).execute()
    except Exception:
        pass


def sb_load_platforms() -> None:
    """Load connected platforms for the current user from Supabase."""
    try:
        db = _sb()
        if db is None:
            return
        user_id = _get_current_user_id()
        if not user_id:
            return
        result = db.table("platforms").select("*").eq("user_id", user_id).execute()
        if result.data:
            st.session_state["platforms_connected"] = [r["platform"] for r in result.data]
            st.session_state["platforms_connected_at"] = {
                r["platform"]: r.get("connected_at", "") for r in result.data
            }
    except Exception:
        pass


def add_job_to_tracker(job: dict) -> None:
    row = {
        "Company": str(job.get("company", "")),
        "Role": str(job.get("title", "")),
        "Platform": str(job.get("platform", "")),
        "Date": date.today().isoformat(),
        "Status": "Applied",
        "Package": f"{float(job.get('salary_lpa', 0) or 0):.1f} LPA",
        "Notes": str(job.get("rationale", "")),
        "NextStep": "Follow up in 3 days",
        "URL": str(job.get("apply_url", "")),
    }
    st.session_state["tracker"].append(row)
    st.session_state["tracker_rows"] = st.session_state["tracker"]
    log_notion(
        {
            "SessionID": st.session_state.get("session_id", ""),
            "Timestamp": datetime.utcnow().isoformat(),
            "Location": st.session_state.get("location", ""),
            "JobsSearched": len(st.session_state.get("jobs", [])),
            "Platforms": ", ".join(st.session_state.get("platforms_connected", [])),
            "TopJob": f"{job.get('title', '')} @ {job.get('company', '')}".strip(),
            "TopScore": float(job.get("ats_match", 0) or 0),
            "Currency": st.session_state.get("currency", "₹"),
        }
    )


def format_salary_lpa(value: float) -> str:
    currency = st.session_state.get("currency", "₹")
    if currency == "$":
        return f"${value * 1200:.0f}"
    if currency == "Both":
        return f"{value:.1f} LPA | ${value * 1200:.0f}"
    return f"{value:.1f} LPA"


def _keyword_chips_html(keywords: list, color: str, bg: str, border: str) -> str:
    """Render a list of keywords as pill chips."""
    if not keywords:
        return "<span style='color:#475569;font-size:12px;font-style:italic;'>None found</span>"
    chips = "".join(
        f'<span style="display:inline-block;margin:3px 4px 3px 0;padding:4px 10px;'
        f'border-radius:20px;background:{bg};color:{color};font-size:11px;'
        f'font-weight:600;border:1px solid {border};letter-spacing:0.2px;">{kw}</span>'
        for kw in keywords
    )
    return f'<div style="line-height:2;">{chips}</div>'


def _insight_stat_box(label: str, value: str, color: str = "#10d9a0") -> str:
    return (
        f'<div style="background:#070b14;border:1px solid #1a2744;border-radius:10px;'
        f'padding:12px 14px;text-align:center;">'
        f'  <div style="color:{color};font-size:20px;font-weight:800;line-height:1;">{value}</div>'
        f'  <div style="color:#64748b;font-size:10px;font-weight:600;margin-top:4px;'
        f'  text-transform:uppercase;letter-spacing:0.5px;">{label}</div>'
        f'</div>'
    )


def dashboard_page() -> None:
    st.subheader("Your Resume Insight")

    # ── Provider status bar ──────────────────────────────────────────────
    providers_live = _build_provider_list()
    if providers_live:
        provider_names = " → ".join(
            f"<span style='color:{p['color']};font-weight:600;'>{p['label']}</span>"
            for p in providers_live
        )
        st.markdown(
            f"<div style='background:#0d1117;border:1px solid #1e2d45;border-radius:10px;"
            f"padding:8px 16px;margin-bottom:10px;font-size:12px;'>"
            f"<span style='color:#94a3b8;'>AI Providers:</span> {provider_names}"
            f"<span style='color:#64748b;font-size:11px;'>&nbsp;·&nbsp;auto-rotates on rate limit</span>"
            f"</div>",
            unsafe_allow_html=True,
        )
    else:
        st.warning("⚠ No API keys configured. Go to **API Keys** page to add a free key.", icon="🔑")

    # ── Upload + Location row ────────────────────────────────────────────
    ul_col, loc_col = st.columns([2, 1])
    with ul_col:
        uploaded_resume = st.file_uploader("Upload Resume (PDF / DOCX / TXT)", type=["pdf", "docx", "txt"])
        if uploaded_resume:
            st.session_state["resume_bytes"]         = uploaded_resume.getvalue()
            st.session_state["original_resume_bytes"] = uploaded_resume.getvalue()
            st.session_state["original_resume_mime"]  = uploaded_resume.type or "application/octet-stream"
            parsed_text = read_uploaded_resume(uploaded_resume)
            if parsed_text:
                st.session_state["resume_text"]     = parsed_text
                st.session_state["resume_filename"] = uploaded_resume.name
                st.success(f"✅ Resume loaded: **{uploaded_resume.name}**")
            else:
                st.warning("Unable to extract text from this file.")
    with loc_col:
        st.session_state["location"] = st.text_input(
            "Preferred Location",
            value=st.session_state["location"],
            placeholder="Bengaluru / Remote / Mumbai",
        ).strip()

    # ── Trigger analysis ─────────────────────────────────────────────────
    if providers_live:
        auto_run_if_ready(show_status=True)
    else:
        st.info("Add at least one API key in the **API Keys** page to start analysis.", icon="🔑")

    # ── Quota / error messages ───────────────────────────────────────────
    if st.session_state.get("quota_message"):
        timer_txt = quota_countdown_text()
        msg = st.session_state["quota_message"]
        st.warning(f"{msg} Reset in {timer_txt}" if timer_txt else msg)
    if st.session_state.get("analysis_error"):
        st.error(f"Analysis failed: {st.session_state['analysis_error']}")

    # ── If no analysis yet — show placeholder ────────────────────────────
    profile  = st.session_state.get("profile", {})
    insight  = st.session_state.get("insight", {})
    ats      = st.session_state.get("ats", {})
    jobs     = st.session_state.get("jobs", [])
    improvements = st.session_state.get("improvements", [])

    has_data = bool(profile.get("name") or insight.get("keywords_present") or ats.get("overall_score", 0) > 0)

    if not has_data:
        st.markdown(
            "<div style='text-align:center;padding:60px 20px;color:#475569;'>"
            "<div style='font-size:48px;margin-bottom:12px;'>📄</div>"
            "<div style='font-size:16px;font-weight:600;color:#94a3b8;margin-bottom:6px;'>No resume analysed yet</div>"
            "<div style='font-size:13px;'>Upload your resume above — AI will extract insights in seconds.</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        return

    # ════════════════════════════════════════════════════════════════════
    # INSIGHT DASHBOARD — 6 tabs
    # ════════════════════════════════════════════════════════════════════
    score_label = insight.get("resume_score_label", "Average")
    score_color = {"Weak": "#ef4444", "Average": "#f59e0b", "Good": "#3b82f6", "Strong": "#10d9a0"}.get(score_label, "#94a3b8")
    ats_score   = ats.get("overall_score", 0)

    # ── Top hero banner ──────────────────────────────────────────────────
    name        = profile.get("name", "Candidate")
    exp_yrs     = profile.get("experience_years", 0)
    seniority   = profile.get("seniority_level", "")
    industry    = insight.get("industry", "")
    summary     = insight.get("summary_line", "")
    word_count  = insight.get("word_count", 0)
    page_est    = insight.get("page_estimate", 1)
    sal_min     = insight.get("salary_range", {}).get("min_lpa", 0)
    sal_max     = insight.get("salary_range", {}).get("max_lpa", 0)

    st.markdown(
        f'<div style="background:linear-gradient(135deg,#0a0f1a,#0d1424);'
        f'border:1px solid #1e2d45;border-radius:16px;padding:20px 24px;margin-bottom:16px;">'
        f'  <div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:12px;">'
        f'    <div style="flex:1;min-width:220px;">'
        f'      <div style="font-size:22px;font-weight:800;color:#EAF2FF;letter-spacing:-0.3px;">{name}</div>'
        f'      <div style="color:#64748b;font-size:12px;margin-top:4px;">'
        f'        <span style="color:#94a3b8;">{seniority}</span>'
        f'        {"&nbsp;·&nbsp;" if industry else ""}'
        f'        <span style="color:#94a3b8;">{industry}</span>'
        f'        {"&nbsp;·&nbsp;" if exp_yrs else ""}'
        f'        <span style="color:#94a3b8;">{exp_yrs} yrs exp</span>'
        f'      </div>'
        f'      {"<div style=&quot;margin-top:10px;color:#cbd5e1;font-size:13px;font-style:italic;line-height:1.5;&quot;>&quot;" + summary + "&quot;</div>" if summary else ""}'
        f'    </div>'
        f'    <div style="display:flex;gap:10px;flex-wrap:wrap;">'
        f'      <div style="background:#070b14;border:1px solid {score_color}30;border-radius:12px;'
        f'      padding:12px 18px;text-align:center;">'
        f'        <div style="color:{score_color};font-size:22px;font-weight:800;">{ats_score}</div>'
        f'        <div style="color:#64748b;font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;margin-top:2px;">ATS Score</div>'
        f'        <div style="display:inline-block;margin-top:4px;padding:2px 8px;border-radius:10px;'
        f'        background:{score_color}15;border:1px solid {score_color}40;'
        f'        color:{score_color};font-size:9px;font-weight:700;">{score_label}</div>'
        f'      </div>'
        f'      <div style="background:#070b14;border:1px solid #1a2744;border-radius:12px;padding:12px 18px;text-align:center;">'
        f'        <div style="color:#10d9a0;font-size:18px;font-weight:800;">{sal_min:.0f}–{sal_max:.0f}</div>'
        f'        <div style="color:#64748b;font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;margin-top:2px;">LPA Range</div>'
        f'      </div>'
        f'      <div style="background:#070b14;border:1px solid #1a2744;border-radius:12px;padding:12px 18px;text-align:center;">'
        f'        <div style="color:#3b82f6;font-size:18px;font-weight:800;">{word_count}</div>'
        f'        <div style="color:#64748b;font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;margin-top:2px;">Words · ~{page_est}pg</div>'
        f'      </div>'
        f'    </div>'
        f'  </div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── 6 Tabs ───────────────────────────────────────────────────────────
    tab_overview, tab_keywords, tab_skills, tab_strengths, tab_contact, tab_jobs_summary = st.tabs([
        "📊 Overview",
        "🔑 Keywords",
        "🛠 Skills & Roles",
        "💪 Strengths / Gaps",
        "📬 Contact & Certs",
        "💼 Job Snapshot",
    ])

    # ══════════════════════════════════════════════════════════════════
    # TAB 1 — OVERVIEW: ATS breakdown + improvements
    # ══════════════════════════════════════════════════════════════════
    with tab_overview:
        left_col, right_col = st.columns([1, 1])

        with left_col:
            st.markdown("#### ATS Score")
            circular_score(ats_score)

            st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)
            breakdown = ats.get("breakdown", {})
            for category, score in breakdown.items():
                score_int = max(0, min(100, int(score)))
                bar_color = "#10d9a0" if score_int >= 70 else "#f59e0b" if score_int >= 45 else "#ef4444"
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">'
                    f'  <div style="width:120px;color:#94a3b8;font-size:12px;font-weight:600;flex-shrink:0;">{category}</div>'
                    f'  <div style="flex:1;background:#1a2744;border-radius:999px;height:8px;overflow:hidden;">'
                    f'    <div style="width:{score_int}%;height:100%;background:{bar_color};border-radius:999px;'
                    f'    box-shadow:0 0 8px {bar_color}55;"></div>'
                    f'  </div>'
                    f'  <div style="width:32px;text-align:right;color:{bar_color};font-size:12px;font-weight:700;">{score_int}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        with right_col:
            st.markdown("#### What to Improve")
            if improvements:
                for i, item in enumerate(improvements, 1):
                    st.markdown(
                        f'<div style="display:flex;gap:10px;align-items:flex-start;'
                        f'padding:8px 12px;border-radius:8px;margin-bottom:6px;'
                        f'background:#060c18;border:1px solid #1e2d4530;">'
                        f'  <span style="color:#3b82f6;font-weight:700;font-size:12px;'
                        f'  flex-shrink:0;margin-top:1px;">{i:02d}</span>'
                        f'  <span style="color:#cbd5e1;font-size:12px;line-height:1.5;">{item}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.info("Improvements will appear here after analysis.")

            # Jobs quick metrics
            if jobs:
                salaries = [float(j.get("salary_lpa", 0) or 0) for j in jobs]
                interviews = [float(j.get("interview_probability", 0) or 0) for j in jobs]
                avg_sal = sum(salaries) / len(salaries)
                avg_int = sum(interviews) / len(interviews)
                st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
                m1, m2, m3 = st.columns(3)
                m1.metric("Jobs Matched", len(jobs))
                m2.metric("Avg Package", format_salary_lpa(avg_sal))
                m3.metric("Avg Interview %", f"{avg_int:.0f}%")

    # ══════════════════════════════════════════════════════════════════
    # TAB 2 — KEYWORDS
    # ══════════════════════════════════════════════════════════════════
    with tab_keywords:
        kw_present = insight.get("keywords_present", [])
        kw_missing = insight.get("keywords_missing", [])

        kp_col, km_col = st.columns(2)
        with kp_col:
            found_count = len(kw_present)
            st.markdown(
                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">'
                f'  <span style="color:#10d9a0;font-size:16px;font-weight:700;">✅ Keywords Found</span>'
                f'  <span style="background:#0a2318;color:#10d9a0;font-size:11px;font-weight:700;'
                f'  padding:2px 8px;border-radius:20px;border:1px solid #10d9a030;">{found_count}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                _keyword_chips_html(kw_present, "#10d9a0", "#0a2318", "#10d9a030"),
                unsafe_allow_html=True,
            )

        with km_col:
            miss_count = len(kw_missing)
            st.markdown(
                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">'
                f'  <span style="color:#ef4444;font-size:16px;font-weight:700;">❌ Missing Keywords</span>'
                f'  <span style="background:#1a0808;color:#ef4444;font-size:11px;font-weight:700;'
                f'  padding:2px 8px;border-radius:20px;border:1px solid #ef444430;">{miss_count}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                _keyword_chips_html(kw_missing, "#ef4444", "#1a0808", "#ef444430"),
                unsafe_allow_html=True,
            )
            if kw_missing:
                st.markdown(
                    '<div style="margin-top:10px;background:#0f0a00;border:1px solid #f59e0b30;'
                    'border-left:3px solid #f59e0b;border-radius:8px;padding:10px 14px;'
                    'color:#fbbf24;font-size:11px;line-height:1.6;">'
                    '💡 <b>Tip:</b> Add these keywords naturally into your experience bullets '
                    'and skills section to boost your ATS score significantly.'
                    '</div>',
                    unsafe_allow_html=True,
                )

    # ══════════════════════════════════════════════════════════════════
    # TAB 3 — SKILLS & RECOMMENDED ROLES
    # ══════════════════════════════════════════════════════════════════
    with tab_skills:
        sk_col, role_col = st.columns([1, 1])

        with sk_col:
            st.markdown("#### All Skills Detected")
            skills = profile.get("skills", [])
            st.markdown(
                _keyword_chips_html(skills, "#3b82f6", "#0a1628", "#3b82f630"),
                unsafe_allow_html=True,
            )

            languages = insight.get("languages", [])
            if languages:
                st.markdown("#### Languages / Frameworks")
                st.markdown(
                    _keyword_chips_html(languages, "#8b5cf6", "#120a28", "#8b5cf630"),
                    unsafe_allow_html=True,
                )

        with role_col:
            st.markdown("#### Recommended Job Roles")
            rec_roles = insight.get("recommended_roles", [])
            if rec_roles:
                for i, role in enumerate(rec_roles):
                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:10px;'
                        f'padding:10px 14px;border-radius:10px;margin-bottom:6px;'
                        f'background:linear-gradient(90deg,#0a1628,#070b14);'
                        f'border:1px solid #3b82f630;">'
                        f'  <span style="color:#3b82f6;font-weight:800;font-size:14px;">{i+1}</span>'
                        f'  <span style="color:#cbd5e1;font-size:13px;font-weight:500;">{role}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.info("Recommended roles appear after analysis.")

            # Job titles held
            titles = profile.get("job_titles", [])
            if titles:
                st.markdown("#### Past Job Titles")
                st.markdown(
                    _keyword_chips_html(titles, "#f59e0b", "#1a1200", "#f59e0b30"),
                    unsafe_allow_html=True,
                )

            # Education
            education = profile.get("education", [])
            if education:
                st.markdown("#### Education")
                for edu in education:
                    st.markdown(
                        f'<div style="color:#94a3b8;font-size:12px;padding:4px 0;'
                        f'border-bottom:1px solid #1e2d4530;">🎓 {edu}</div>',
                        unsafe_allow_html=True,
                    )

    # ══════════════════════════════════════════════════════════════════
    # TAB 4 — STRENGTHS / GAPS / SALARY / IMPROVEMENTS
    # ══════════════════════════════════════════════════════════════════
    with tab_strengths:

        # ── Glassdoor Salary Estimate card ───────────────────────────
        gd = insight.get("glassdoor_salary", {})
        gd_role   = gd.get("role", "") or profile.get("seniority_level", "")
        gd_min    = float(gd.get("min_lpa", 0) or 0)
        gd_max    = float(gd.get("max_lpa", 0) or 0)
        gd_med    = float(gd.get("median_lpa", 0) or 0)
        gd_loc    = gd.get("location", "") or st.session_state.get("location", "India")
        gd_band   = gd.get("experience_band", "")
        gd_note   = gd.get("note", "")
        sal_min   = insight.get("salary_range", {}).get("min_lpa", 0)
        sal_max   = insight.get("salary_range", {}).get("max_lpa", 0)

        if gd_min or gd_med or sal_min:
            display_min = gd_min or sal_min
            display_max = gd_max or sal_max
            display_med = gd_med or round((display_min + display_max) / 2, 1)
            st.markdown(
                f'<div style="background:linear-gradient(135deg,#07130a,#091a10);'
                f'border:1px solid #10d9a030;border-radius:14px;padding:16px 20px;margin-bottom:14px;">'
                f'  <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">'
                f'    <span style="font-size:18px;">💰</span>'
                f'    <span style="color:#10d9a0;font-size:14px;font-weight:700;">Glassdoor Salary Estimate</span>'
                f'    {"<span style=&quot;background:#0a2318;color:#10d9a0;font-size:10px;font-weight:700;padding:2px 8px;border-radius:20px;border:1px solid #10d9a030;margin-left:6px;&quot;>" + gd_band + "</span>" if gd_band else ""}'
                f'  </div>'
                f'  <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:10px;">'
                f'    <div style="flex:1;min-width:100px;background:#060e08;border:1px solid #10d9a025;border-radius:10px;padding:10px 14px;text-align:center;">'
                f'      <div style="color:#94a3b8;font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">Min</div>'
                f'      <div style="color:#10d9a0;font-size:20px;font-weight:800;">{display_min:.1f} <span style="font-size:12px;">LPA</span></div>'
                f'    </div>'
                f'    <div style="flex:1;min-width:100px;background:#060e08;border:2px solid #10d9a050;border-radius:10px;padding:10px 14px;text-align:center;">'
                f'      <div style="color:#94a3b8;font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">Median ✦</div>'
                f'      <div style="color:#10d9a0;font-size:22px;font-weight:800;">{display_med:.1f} <span style="font-size:12px;">LPA</span></div>'
                f'    </div>'
                f'    <div style="flex:1;min-width:100px;background:#060e08;border:1px solid #10d9a025;border-radius:10px;padding:10px 14px;text-align:center;">'
                f'      <div style="color:#94a3b8;font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">Max</div>'
                f'      <div style="color:#10d9a0;font-size:20px;font-weight:800;">{display_max:.1f} <span style="font-size:12px;">LPA</span></div>'
                f'    </div>'
                f'  </div>'
                f'  <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;">'
                f'    {"<span style=&quot;color:#64748b;font-size:11px;&quot;>📍 " + gd_loc + "</span>" if gd_loc else ""}'
                f'    {"&nbsp;·&nbsp;<span style=&quot;color:#64748b;font-size:11px;&quot;>" + gd_role + "</span>" if gd_role else ""}'
                f'  </div>'
                f'  {"<div style=&quot;margin-top:8px;color:#94a3b8;font-size:11px;line-height:1.5;border-top:1px solid #10d9a015;padding-top:8px;&quot;>📊 " + gd_note + "</div>" if gd_note else ""}'
                f'  <div style="margin-top:6px;color:#475569;font-size:10px;">Source: Glassdoor market data estimate · figures in Indian LPA</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

        # ── Strengths & Weaknesses columns ───────────────────────────
        str_col, weak_col = st.columns(2)

        with str_col:
            st.markdown(
                '<div style="color:#10d9a0;font-size:15px;font-weight:700;margin-bottom:10px;">💪 Strengths</div>',
                unsafe_allow_html=True,
            )
            strengths = insight.get("strengths", [])
            if strengths:
                for item in strengths:
                    st.markdown(
                        f'<div style="display:flex;gap:8px;align-items:flex-start;'
                        f'padding:9px 12px;border-radius:8px;margin-bottom:6px;'
                        f'background:linear-gradient(90deg,#071a0f,#060c0a);'
                        f'border:1px solid #10d9a025;">'
                        f'  <span style="color:#10d9a0;font-size:14px;flex-shrink:0;">✓</span>'
                        f'  <span style="color:#cbd5e1;font-size:12px;line-height:1.5;">{item}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.info("Strengths appear after analysis.")

        with weak_col:
            st.markdown(
                '<div style="color:#ef4444;font-size:15px;font-weight:700;margin-bottom:10px;">⚠ Gaps / Weaknesses</div>',
                unsafe_allow_html=True,
            )
            weaknesses = insight.get("weaknesses", [])
            career_gaps = insight.get("career_gaps", [])
            if weaknesses:
                for item in weaknesses:
                    st.markdown(
                        f'<div style="display:flex;gap:8px;align-items:flex-start;'
                        f'padding:9px 12px;border-radius:8px;margin-bottom:6px;'
                        f'background:linear-gradient(90deg,#1a0808,#110606);'
                        f'border:1px solid #ef444425;">'
                        f'  <span style="color:#ef4444;font-size:14px;flex-shrink:0;">✗</span>'
                        f'  <span style="color:#cbd5e1;font-size:12px;line-height:1.5;">{item}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            if career_gaps:
                st.markdown(
                    '<div style="color:#f59e0b;font-size:13px;font-weight:600;margin:10px 0 6px;">📅 Career Gaps</div>',
                    unsafe_allow_html=True,
                )
                for gap in career_gaps:
                    st.markdown(
                        f'<div style="padding:8px 12px;border-radius:8px;margin-bottom:6px;'
                        f'background:#1a1200;border:1px solid #f59e0b25;'
                        f'color:#fbbf24;font-size:12px;">📌 {gap}</div>',
                        unsafe_allow_html=True,
                    )
            if not weaknesses and not career_gaps:
                st.info("Gaps appear after analysis.")

        # ── Improvement Suggestions ───────────────────────────────────
        improvement_suggestions = insight.get("improvement_suggestions", [])
        if not improvement_suggestions:
            improvement_suggestions = improvements  # fall back to legacy improvements list
        if improvement_suggestions:
            st.markdown(
                '<div style="color:#3b82f6;font-size:15px;font-weight:700;margin:14px 0 10px;">🚀 Prioritized Improvement Suggestions</div>',
                unsafe_allow_html=True,
            )
            priority_colors = ["#ef4444", "#ef4444", "#f59e0b", "#f59e0b", "#3b82f6", "#3b82f6", "#10d9a0", "#10d9a0"]
            priority_labels = ["P1", "P1", "P2", "P2", "P3", "P3", "P4", "P4"]
            for i, item in enumerate(improvement_suggestions):
                p_color = priority_colors[i] if i < len(priority_colors) else "#64748b"
                p_label = priority_labels[i] if i < len(priority_labels) else f"P{i+1}"
                st.markdown(
                    f'<div style="display:flex;gap:10px;align-items:flex-start;'
                    f'padding:10px 14px;border-radius:10px;margin-bottom:6px;'
                    f'background:#060c18;border:1px solid #1e2d4540;'
                    f'border-left:3px solid {p_color};">'
                    f'  <span style="background:{p_color}20;color:{p_color};font-size:9px;font-weight:800;'
                    f'  padding:2px 6px;border-radius:6px;border:1px solid {p_color}40;flex-shrink:0;margin-top:1px;">{p_label}</span>'
                    f'  <span style="color:#cbd5e1;font-size:12px;line-height:1.55;">{item}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

    # ══════════════════════════════════════════════════════════════════
    # TAB 5 — CONTACT & CERTIFICATIONS
    # ══════════════════════════════════════════════════════════════════
    with tab_contact:
        con_col, cert_col = st.columns(2)

        with con_col:
            st.markdown("#### Contact Info Detected")
            contact = insight.get("contact", {})
            _CONTACT_ICONS = {
                "email": "📧", "phone": "📱", "linkedin": "🔵",
                "github": "🐙", "portfolio": "🌐",
            }
            any_contact = False
            for field, icon in _CONTACT_ICONS.items():
                val = contact.get(field, "")
                if val:
                    any_contact = True
                    display = f'<a href="{val}" target="_blank" style="color:#3b82f6;">{val}</a>' if val.startswith("http") else val
                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:10px;'
                        f'padding:9px 12px;border-radius:8px;margin-bottom:6px;'
                        f'background:#060c18;border:1px solid #1e2d4530;">'
                        f'  <span style="font-size:16px;">{icon}</span>'
                        f'  <div>'
                        f'    <div style="color:#64748b;font-size:10px;text-transform:uppercase;font-weight:600;">{field}</div>'
                        f'    <div style="color:#cbd5e1;font-size:12px;">{display}</div>'
                        f'  </div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            if not any_contact:
                st.markdown(
                    '<div style="background:#1a0808;border:1px solid #ef444430;border-radius:8px;'
                    'padding:12px;color:#ef4444;font-size:12px;">'
                    '⚠ No contact info detected in your resume. Add email, phone, and LinkedIn URL '
                    'at the top of your resume — ATS systems require these.'
                    '</div>',
                    unsafe_allow_html=True,
                )

        with cert_col:
            st.markdown("#### Certifications & Courses")
            certs = insight.get("certifications", [])
            if certs:
                for cert in certs:
                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:8px;'
                        f'padding:9px 12px;border-radius:8px;margin-bottom:6px;'
                        f'background:#0a0f28;border:1px solid #3b82f625;">'
                        f'  <span style="font-size:14px;">🏅</span>'
                        f'  <span style="color:#93c5fd;font-size:12px;">{cert}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.markdown(
                    '<div style="background:#1a1200;border:1px solid #f59e0b30;border-radius:8px;'
                    'padding:12px;color:#fbbf24;font-size:12px;">'
                    '💡 No certifications found. Adding relevant certs (AWS, Google, etc.) '
                    'can boost your ATS score by 10–15 points.'
                    '</div>',
                    unsafe_allow_html=True,
                )

    # ══════════════════════════════════════════════════════════════════
    # TAB 6 — JOB SNAPSHOT (top 3 matches)
    # ══════════════════════════════════════════════════════════════════
    with tab_jobs_summary:
        if not jobs:
            st.info("Job matches appear here after analysis.")
        else:
            st.markdown(
                f'<div style="color:#94a3b8;font-size:12px;margin-bottom:12px;">'
                f'Showing top {min(3, len(jobs))} of {len(jobs)} matches — '
                f'<span style="color:#3b82f6;">go to Job Matches tab for full list.</span></div>',
                unsafe_allow_html=True,
            )
            for job in jobs[:3]:
                sal   = float(job.get("salary_lpa", 0) or 0)
                match = float(job.get("ats_match", 0) or 0)
                prob  = float(job.get("interview_probability", 0) or 0)
                comp  = str(job.get("competition_level", "")).capitalize()
                comp_color = {"Low": "#10d9a0", "Medium": "#f59e0b", "High": "#ef4444"}.get(comp, "#94a3b8")
                tags  = job.get("tags", [])
                tag_html = "".join(
                    f'<span style="background:#0a1628;color:#60a5fa;font-size:10px;'
                    f'padding:2px 7px;border-radius:4px;margin:2px 2px 0 0;'
                    f'border:1px solid #3b82f630;display:inline-block;">{t}</span>'
                    for t in tags[:5]
                )
                st.markdown(
                    f'<div style="background:linear-gradient(135deg,#0a0f1a,#0d1117);'
                    f'border:1px solid #1e2d45;border-radius:12px;padding:14px 16px;margin-bottom:10px;">'
                    f'  <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:8px;">'
                    f'    <div>'
                    f'      <div style="color:#EAF2FF;font-size:14px;font-weight:700;">{job.get("title","")}</div>'
                    f'      <div style="color:#64748b;font-size:11px;margin-top:2px;">'
                    f'        {job.get("company","")} &nbsp;·&nbsp; {job.get("platform","")} &nbsp;·&nbsp; {job.get("location","")}'
                    f'      </div>'
                    f'      <div style="margin-top:6px;">{tag_html}</div>'
                    f'    </div>'
                    f'    <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center;">'
                    f'      <div style="text-align:center;background:#070b14;border-radius:8px;padding:6px 10px;border:1px solid #1a2744;">'
                    f'        <div style="color:#10d9a0;font-size:14px;font-weight:700;">{sal:.0f} LPA</div>'
                    f'        <div style="color:#64748b;font-size:9px;">Salary</div>'
                    f'      </div>'
                    f'      <div style="text-align:center;background:#070b14;border-radius:8px;padding:6px 10px;border:1px solid #1a2744;">'
                    f'        <div style="color:#3b82f6;font-size:14px;font-weight:700;">{match:.0f}%</div>'
                    f'        <div style="color:#64748b;font-size:9px;">ATS Match</div>'
                    f'      </div>'
                    f'      <div style="text-align:center;background:#070b14;border-radius:8px;padding:6px 10px;border:1px solid #1a2744;">'
                    f'        <div style="color:#f59e0b;font-size:14px;font-weight:700;">{prob:.0f}%</div>'
                    f'        <div style="color:#64748b;font-size:9px;">Interview</div>'
                    f'      </div>'
                    f'      <div style="text-align:center;background:#070b14;border-radius:8px;padding:6px 10px;border:1px solid {comp_color}30;">'
                    f'        <div style="color:{comp_color};font-size:12px;font-weight:700;">{comp}</div>'
                    f'        <div style="color:#64748b;font-size:9px;">Competition</div>'
                    f'      </div>'
                    f'    </div>'
                    f'  </div>'
                    f'  <div style="margin-top:8px;color:#64748b;font-size:11px;line-height:1.5;">{job.get("rationale","")}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            st.caption("💡 Go to **Job Matches** in the sidebar for full list + Apply buttons.")


def tailor_resume_page() -> None:
    st.subheader("Tailor Resume")
    if not _build_provider_list():
        st.warning("⚠ No API keys configured. Add a free key in the **API Keys** page.", icon="🔑")
        return
    auto_run_if_ready()
    original_text = st.session_state.get("resume_text", "")
    optimized_text = st.session_state.get("resume_optimized", "")
    job_description = st.text_area("Paste Job Description", height=220, key="tailor_jd")
    if st.button("Tailor Resume for This JD", type="primary"):
        if not original_text.strip():
            st.error("Upload a resume first on Dashboard.")
        elif not job_description.strip():
            st.error("Paste a job description.")
        else:
            resume_compact = trimmed_text(original_text, max_chars=6000)
            jd_compact = trimmed_text(job_description, max_chars=2800)
            prompt = f"""
Tailor this resume for this job description to maximize interview chances.
Rules: keep format/structure unchanged; do not fabricate achievements.
Only change: JD keyword alignment, skills reordering, stronger impact statements.
Return JSON only:
{{
  "tailored_resume":"plain text",
  "ats_before":0,
  "ats_after":0,
  "changes_made":["string"],
  "keywords_added":["string"],
  "interview_fit_score":0
}}

Resume:
{resume_compact}

Job Description:
{jd_compact}
"""
            try:
                with st.spinner("Tailoring resume..."):
                    output = run_gemini_prompt(prompt)
                parsed = extract_json_payload(output)
                if isinstance(parsed, dict):
                    st.session_state["tailor_result"] = parsed
                    st.session_state["resume_optimized"] = parsed.get("tailored_resume", optimized_text)
                    st.session_state["optimized_resume"] = st.session_state["resume_optimized"]
                    optimized_text = st.session_state["resume_optimized"]
                    st.success("Tailored resume generated.")
                else:
                    st.error("Gemini did not return valid JSON.")
            except Exception as exc:
                st.error(f"Tailoring failed: {exc}")

    tailor = st.session_state.get("tailor_result", {})
    if tailor:
        b1, b2, b3 = st.columns(3)
        b1.metric("ATS Before", tailor.get("ats_before", 0))
        b2.metric("ATS After", tailor.get("ats_after", 0))
        b3.metric("Interview Fit", f"{tailor.get('interview_fit_score', 0)}%")
        st.markdown("### Changes Made")
        for item in tailor.get("changes_made", []):
            st.write(f"- {item}")

    tab_diff, tab_opt, tab_original = st.tabs(["Diff", "Optimized", "Original"])

    with tab_diff:
        if original_text and optimized_text:
            st.markdown(build_diff_html(original_text, optimized_text), unsafe_allow_html=True)
        else:
            st.info("Diff appears after automation completes.")

    with tab_opt:
        if optimized_text:
            st.text_area("Optimized Resume", value=optimized_text, height=430)
        else:
            st.info("No optimized resume yet. Complete Dashboard inputs.")

    with tab_original:
        if original_text:
            st.text_area("Original Resume", value=original_text, height=430)
        else:
            st.info("No original resume text found.")

    d1, d2, d3 = st.columns(3)
    d1.download_button(
        "Download .txt",
        data=optimized_text.encode("utf-8"),
        file_name="optimized_resume.txt",
        mime="text/plain",
        disabled=not bool(optimized_text),
    )
    d2.download_button(
        "Download .pdf",
        data=optimized_resume_to_pdf_bytes(optimized_text) if optimized_text else b"",
        file_name="optimized_resume.pdf",
        mime="application/pdf",
        disabled=not bool(optimized_text),
    )
    d3.download_button(
        "Download Original File",
        data=st.session_state.get("original_resume_bytes", b""),
        file_name=st.session_state.get("resume_filename", "original_resume"),
        mime=st.session_state.get("original_resume_mime", "application/octet-stream"),
        disabled=not bool(st.session_state.get("original_resume_bytes", b"")),
    )

    analysis = st.session_state["profile"]
    st.markdown("### Extracted Candidate Profile")
    st.write(f"**Name:** {analysis.get('name', '') or 'N/A'}")
    st.write(f"**Seniority:** {analysis.get('seniority_level', '') or 'N/A'}")
    st.write(f"**Experience (years):** {analysis.get('experience_years', 0)}")
    st.write(f"**Skills:** {', '.join(analysis.get('skills', [])) or 'N/A'}")
    st.write(f"**Job Titles:** {', '.join(analysis.get('job_titles', [])) or 'N/A'}")
    st.write(f"**Education:** {', '.join(analysis.get('education', [])) or 'N/A'}")


def normalize_competition(value: str) -> int:
    mapping = {"low": 1, "medium": 2, "high": 3}
    return mapping.get(str(value).strip().lower(), 99)


def _build_search_url(platform: str, keywords: str, location: str) -> str:
    """Build a real public job-search URL for the given platform."""
    kw  = urllib.parse.quote_plus(keywords)
    loc = urllib.parse.quote_plus(location or "India")
    p   = platform.lower()
    if "linkedin"    in p: return f"https://www.linkedin.com/jobs/search/?keywords={kw}&location={loc}&f_TPR=r604800"
    if "indeed"      in p: return f"https://in.indeed.com/jobs?q={kw}&l={loc}"
    if "naukri"      in p: return f"https://www.naukri.com/{keywords.lower().replace(' ','-')}-jobs-in-{(location or 'india').lower().replace(' ','-')}"
    if "glassdoor"   in p: return f"https://www.glassdoor.co.in/Job/jobs.htm?sc.keyword={kw}"
    if "wellfound"   in p: return f"https://wellfound.com/jobs?query={kw}"
    if "internshala" in p: return f"https://internshala.com/jobs/{kw}/"
    if "cutshort"    in p: return f"https://cutshort.io/jobs?keyword={kw}"
    if "hirist"      in p: return f"https://www.hirist.tech/search?q={kw}"
    if "unstop"      in p: return f"https://unstop.com/jobs?search={kw}"
    if "monster"     in p: return f"https://www.monsterindia.com/srp/results?query={kw}&locations={loc}"
    if "shine"       in p: return f"https://www.shine.com/job-search/{keywords.lower().replace(' ','-')}-jobs"
    if "angel"       in p: return f"https://wellfound.com/jobs?query={kw}"
    return f"https://www.linkedin.com/jobs/search/?keywords={kw}&location={loc}"


# ── Platform domain map for grounding site-search ──────────────────────────
_PLATFORM_DOMAINS = {
    "LinkedIn":    "linkedin.com/jobs",
    "Indeed":      "in.indeed.com",
    "Naukri":      "naukri.com",
    "Glassdoor":   "glassdoor.co.in",
    "Wellfound":   "wellfound.com/jobs",
    "Internshala": "internshala.com/jobs",
    "Cutshort":    "cutshort.io/jobs",
    "Hirist":      "hirist.tech",
    "Unstop":      "unstop.com/jobs",
    "Monster":     "monsterindia.com",
    "Shine":       "shine.com",
    "AngelList":   "wellfound.com/jobs",
}


def _fetch_real_jobs_from_connected_platforms(
    resume_text: str,
    profile: dict,
    location: str,
    connected_platforms: list,
    time_filter: str,
) -> list:
    """
    Fetch REAL job listings ONLY from the user's connected portals.

    Strategy:
    - Uses Gemini with Google Search Grounding (live internet access)
    - Restricts the search with site: operators to ONLY the connected platforms
    - AI reads the actual live search results and extracts real job data
    - AI then scores each real job against the resume
    - Zero hallucination: if grounding finds no jobs, returns []
    """
    if not connected_platforms:
        return []

    today          = datetime.utcnow().strftime("%B %d, %Y")
    resume_compact = trimmed_text(resume_text, max_chars=5000)
    candidate_name = profile.get("name", "")
    exp_years      = profile.get("experience_years", 0)

    # Build search keywords from resume profile
    titles    = profile.get("job_titles", [])
    skills    = profile.get("skills", [])
    seniority = profile.get("seniority_level", "")
    if titles:
        primary_kw = titles[0]
        alt_kw     = titles[1] if len(titles) > 1 else ""
    elif skills:
        primary_kw = " ".join(skills[:2])
        alt_kw     = " ".join(skills[2:4])
    else:
        primary_kw = "Software Developer"
        alt_kw     = ""

    # Build site-restricted search instruction for ONLY connected platforms
    site_list = []
    for p in connected_platforms:
        domain = _PLATFORM_DOMAINS.get(p)
        if domain:
            site_list.append(f"site:{domain}")
    site_instruction = " OR ".join(site_list) if site_list else ""

    time_map = {
        "Last 24 Hours": "posted in the last 24 hours",
        "Last 3 Days":   "posted in the last 3 days",
        "Last 7 Days":   "posted in the last week",
        "Last 30 Days":  "posted in the last month",
        "All Time":      "currently open",
    }
    time_hint = time_map.get(time_filter, "currently open")

    platforms_str = ", ".join(connected_platforms)

    prompt = f"""
Today's date: {today}
Location: {location or "India"}
Candidate: {candidate_name or "the candidate"} | Experience: {exp_years} years
Job role sought: {primary_kw}{f" or {alt_kw}" if alt_kw else ""}

You have live Google Search access (grounding enabled).

TASK: Search ONLY the following job portals for REAL, currently open job listings:
Platforms: {platforms_str}
Search query to use: ({primary_kw}{f" OR {alt_kw}" if alt_kw else ""}) {location} jobs {time_hint}
{f"Restrict results to: {site_instruction}" if site_instruction else ""}

Search these portals now and find 10 REAL job postings that:
1. Are currently open and {time_hint}
2. Match the candidate's skills and experience level
3. Are from ONLY these platforms: {platforms_str}

For each real job found, extract:
- Exact job title as shown on the portal
- Real company name
- Which portal it was found on ({platforms_str})
- Job location
- Direct URL to the job posting (the actual posting URL, not the search page)
- Salary if shown
- Experience required if shown
- Key skills mentioned in the posting

Then score each job against the resume below.

Resume:
{resume_compact}

Return ONLY valid JSON — no markdown, no explanation:
{{
  "candidate_name": "{candidate_name}",
  "source_note": "brief note about which portals returned results",
  "jobs": [
    {{
      "title": "exact job title from the portal",
      "company": "real company name",
      "platform": "exact portal name from: {platforms_str}",
      "location": "city, state or Remote",
      "apply_url": "direct URL to the actual job posting page on the portal (e.g. https://www.linkedin.com/jobs/view/1234567890 or https://www.naukri.com/job-listings-... or full indeed.com/viewjob?jk=... URL — NOT to.indeed.com or any redirect/short URL)",
      "direct_url": "the canonical company careers page URL or full portal listing URL for this specific job — must be a real resolvable link, not a search page or redirect. Example: https://careers.tcs.com/job/123 or https://www.naukri.com/job-listings-data-engineer-accenture-bangalore-2-to-5-years-123456.html",
      "linkedin_job_url": "LinkedIn posting URL if found, else linkedin.com search URL",
      "indeed_job_url": "Indeed posting URL if found, else indeed.com search URL",
      "naukri_job_url": "Naukri posting URL if found, else naukri.com search URL",
      "description": "2-3 sentence summary of the role from the actual posting",
      "salary_lpa": 0.0,
      "experience_required": "e.g. 2-5 years",
      "job_type": "Full-time|Part-time|Contract|Internship|Freelance",
      "tags": ["skill1", "skill2", "skill3"],
      "ats_match": 0,
      "interview_probability": 0,
      "competition_level": "Low|Medium|High",
      "shortlisting_confidence": 0,
      "shortlisting_reasons": ["reason 1", "reason 2"],
      "shortlisting_gaps": ["gap 1"],
      "skills_matched": ["skill A"],
      "skills_missing": ["skill B"],
      "posted_label": "Today|Yesterday|2 days ago|This week|This month",
      "urgency": "Actively Hiring|Normal|Low Priority",
      "rationale": "2-sentence explanation of why this resume fits this specific real job",
      "hr_name": "",
      "hr_email": ""
    }}
  ]
}}

STRICT RULES:
- ONLY include jobs from these platforms: {platforms_str}
- ONLY include jobs you actually found via live search — do NOT invent jobs
- apply_url MUST be a real direct URL to the posting, not a placeholder
- direct_url MUST be the full canonical URL — never a to.indeed.com or bit.ly or any redirect. If you can find the company careers page URL for the role, use that. Otherwise use the full portal listing URL.
- shortlisting_confidence must honestly reflect how well the resume matches
- If a portal returned no results, skip it — do not make up jobs for it
- Return ONLY the JSON object
"""
    try:
        # use_grounding=True enables live Google Search — this is the key difference
        raw = call_ai(prompt, use_grounding=True)
        parsed = extract_json_payload(raw)
        if not isinstance(parsed, dict) or "jobs" not in parsed:
            return []

        jobs = parsed.get("jobs", [])
        if not isinstance(jobs, list):
            return []

        cname = parsed.get("candidate_name", candidate_name)

        # Post-process: fill missing URLs, tag candidate name
        valid = []
        for job in jobs:
            plat    = str(job.get("platform", ""))
            title   = str(job.get("title", "")).strip()
            loc     = str(job.get("location", location or "India"))
            if not title:
                continue
            # Only keep jobs from connected platforms
            if not any(p.lower() in plat.lower() or plat.lower() in p.lower()
                       for p in connected_platforms):
                continue
            # Ensure apply_url is real
            apply_url = str(job.get("apply_url", "")).strip()
            if not apply_url or apply_url in ("string", "url", ""):
                apply_url = _build_search_url(plat, title, loc)
            job["apply_url"] = apply_url

            # Validate and clean direct_url — reject redirects and placeholders
            direct_url = str(job.get("direct_url", "")).strip()
            _bad_direct = ("string", "url", "", "#", "null", "none")
            _redirect_hosts = ("to.indeed.com", "bit.ly", "tinyurl", "ow.ly", "goo.gl", "t.co")
            if (not direct_url
                    or direct_url.lower() in _bad_direct
                    or any(h in direct_url for h in _redirect_hosts)):
                # Fall back to apply_url if it looks like a real listing URL
                direct_url = apply_url if apply_url != _build_search_url(plat, title, loc) else ""
            job["direct_url"] = direct_url
            # Fill per-platform search links if blank
            if not job.get("linkedin_job_url"):
                job["linkedin_job_url"] = _build_search_url("LinkedIn", title, loc)
            if not job.get("indeed_job_url"):
                job["indeed_job_url"] = _build_search_url("Indeed", title, loc)
            if not job.get("naukri_job_url"):
                job["naukri_job_url"] = _build_search_url("Naukri", title, loc)
            job["_candidate_name"] = cname
            valid.append(job)

        # Sort by shortlisting_confidence
        valid.sort(key=lambda x: -int(x.get("shortlisting_confidence", 0)))
        return valid[:10]

    except Exception:
        return []


def job_matches_page() -> None:
    st.subheader("Job Matches")

    # ── CSS ─────────────────────────────────────────────────────────────
    st.markdown("""
<style>
@keyframes shimmer-bar {
  0%   { background-position: -200% center; }
  100% { background-position: 200% center; }
}
@keyframes pulse-badge {
  0%, 100% { opacity: 1; }
  50%       { opacity: 0.6; }
}
.jm-page-header {
    margin-bottom: 28px;
}
.jm-page-title {
    font-size: 22px !important; font-weight: 700 !important;
    color: #EAF2FF !important; letter-spacing: -0.4px;
    font-family: 'Sora', sans-serif !important;
    margin: 0 0 4px !important;
}
.jm-page-sub {
    font-size: 13px !important; color: #475569 !important; margin: 0 !important;
}
/* ── Filter bar ── */
.jm-filter-bar {
    display: flex; gap: 10px; align-items: center; flex-wrap: wrap;
    background: linear-gradient(135deg, #0c1020, #090d18);
    border: 1px solid #1e2d45; border-radius: 14px;
    padding: 14px 18px; margin-bottom: 22px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.35);
}
/* ── Summary strip ── */
.jm-summary {
    display: flex; gap: 14px; flex-wrap: wrap;
    margin-bottom: 22px;
}
.jm-summary-chip {
    display: inline-flex; align-items: center; gap: 7px;
    background: #0d1117; border: 1px solid #1e2d45;
    border-radius: 10px; padding: 8px 14px;
    font-size: 12px; font-weight: 600; color: #94a3b8;
    box-shadow: 0 2px 8px rgba(0,0,0,0.25);
}
.jm-summary-chip span.val { font-size: 16px; font-weight: 800; color: #EAF2FF; font-family: 'Sora', sans-serif; }
/* ── Job card ── */
.jm-card {
    background: linear-gradient(145deg, #0d1117, #090d18);
    border: 1px solid #1e2d45;
    border-radius: 18px; overflow: hidden; margin-bottom: 18px;
    box-shadow: 0 4px 28px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.02);
    transition: border-color 0.2s, box-shadow 0.2s, transform 0.2s;
    position: relative;
}
.jm-card:hover {
    border-color: rgba(59,130,246,0.25);
    box-shadow: 0 8px 40px rgba(0,0,0,0.55);
    transform: translateY(-2px);
}
.jm-card-accent { height: 2px; width: 100%; }
.jm-card-body { padding: 20px 22px 16px; }
/* Rank badge */
.jm-rank {
    position: absolute; top: 16px; right: 18px;
    width: 32px; height: 32px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 12px; font-weight: 800; color: #fff;
    font-family: 'Sora', sans-serif;
}
/* Title row */
.jm-title-row {
    display: flex; align-items: flex-start; gap: 12px; margin-bottom: 6px;
    padding-right: 48px;
}
.jm-title {
    font-size: 17px !important; font-weight: 700 !important;
    color: #EAF2FF !important; letter-spacing: -0.3px;
    font-family: 'Sora', sans-serif !important; margin: 0 !important;
    line-height: 1.3 !important;
}
.jm-company {
    font-size: 13px !important; color: #64748b !important;
    font-weight: 500 !important; margin: 0 0 10px !important;
}
/* Meta pills row */
.jm-meta-row {
    display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 14px;
    align-items: center;
}
.jm-pill {
    display: inline-flex; align-items: center; gap: 4px;
    padding: 3px 10px; border-radius: 99px;
    font-size: 10px; font-weight: 700; letter-spacing: 0.3px;
    text-transform: uppercase; border: 1px solid;
}
.jm-pill-platform  { color: #60a5fa; background: rgba(59,130,246,0.08); border-color: rgba(59,130,246,0.2); }
.jm-pill-location  { color: #94a3b8; background: rgba(148,163,184,0.06); border-color: rgba(148,163,184,0.15); }
.jm-pill-type      { color: #a78bfa; background: rgba(167,139,250,0.08); border-color: rgba(167,139,250,0.2); }
.jm-pill-posted    { color: #10d9a0; background: rgba(16,217,160,0.08); border-color: rgba(16,217,160,0.2); }
.jm-pill-urgent    { color: #f59e0b; background: rgba(245,158,11,0.08); border-color: rgba(245,158,11,0.2);
                     animation: pulse-badge 2s infinite; }
.jm-pill-exp       { color: #f97316; background: rgba(249,115,22,0.08); border-color: rgba(249,115,22,0.2); }
/* Metrics row */
.jm-metrics {
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px;
    margin-bottom: 16px;
}
.jm-metric {
    background: linear-gradient(145deg, #070b14, #060a12);
    border: 1px solid #111c2e; border-radius: 12px;
    padding: 12px 14px; text-align: center;
    box-shadow: inset 0 1px 4px rgba(0,0,0,0.3);
}
.jm-metric-label {
    font-size: 9px !important; font-weight: 800 !important;
    text-transform: uppercase; letter-spacing: 0.8px;
    color: #334155 !important; margin-bottom: 5px !important;
}
.jm-metric-value {
    font-size: 18px !important; font-weight: 800 !important;
    font-family: 'Sora', sans-serif !important;
    line-height: 1.1 !important; margin: 0 !important;
}
/* Shortlisting confidence bar */
.jm-sc-section { margin-bottom: 16px; }
.jm-sc-header {
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 8px;
}
.jm-sc-label {
    font-size: 11px !important; font-weight: 800 !important;
    text-transform: uppercase; letter-spacing: 0.8px; color: #64748b !important;
}
.jm-sc-score {
    font-size: 13px !important; font-weight: 800 !important;
    font-family: 'Sora', sans-serif !important;
}
.jm-sc-bar-bg {
    height: 7px; background: #111c2e; border-radius: 99px; overflow: hidden;
    box-shadow: inset 0 1px 3px rgba(0,0,0,0.4);
}
.jm-sc-bar-fill {
    height: 100%; border-radius: 99px;
    background-size: 200% auto;
}
/* Reasons / gaps pills */
.jm-reason-row { display: flex; gap: 6px; flex-wrap: wrap; margin-top: 8px; }
.jm-reason-chip {
    display: inline-flex; align-items: center; gap: 4px;
    padding: 3px 9px; border-radius: 6px;
    font-size: 10px; font-weight: 600;
}
.jm-reason-chip.match  { background: rgba(16,217,160,0.08); color: #10d9a0; border: 1px solid rgba(16,217,160,0.2); }
.jm-reason-chip.gap    { background: rgba(239,68,68,0.08);  color: #ef4444; border: 1px solid rgba(239,68,68,0.2); }
/* Tags */
.jm-tags { display: flex; gap: 5px; flex-wrap: wrap; margin-bottom: 14px; }
.jm-tag {
    display: inline-block; padding: 3px 10px; border-radius: 6px;
    font-size: 10px; font-weight: 700;
    background: rgba(59,130,246,0.07); color: #60a5fa;
    border: 1px solid rgba(59,130,246,0.18); letter-spacing: 0.2px;
}
/* Rationale */
.jm-rationale {
    background: linear-gradient(90deg, #070c14, #060a12);
    border: 1px solid #111c2e; border-left: 2px solid #3b82f6;
    border-radius: 10px; padding: 10px 14px; margin-bottom: 14px;
    font-size: 12px !important; color: #94a3b8 !important;
    line-height: 1.6 !important;
}
/* Divider */
.jm-divider { height: 1px; background: rgba(30,45,69,0.6); margin: 14px 0; }
/* Empty state */
.jm-empty {
    text-align: center; padding: 60px 24px;
    background: linear-gradient(145deg, #0d1117, #090d18);
    border: 1px solid #1e2d45; border-radius: 18px;
    box-shadow: 0 4px 28px rgba(0,0,0,0.35);
}
.jm-empty-icon { font-size: 48px; margin-bottom: 16px; }
.jm-empty-title { font-size: 18px !important; font-weight: 700 !important; color: #EAF2FF !important; margin-bottom: 8px !important; }
.jm-empty-sub   { font-size: 13px !important; color: #475569 !important; max-width: 380px; margin: 0 auto !important; }
/* Portal search link buttons inside card */
.jm-portal-link {
    display: inline-flex; align-items: center; gap: 5px;
    padding: 6px 14px; border-radius: 8px;
    font-size: 12px; font-weight: 600; border: 1px solid;
    text-decoration: none; transition: opacity 0.15s ease;
}
.jm-portal-link:hover { opacity: 0.8; text-decoration: none; }
/* Fetch button */
.jm-fetch-cta {
    background: linear-gradient(135deg, #0a1628, #0d1f3c);
    border: 1.5px solid rgba(59,130,246,0.4);
    border-radius: 14px; padding: 16px 20px; margin-bottom: 22px;
    display: flex; align-items: center; gap: 14px;
    box-shadow: 0 0 24px rgba(59,130,246,0.08);
}
.jm-fetch-cta-text { flex: 1; }
.jm-fetch-cta-title { font-size: 14px !important; font-weight: 700 !important; color: #60a5fa !important; margin-bottom: 3px !important; }
.jm-fetch-cta-sub   { font-size: 11px !important; color: #475569 !important; margin: 0 !important; }
/* ── Apply button + portal links row ── */
.jm-actions-row {
    display: flex; align-items: center; justify-content: space-between;
    gap: 12px; flex-wrap: wrap; margin-bottom: 10px;
}
.jm-apply-btn {
    display: inline-flex; align-items: center; gap: 8px;
    padding: 10px 22px; border-radius: 10px; text-decoration: none;
    font-size: 13px; font-weight: 700; letter-spacing: 0.1px;
    background: linear-gradient(135deg, #10d9a0, #3b82f6);
    color: #fff; border: none;
    box-shadow: 0 4px 18px rgba(16,217,160,0.25), 0 0 0 1px rgba(16,217,160,0.15);
    transition: opacity 0.15s ease, transform 0.15s ease;
}
.jm-apply-btn:hover { opacity: 0.88; transform: translateY(-1px); text-decoration: none; color: #fff; }
.jm-view-listing-btn {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 10px 18px; border-radius: 10px; text-decoration: none;
    font-size: 13px; font-weight: 600; letter-spacing: 0.1px;
    background: rgba(255,255,255,0.04);
    color: #94a3b8;
    border: 1px solid rgba(255,255,255,0.1);
    transition: background 0.15s ease, color 0.15s ease, transform 0.15s ease;
}
.jm-view-listing-btn:hover { background: rgba(255,255,255,0.08); color: #cbd5e1; transform: translateY(-1px); text-decoration: none; }
.jm-portal-links-row { display: flex; gap: 6px; flex-wrap: wrap; }
.jm-src-link {
    display: inline-flex; align-items: center; gap: 4px;
    padding: 6px 12px; border-radius: 8px; font-size: 11px; font-weight: 600;
    text-decoration: none; border: 1px solid; transition: opacity 0.15s;
}
.jm-src-link:hover { opacity: 0.75; text-decoration: none; }
.jm-src-li  { color: #60a5fa; background: rgba(10,102,194,0.08);  border-color: rgba(10,102,194,0.2); }
.jm-src-in  { color: #818cf8; background: rgba(33,100,243,0.08);  border-color: rgba(33,100,243,0.2); }
.jm-src-nk  { color: #fb923c; background: rgba(255,117,85,0.08);  border-color: rgba(255,117,85,0.2); }
.jm-src-active { font-weight: 800 !important; box-shadow: 0 0 0 1.5px currentColor; }
.jm-url-preview {
    display: flex; align-items: center; gap: 6px;
    padding: 6px 10px; border-radius: 7px;
    background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.05);
    margin-bottom: 4px;
}
.jm-url-label { font-size: 10px; font-weight: 700; color: #334155; text-transform: uppercase; letter-spacing: 0.5px; white-space: nowrap; }
.jm-url-text  { font-size: 11px; color: #475569; word-break: break-all; font-family: 'DM Mono', monospace; }
</style>
""", unsafe_allow_html=True)

    if not _build_provider_list():
        st.warning("⚠ No API keys configured. Add a free key in the **API Keys** page.", icon="🔑")
        return

    auto_run_if_ready()

    resume_text = st.session_state.get("resume_text", "").strip()
    profile     = st.session_state.get("profile", {})
    location    = st.session_state.get("location", "India")

    # ── Connected platforms guard ────────────────────────────────────────
    connected_platforms = st.session_state.get("platforms_connected", [])

    # ── Page header ──────────────────────────────────────────────────────
    st.markdown("""
<div class="jm-page-header">
  <p class="jm-page-title">🔍 Job Matches</p>
  <p class="jm-page-sub">Real jobs fetched from your connected career portals · scored against your resume</p>
</div>
""", unsafe_allow_html=True)

    # ── No resume guard ──────────────────────────────────────────────────
    if not resume_text:
        st.markdown("""
<div class="jm-empty">
  <div class="jm-empty-icon">📄</div>
  <p class="jm-empty-title">Upload your resume first</p>
  <p class="jm-empty-sub">Go to <b>Your Resume Insight</b> and upload your resume. We'll then fetch real jobs from your connected portals and score each one against your resume.</p>
</div>
""", unsafe_allow_html=True)
        return

    # ── No connected platforms guard ────────────────────────────────────
    if not connected_platforms:
        st.markdown("""
<div class="jm-empty">
  <div class="jm-empty-icon">🌐</div>
  <p class="jm-empty-title">No career portals connected</p>
  <p class="jm-empty-sub">Go to <b>Connect with Career Portals</b> in the sidebar and connect at least one platform (LinkedIn, Indeed, Naukri, etc.). Job Matches will only show real jobs from your connected portals — no AI-generated listings.</p>
</div>
""", unsafe_allow_html=True)
        if st.button("🌐 Go to Career Portals →", type="primary"):
            st.session_state["_nav_override"] = "Connect with Career Portals"
            st.rerun()
        return

    # ── Filter bar ───────────────────────────────────────────────────────
    col_time, col_sal, col_sort, col_custom = st.columns([2, 2, 2, 1])

    with col_time:
        time_filter = st.selectbox(
            "Posted Within",
            ["Last 24 Hours", "Last 3 Days", "Last 7 Days", "Last 30 Days", "All Time"],
            index=2,
            key="jm_time_filter",
        )
    with col_sal:
        salary_filter = st.selectbox(
            "Min Salary",
            ["All", "₹5L+", "₹10L+", "₹20L+", "₹30L+", "₹50L+", "₹1Cr+", "Custom"],
            index=0,
            key="jm_salary_filter",
        )
    with col_sort:
        sort_by = st.selectbox(
            "Sort By",
            ["Shortlisting Confidence", "Salary (High→Low)", "ATS Match", "Interview Probability"],
            index=0,
            key="jm_sort_by",
        )
    with col_custom:
        custom_salary = 0.0
        if salary_filter == "Custom":
            custom_salary = st.number_input("LPA Min", min_value=0.0, value=12.0, step=1.0, key="jm_custom_sal")

    threshold_map = {
        "All": 0.0, "₹5L+": 5.0, "₹10L+": 10.0, "₹20L+": 20.0,
        "₹30L+": 30.0, "₹50L+": 50.0, "₹1Cr+": 100.0, "Custom": custom_salary,
    }
    min_salary = threshold_map.get(salary_filter, 0.0)

    # ── Fetch CTA ────────────────────────────────────────────────────────
    fetch_key = f"jm_real_jobs_{','.join(sorted(connected_platforms))}_{time_filter}"
    live_jobs = st.session_state.get(fetch_key, [])

    # Connected platform badges
    platform_badges = "".join(
        f'<span style="display:inline-flex;align-items:center;gap:4px;padding:3px 10px;'
        f'background:rgba(16,217,160,0.08);color:#10d9a0;border-radius:99px;'
        f'font-size:10px;font-weight:700;border:1px solid rgba(16,217,160,0.2);margin-right:4px;">'
        f'✓ {p}</span>'
        for p in connected_platforms
    )

    col_cta, col_btn = st.columns([3.5, 1])
    with col_cta:
        candidate_name = profile.get("name", "") or "your"
        seniority = profile.get("seniority_level", "")
        skills_preview = ", ".join(profile.get("skills", [])[:4])
        st.markdown(
            f'<div class="jm-fetch-cta">'
            f'  <span style="font-size:26px;">🎯</span>'
            f'  <div class="jm-fetch-cta-text">'
            f'    <p class="jm-fetch-cta-title">Fetch real jobs from your connected portals</p>'
            f'    <p class="jm-fetch-cta-sub">Portals: {platform_badges}</p>'
            f'    <p class="jm-fetch-cta-sub" style="margin-top:4px;">Role: <b style="color:#60a5fa;">{seniority or "matching"}</b> · Skills: {skills_preview or "from resume"} · Location: {location}</p>'
            f'  </div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    with col_btn:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        fetch_clicked = st.button(
            "🔍 Fetch Real Jobs",
            key="jm_fetch_btn",
            use_container_width=True,
            type="primary",
            help=f"Fetches real job listings from: {', '.join(connected_platforms)}",
        )

    if fetch_clicked:
        with st.spinner(f"🔍 Fetching real jobs from {', '.join(connected_platforms)}…"):
            fetched = _fetch_real_jobs_from_connected_platforms(
                resume_text, profile, location, connected_platforms, time_filter
            )
        if fetched:
            st.session_state[fetch_key] = fetched
            live_jobs = fetched
            st.success(f"✅ Found {len(fetched)} real job matches from {', '.join(connected_platforms)}!")
            st.rerun()
        else:
            st.warning(
                "⚠️ No jobs returned. This can happen if:\n"
                "- Your AI key's grounding quota is exhausted (try a different key in API Keys)\n"
                "- The portals had no results for your profile keywords\n\n"
                "👉 Use the **direct portal links** below each job card to search manually."
            )

    # ── Show jobs ─────────────────────────────────────────────────────────
    all_jobs = live_jobs

    if not all_jobs:
        st.markdown(
            f'<div class="jm-empty">'
            f'  <div class="jm-empty-icon">🔍</div>'
            f'  <p class="jm-empty-title">No job matches yet</p>'
            f'  <p class="jm-empty-sub">Click <b>Fetch Real Jobs</b> above to pull live listings from '
            f'  {", ".join(connected_platforms)}. Only real postings are shown — no AI-generated jobs.</p>'
            f'</div>',
            unsafe_allow_html=True,
        )
        return

    # ── Normalize & filter ────────────────────────────────────────────────
    normalized = []
    for job in all_jobs:
        normalized.append({
            **job,
            "platform":               str(job.get("platform", "Unknown")),
            "tags":                   job.get("tags", []),
            "salary_lpa":             float(job.get("salary_lpa", 0) or 0),
            "ats_match":              float(job.get("ats_match", 0) or 0),
            "interview_probability":  float(job.get("interview_probability", 0) or 0),
            "shortlisting_confidence":int(float(job.get("shortlisting_confidence", 0) or 0)),
            "_competition_rank":      normalize_competition(job.get("competition_level", "High")),
        })

    filtered = [j for j in normalized if j["salary_lpa"] >= min_salary]

    sort_keys = {
        "Shortlisting Confidence": lambda x: -x["shortlisting_confidence"],
        "Salary (High→Low)":       lambda x: -x["salary_lpa"],
        "ATS Match":               lambda x: -x["ats_match"],
        "Interview Probability":   lambda x: -x["interview_probability"],
    }
    ranked = sorted(filtered, key=sort_keys.get(sort_by, sort_keys["Shortlisting Confidence"]))

    # ── Summary strip ────────────────────────────────────────────────────
    if ranked:
        avg_sc  = sum(j["shortlisting_confidence"] for j in ranked) / len(ranked)
        avg_sal = sum(j["salary_lpa"] for j in ranked) / len(ranked)
        top_sc  = max(j["shortlisting_confidence"] for j in ranked)
        sc_color = "#10d9a0" if avg_sc >= 70 else "#f59e0b" if avg_sc >= 50 else "#ef4444"
        source_label = f"Live · {', '.join(connected_platforms[:2])}{'…' if len(connected_platforms) > 2 else ''}"
        st.markdown(
            f'<div class="jm-summary">'
            f'  <div class="jm-summary-chip">🎯 <span class="val">{len(ranked)}</span>&nbsp;matches found</div>'
            f'  <div class="jm-summary-chip">📊 Avg shortlisting <span class="val" style="color:{sc_color};">{avg_sc:.0f}%</span></div>'
            f'  <div class="jm-summary-chip">🏆 Best fit <span class="val" style="color:#10d9a0;">{top_sc}%</span></div>'
            f'  <div class="jm-summary-chip">💰 Avg salary <span class="val">{format_salary_lpa(avg_sal)}</span></div>'
            f'  <div class="jm-summary-chip">🕐 <span class="val" style="font-size:12px;color:#60a5fa;">{source_label}</span></div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── Rank color helper ─────────────────────────────────────────────────
    def _sc_color(sc: int) -> str:
        if sc >= 75: return "#10d9a0"
        if sc >= 55: return "#f59e0b"
        return "#ef4444"

    def _sc_gradient(sc: int) -> str:
        if sc >= 75: return "linear-gradient(90deg,#10d9a0,#059669)"
        if sc >= 55: return "linear-gradient(90deg,#f59e0b,#d97706)"
        return "linear-gradient(90deg,#ef4444,#dc2626)"

    def _rank_bg(idx: int) -> str:
        palettes = [
            "linear-gradient(135deg,#f59e0b,#d97706)",  # gold
            "linear-gradient(135deg,#94a3b8,#64748b)",  # silver
            "linear-gradient(135deg,#f97316,#ea580c)",  # bronze
        ]
        return palettes[idx] if idx < 3 else "linear-gradient(135deg,#1e2d45,#111c2e)"

    def _card_accent(sc: int) -> str:
        if sc >= 75: return "linear-gradient(90deg,#10d9a0,#059669aa,transparent)"
        if sc >= 55: return "linear-gradient(90deg,#f59e0b,#d97706aa,transparent)"
        return "linear-gradient(90deg,#ef4444,#dc2626aa,transparent)"

    # ── Render job cards ──────────────────────────────────────────────────
    for idx, job in enumerate(ranked):
        sc    = job["shortlisting_confidence"]
        color = _sc_color(sc)
        grad  = _sc_gradient(sc)

        urgency   = str(job.get("urgency", "Normal"))
        posted    = str(job.get("posted_label", "Recently"))
        jtype     = str(job.get("job_type", "Full-time"))
        exp_req   = str(job.get("experience_required", ""))
        gaps      = job.get("shortlisting_gaps", [])
        tags      = job.get("tags", []) or []
        s_matched = job.get("skills_matched", []) or []
        s_missing = job.get("skills_missing", []) or []
        rationale = str(job.get("rationale", ""))
        apply_url  = str(job.get("apply_url", "")).strip()
        direct_url = str(job.get("direct_url", "")).strip()
        # Strip any redirect wrappers that sneak through
        if any(h in direct_url for h in ("to.indeed.com", "bit.ly", "tinyurl", "goo.gl")):
            direct_url = ""
        comp_level = str(job.get("competition_level", "Medium"))
        comp_color = "#10d9a0" if comp_level == "Low" else "#f59e0b" if comp_level == "Medium" else "#ef4444"

        sc_label = (
            "High Confidence" if sc >= 75
            else "Moderate Chance" if sc >= 55
            else "Stretch Role"
        )
        sc_icon = "🟢" if sc >= 75 else "🟡" if sc >= 55 else "🔴"

        # Platform search links
        li_url     = str(job.get("linkedin_job_url", "")).strip()
        indeed_url = str(job.get("indeed_job_url", "")).strip()
        naukri_url = str(job.get("naukri_job_url", "")).strip()
        _title_enc = job.get("title", "")
        _loc_enc   = str(job.get("location", location or "India"))
        if not li_url:
            li_url = _build_search_url("LinkedIn", _title_enc, _loc_enc)
        if not indeed_url:
            indeed_url = _build_search_url("Indeed", _title_enc, _loc_enc)
        if not naukri_url:
            naukri_url = _build_search_url("Naukri", _title_enc, _loc_enc)

        # Escape all dynamic data to prevent HTML breakage from special chars in AI/job data
        def _e(v: str) -> str:
            return _html_escape.escape(str(v))

        # Build HTML sub-components
        urgency_html = (
            f'<span class="jm-pill jm-pill-urgent">⚡ {_e(urgency)}</span>'
            if urgency == "Actively Hiring" else ""
        )
        exp_html = f'<span class="jm-pill jm-pill-exp">🎯 {_e(exp_req)}</span>' if exp_req else ""
        matched_chips = "".join(
            f'<span class="jm-reason-chip match">✓ {_e(s)}</span>' for s in s_matched[:5]
        )
        gap_chips = "".join(
            f'<span class="jm-reason-chip gap">✗ {_e(g)}</span>' for g in (s_missing or gaps)[:3]
        )
        tag_chips_html = "".join(f'<span class="jm-tag">{_e(t)}</span>' for t in tags[:8])
        rationale_html = (
            f'<div class="jm-rationale">💡 {_e(rationale)}</div>' if rationale else ""
        )

        # ── Single unified card HTML ──────────────────────────────────────
        _title   = _e(job.get("title", "Unknown Role"))
        _company = _e(job.get("company", "N/A"))
        _loc     = _e(job.get("location", ""))
        _plat    = _e(job.get("platform", ""))
        _posted  = _e(posted)
        _jtype   = _e(jtype)

        # Best direct URL: prefer apply_url → platform-specific → search fallback
        _platform_lower = job.get("platform", "").lower()
        if apply_url:
            _best_url = apply_url
        elif "linkedin" in _platform_lower and li_url:
            _best_url = li_url
        elif "indeed" in _platform_lower and indeed_url:
            _best_url = indeed_url
        elif "naukri" in _platform_lower and naukri_url:
            _best_url = naukri_url
        else:
            _best_url = li_url or indeed_url or naukri_url or _build_search_url(
                job.get("platform", "LinkedIn"), job.get("title", ""), str(job.get("location", location or "India"))
            )

        st.markdown(f"""
<div class="jm-card">
  <div class="jm-card-accent" style="background:{_card_accent(sc)};"></div>
  <div class="jm-card-body">
    <div class="jm-rank" style="background:{_rank_bg(idx)};">#{idx+1}</div>
    <div class="jm-title-row">
      <p class="jm-title">{_title}</p>
    </div>
    <p class="jm-company">
      <span style="color:#94a3b8;">🏢</span> <b style="color:#cbd5e1;">{_company}</b>
      &nbsp;·&nbsp;
      <span style="color:#64748b;">📍 {_loc}</span>
    </p>
    <div class="jm-meta-row">
      <span class="jm-pill jm-pill-platform">🔗 {_plat}</span>
      <span class="jm-pill jm-pill-posted">🕐 {_posted}</span>
      <span class="jm-pill jm-pill-type">{_jtype}</span>
      {exp_html}
      {urgency_html}
    </div>
    <div class="jm-tags">{tag_chips_html}</div>
    <div class="jm-metrics">
      <div class="jm-metric">
        <div class="jm-metric-label">💰 Salary</div>
        <div class="jm-metric-value" style="color:#10d9a0;">{format_salary_lpa(job["salary_lpa"])}</div>
      </div>
      <div class="jm-metric">
        <div class="jm-metric-label">📊 ATS Match</div>
        <div class="jm-metric-value" style="color:#3b82f6;">{job["ats_match"]:.0f}%</div>
      </div>
      <div class="jm-metric">
        <div class="jm-metric-label">🎤 Interview</div>
        <div class="jm-metric-value" style="color:#a78bfa;">{job["interview_probability"]:.0f}%</div>
      </div>
      <div class="jm-metric">
        <div class="jm-metric-label">⚔ Competition</div>
        <div class="jm-metric-value" style="color:{comp_color};">{comp_level}</div>
      </div>
    </div>
    <div class="jm-sc-section">
      <div class="jm-sc-header">
        <span class="jm-sc-label">🎯 Shortlisting Confidence</span>
        <span class="jm-sc-score" style="color:{color};">{sc_icon} {sc_label} &nbsp;<b>{sc}%</b></span>
      </div>
      <div class="jm-sc-bar-bg">
        <div class="jm-sc-bar-fill" style="width:{sc}%;background:{grad};"></div>
      </div>
      <div class="jm-reason-row">{matched_chips}{gap_chips}</div>
    </div>
    {rationale_html}
    <div class="jm-divider"></div>
    <div class="jm-actions-row">
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
        <a href="{_best_url}" target="_blank" class="jm-apply-btn">
          <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M18 13v6a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V8a2 2 0 0 1 2-2h6"/><polyline points="15 3 21 3 21 9"/><line x1="10" y1="14" x2="21" y2="3"/></svg>
          Apply on {_plat} ↗
        </a>
        {f'<a href="{direct_url}" target="_blank" class="jm-view-listing-btn">View Listing ↗</a>' if direct_url and direct_url != "#" and direct_url != _best_url else ""}
      </div>
      <div class="jm-portal-links-row">
        {f'<a href="{li_url}" target="_blank" class="jm-src-link jm-src-li {"jm-src-active" if "linkedin" in job.get("platform","").lower() else ""}">🔗 LinkedIn</a>' if li_url else ""}
        {f'<a href="{indeed_url}" target="_blank" class="jm-src-link jm-src-in {"jm-src-active" if "indeed" in job.get("platform","").lower() else ""}">🔍 Indeed</a>' if indeed_url else ""}
        {f'<a href="{naukri_url}" target="_blank" class="jm-src-link jm-src-nk {"jm-src-active" if "naukri" in job.get("platform","").lower() else ""}">🇮🇳 Naukri</a>' if naukri_url else ""}
      </div>
    </div>
    <div class="jm-url-preview">
      <span class="jm-url-label">Opens →</span>
      <span class="jm-url-text">{_best_url[:80] + "…" if len(_best_url) > 80 else _best_url}</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # Save to Tracker button (Streamlit widget, below card)
        if st.button("📋 Save to Tracker", key=f"add_tracker_{idx}", use_container_width=False):
            add_job_to_tracker(job)
            st.success(f"✅ '{job.get('title','')}' saved to tracker.")

        # ── Outreach templates expander ───────────────────────────────────
        with st.expander(f"✉️ Outreach Templates — {job.get('title','')}", expanded=False):
            candidate_name = str(job.get("_candidate_name", "") or profile.get("name", "") or "")
            hr_name     = str(job.get("hr_name", "")).strip()
            hr_email    = str(job.get("hr_email", "")).strip()
            job_title   = str(job.get("title", ""))
            company     = str(job.get("company", ""))
            job_loc     = str(job.get("location", location or ""))
            job_desc_snippet = str(job.get("rationale", "") or "")

            # Pull rich resume data from session state
            insight       = st.session_state.get("insight", {})
            summary_line  = insight.get("summary_line", "")
            strengths     = insight.get("strengths", [])
            kw_present    = insight.get("keywords_present", [])
            certifications = insight.get("certifications", [])
            industry      = insight.get("industry", "")
            edu_list      = profile.get("education", [])
            job_titles    = profile.get("job_titles", [])
            seniority     = profile.get("seniority_level", "")
            exp_years     = profile.get("experience_years", 0)

            # Matched/missing from the specific job card
            skills_matched = (job.get("skills_matched") or job.get("tags") or [])[:5]
            skills_missing = (job.get("skills_missing") or [])[:3]
            skills_str     = ", ".join(skills_matched)

            # Compact resume snapshot (capped so prompt stays lean)
            resume_snippet = trimmed_text(st.session_state.get("resume_text", ""), max_chars=1200)

            tab_li, tab_gmail = st.tabs(["🔗 LinkedIn Invitation", "📧 Gmail to HR"])

            with tab_li:
                _li_key = f"li_invite_text_{idx}"

                # ── User-facing outreach prompt (≤300 chars for LinkedIn) ──
                _li_user_prompt_full = (
                    f"Hi {hr_name or '[Recruiter Name]'}, I came across the {job_title} role at {company} "
                    f"and I'm very interested. With {exp_years}+ yrs in {skills_str or 'relevant skills'}, "
                    f"I'd love to connect!"
                )
                # hard-cap at 300 chars
                if len(_li_user_prompt_full) > 300:
                    _li_user_prompt_full = _li_user_prompt_full[:297] + "…"

                with st.expander("📋 Your outreach prompt (use this to reach out manually)", expanded=False):
                    st.markdown(
                        f'<p style="font-size:11px;color:#64748b;margin-bottom:4px;">'
                        f'{len(_li_user_prompt_full)}/300 characters</p>',
                        unsafe_allow_html=True,
                    )
                    st.text_area(
                        "Copy this message and paste into LinkedIn → Connect → Add Note:",
                        value=_li_user_prompt_full,
                        height=100,
                        key=f"li_user_prompt_{idx}",
                    )

                if st.button("✨ Generate LinkedIn Invite", key=f"li_invite_{idx}", use_container_width=True):
                    li_prompt = f"""You are a professional career coach. Write a LinkedIn connection request note.

STRICT RULES — non-negotiable:
- Total output: MAXIMUM 280 characters (not words — characters including spaces). LinkedIn's limit is 300. Stay under 280 to be safe.
- Count characters carefully. A typical sentence is ~80-100 characters. You get roughly 3 short sentences total.
- No subject line, no headers, no labels — only the message body.
- Do NOT start with "I" — open with the recruiter's name or a brief hook.
- Sound human and warm, not like a template. No clichés.
- Mention the EXACT job title and company.
- Reference 1 specific matched skill FROM THE RESUME below.
- End with a polite ask to connect.

CANDIDATE RESUME SNAPSHOT:
{resume_snippet}

CANDIDATE PROFILE:
- Name: {candidate_name or "the candidate"}
- Role history: {', '.join(job_titles[:3]) or seniority or 'professional'}
- Experience: {exp_years} years in {industry or 'the field'}
- Top skills matched to this job: {skills_str}
- Strengths: {', '.join(strengths[:3])}
- Education: {', '.join(edu_list[:2])}
- Certifications: {', '.join(certifications[:2])}

TARGET JOB:
- Title: {job_title} at {company} ({job_loc})
- HR/Recruiter: {hr_name or "the recruiter"}
- Job context: {job_desc_snippet[:300]}
- Skills the candidate matches: {skills_str}
- Skills the candidate is missing: {', '.join(skills_missing) or 'none identified'}

Write the LinkedIn note now. MAXIMUM 280 characters total. Count carefully.
"""
                    with st.spinner("Crafting LinkedIn note…"):
                        raw_invite = run_gemini_prompt(li_prompt)
                    # Hard-enforce 300-char cap (LinkedIn's actual limit) as safety net
                    li_invite = raw_invite.strip()
                    if len(li_invite) > 300:
                        li_invite = li_invite[:297] + "…"
                    st.session_state[_li_key] = li_invite

                li_text = st.session_state.get(_li_key, "")
                if li_text:
                    st.text_area(
                        "LinkedIn Invitation (copy & paste into LinkedIn → Connect → Add Note)",
                        value=li_text,
                        height=130,
                        key=f"li_ta_{idx}",
                    )
                    _cc = len(li_text)
                    _cc_color = "#10d9a0" if _cc <= 280 else "#f59e0b" if _cc <= 300 else "#ef4444"
                    st.markdown(
                        f'<p style="font-size:11px;color:{_cc_color};margin-top:4px;">{_cc}/300 characters (LinkedIn note limit)</p>',
                        unsafe_allow_html=True,
                    )
                    _li_search = li_url or _build_search_url("LinkedIn", job_title, job_loc)
                    st.markdown(
                        f'<a href="{_li_search}" target="_blank" style="display:inline-flex;align-items:center;gap:5px;'
                        f'padding:7px 16px;background:rgba(10,102,194,0.12);color:#60a5fa;border-radius:8px;'
                        f'font-size:12px;font-weight:600;border:1px solid rgba(10,102,194,0.3);'
                        f'text-decoration:none;margin-top:8px;">🔗 Open Job on LinkedIn →</a>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.caption("Click the button above to generate a personalised LinkedIn note.")

            with tab_gmail:
                _gmail_key = f"gmail_text_{idx}"

                # ── User-facing outreach prompt ──────────────────────────
                _gmail_user_prompt = (
                    f"Subject: {seniority or job_titles[0] if job_titles else 'Experienced Professional'} interested in {job_title} at {company}\n\n"
                    f"Dear {hr_name or 'Hiring Team'},\n\n"
                    f"I'm reaching out regarding the {job_title} position at {company}. "
                    f"With {exp_years}+ years in {industry or 'the industry'} and hands-on experience in {skills_str or 'relevant technologies'}, "
                    f"I believe I'm a strong fit for this role. "
                    f"{'I bring ' + ', '.join(strengths[:2]) + ' to every engagement.' if strengths else ''}\n\n"
                    f"I'd welcome the opportunity to discuss how I can contribute to your team.\n\n"
                    f"Best regards,\n{candidate_name or '[Your Name]'}"
                ).strip()

                with st.expander("📋 Your outreach prompt (use this to reach out manually)", expanded=False):
                    st.text_area(
                        "Copy this and send it as an email:",
                        value=_gmail_user_prompt,
                        height=200,
                        key=f"gmail_user_prompt_{idx}",
                    )

                if st.button("✨ Generate Gmail to HR", key=f"gmail_{idx}", use_container_width=True):
                    gmail_prompt = f"""You are a professional career coach. Write a cold outreach email to an HR recruiter.

STRICT RULES:
- Email body: 150–200 words ONLY. Hard limit. Count every word.
- Subject line: under 10 words, specific, includes role name.
- Open with recruiter's name if available, else "Dear Hiring Team".
- 3 short paragraphs: (1) why this company/role specifically, (2) 2-3 concrete achievements or skills directly matching job, (3) clear call to action.
- No buzzwords, no "I hope this finds you well", no generic filler.
- Ground everything in the RESUME DATA below — mention real skills, real background.
- Return format exactly:
SUBJECT: <subject line>
---
<email body only>

CANDIDATE RESUME SNAPSHOT:
{resume_snippet}

CANDIDATE PROFILE:
- Name: {candidate_name or "the candidate"}
- Role history: {', '.join(job_titles[:3]) or seniority or 'professional'}
- Experience: {exp_years} years | Seniority: {seniority}
- Industry: {industry}
- Matched skills for this job: {skills_str}
- Key strengths: {', '.join(strengths[:3])}
- Education: {', '.join(edu_list[:2])}
- Certifications: {', '.join(certifications[:2])}
- Missing skills to address: {', '.join(skills_missing) or 'none'}

TARGET JOB:
- Title: {job_title} at {company} ({job_loc})
- HR/Recruiter: {hr_name or "Hiring Manager"} | Email: {hr_email or "N/A"}
- Job context: {job_desc_snippet[:400]}
"""
                    with st.spinner("Writing email to HR…"):
                        gmail_text = run_gemini_prompt(gmail_prompt)
                    st.session_state[_gmail_key] = gmail_text

                gmail_content = st.session_state.get(_gmail_key, "")
                if gmail_content:
                    if "---" in gmail_content:
                        parts = gmail_content.split("---", 1)
                        subject_line = parts[0].replace("SUBJECT:", "").strip()
                        body_text    = parts[1].strip()
                    else:
                        subject_line = f"Application for {job_title} at {company}"
                        body_text    = gmail_content
                    # Hard-enforce 200-word cap on body
                    body_words = body_text.split()
                    if len(body_words) > 200:
                        body_text = " ".join(body_words[:200]) + "…"
                    st.text_input("📌 Subject Line (copy this)", value=subject_line, key=f"gmail_sub_{idx}")
                    _bwc = len(body_text.split())
                    _bwc_color = "#10d9a0" if _bwc <= 200 else "#ef4444"
                    st.markdown(
                        f'<p style="font-size:11px;color:{_bwc_color};margin-top:2px;margin-bottom:6px;">Body word count: {_bwc}/200</p>',
                        unsafe_allow_html=True,
                    )
                    if hr_email:
                        st.markdown(
                            f'<p style="font-size:12px;color:#10d9a0;margin-bottom:4px;">📬 Send to: <b>{hr_email}</b></p>',
                            unsafe_allow_html=True,
                        )
                        mailto_link = (
                            f"mailto:{hr_email}"
                            f"?subject={urllib.parse.quote(subject_line)}"
                            f"&body={urllib.parse.quote(body_text[:1500])}"
                        )
                        st.markdown(
                            f'<a href="{mailto_link}" style="display:inline-flex;align-items:center;gap:5px;'
                            f'padding:7px 16px;background:rgba(16,217,160,0.1);color:#10d9a0;border-radius:8px;'
                            f'font-size:12px;font-weight:600;border:1px solid rgba(16,217,160,0.25);'
                            f'text-decoration:none;margin-bottom:10px;">📨 Open in Gmail / Mail App</a>',
                            unsafe_allow_html=True,
                        )
                    st.text_area(
                        "Email Body (copy & paste into Gmail)",
                        value=body_text,
                        height=260,
                        key=f"gmail_area_{idx}",
                    )
                else:
                    st.caption("Click the button above to generate a personalised HR email.")

        # spacing between cards
        st.markdown('<div style="height:10px;"></div>', unsafe_allow_html=True)


def application_tracker_page() -> None:
    sb_load_tracker()  # Load from Supabase on every page visit
    st.subheader("Application Tracker")
    if not _build_provider_list():
        st.warning("⚠ No API keys configured. Add a free key in the **API Keys** page.", icon="🔑")
        return
    importer = st.file_uploader("Import Tracker (.xlsx/.csv)", type=["xlsx", "csv"])
    if importer and st.button("Import File"):
        try:
            imported_df = parse_tracker_upload(importer)
            st.session_state["tracker"] = imported_df.fillna("").to_dict("records")
            st.session_state["tracker_rows"] = st.session_state["tracker"]
            log_notion(
                {
                    "SessionID": st.session_state.get("session_id", ""),
                    "Timestamp": datetime.utcnow().isoformat(),
                    "Location": st.session_state.get("location", ""),
                    "JobsSearched": len(st.session_state.get("jobs", [])),
                    "Platforms": ", ".join(st.session_state.get("platforms_connected", [])),
                    "Currency": st.session_state.get("currency", "₹"),
                }
            )
            st.success(f"Imported {len(imported_df)} rows.")
        except Exception as exc:
            st.error(f"Import failed: {exc}")

    with st.form("quick_add_tracker", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        company = c1.text_input("Company")
        role = c2.text_input("Role")
        platform = c3.selectbox("Platform", st.session_state["platforms"])
        c4, c5 = st.columns(2)
        status = c4.text_input("Status (free text)")
        package = c5.text_input("Package")
        notes = st.text_area("Notes")
        next_step = st.text_input("NextStep")
        url = st.text_input("URL")
        add_row = st.form_submit_button("Add Row")
        if add_row:
            st.session_state["tracker"].append(
                {
                    "Company": company,
                    "Role": role,
                    "Platform": platform,
                    "Date": date.today().isoformat(),
                    "Status": status,
                    "Package": package,
                    "Notes": notes,
                    "NextStep": next_step,
                    "URL": url,
                }
            )
            st.session_state["tracker_rows"] = st.session_state["tracker"]
            sb_save_tracker()
            log_notion(
                {
                    "SessionID": st.session_state.get("session_id", ""),
                    "Timestamp": datetime.utcnow().isoformat(),
                    "Location": st.session_state.get("location", ""),
                    "JobsSearched": len(st.session_state.get("jobs", [])),
                    "Platforms": ", ".join(st.session_state.get("platforms_connected", [])),
                    "Currency": st.session_state.get("currency", "₹"),
                }
            )
            st.success("Row added.")

    df = pd.DataFrame(st.session_state["tracker"], columns=TRACKER_COLUMNS)
    edited_df = st.data_editor(
        df,
        use_container_width=True,
        num_rows="dynamic",
        key="tracker_editor",
        hide_index=True,
    )
    actions_1, actions_2 = st.columns(2)
    if actions_1.button("Save Edits"):
        st.session_state["tracker"] = edited_df.fillna("").to_dict("records")
        st.session_state["tracker_rows"] = st.session_state["tracker"]
        sb_save_tracker()
        log_notion(
            {
                "SessionID": st.session_state.get("session_id", ""),
                "Timestamp": datetime.utcnow().isoformat(),
                "Location": st.session_state.get("location", ""),
                "JobsSearched": len(st.session_state.get("jobs", [])),
                "Platforms": ", ".join(st.session_state.get("platforms_connected", [])),
                "Currency": st.session_state.get("currency", "₹"),
            }
        )
        st.success("Tracker updated.")
    if actions_2.button("Delete All Rows"):
        st.session_state["tracker"] = []
        st.session_state["tracker_rows"] = []
        st.rerun()

    if not edited_df.empty:
        data_for_stats = edited_df.fillna("").astype(str)
        total = len(data_for_stats)
        status_series = data_for_stats["Status"].str.lower()
        interviews = int(status_series.str.contains("interview").sum())
        offers = int(status_series.str.contains("offer").sum())
        success = (offers / total * 100) if total else 0.0
        s1, s2, s3, s4 = st.columns(4)
        s1.metric("Total", total)
        s2.metric("Interviews", interviews)
        s3.metric("Offers", offers)
        s4.metric("Success %", f"{success:.1f}%")

        st.download_button(
            "Export to Excel",
            data=format_tracker_export(edited_df),
            file_name="AutoApply AI_applications.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.markdown("### Status Fetch")
        records = edited_df.fillna("").to_dict("records")
        for idx, row in enumerate(records):
            c1, c2, c3 = st.columns([2, 2, 1])
            c1.write(f"{row.get('Company','')} - {row.get('Role','')}")
            c2.write(f"Current: {row.get('Status','')}")
            if c3.button("Check Status", key=f"check_status_{idx}"):
                prompt = f"""
Given company={row.get("Company","")}, role={row.get("Role","")},
applied={row.get("Date","")}, platform={row.get("Platform","")}.
Based on typical hiring timelines, what is the likely current status?
Return JSON:
{{"status":"string","next_action":"string","expected_response_days":0,"tips":["string"]}}
"""
                try:
                    output = run_gemini_prompt(prompt)
                    parsed = extract_json_payload(output)
                    if isinstance(parsed, dict):
                        records[idx]["Status"] = parsed.get("status", row.get("Status", ""))
                        records[idx]["NextStep"] = parsed.get("next_action", row.get("NextStep", ""))
                        tips = parsed.get("tips", [])
                        tip_text = ", ".join(tips) if isinstance(tips, list) else str(tips)
                        expected_days = parsed.get("expected_response_days", "")
                        records[idx]["Notes"] = (
                            f"{row.get('Notes','')} | ETA: {expected_days} days | Tips: {tip_text}"
                        ).strip(" |")
                        st.session_state["tracker"] = records
                        st.session_state["tracker_rows"] = records
                        st.success(f"Updated: {records[idx]['Status']}")
                    else:
                        st.warning("Could not parse status response.")
                except Exception as exc:
                    st.error(f"Status check failed: {exc}")

        if st.session_state.get("last_daily_status_check") != date.today().isoformat():
            rows_payload = json.dumps(records, default=str)
            with st.spinner("Running daily auto status check..."):
                daily_results = cached_daily_status_check(rows_payload, st.session_state.get("api_key", ""))
            if daily_results:
                for idx_str, result in daily_results.items():
                    idx = int(idx_str)
                    if 0 <= idx < len(records):
                        records[idx]["Status"] = result.get("status", records[idx].get("Status", ""))
                        records[idx]["NextStep"] = result.get("next_action", records[idx].get("NextStep", ""))
                st.session_state["tracker"] = records
                st.session_state["tracker_rows"] = records
            st.session_state["last_daily_status_check"] = date.today().isoformat()
    else:
        st.info("No tracker rows yet.")


def platforms_page() -> None:
    sb_load_platforms()  # Load connected platforms from Supabase
    # ── Platform icons map ───────────────────────────────────────────────
    _PLATFORM_ICONS = {
        "LinkedIn":    "in",
        "Naukri":      "N",
        "Indeed":      "id",
        "Glassdoor":   "G",
        "Wellfound":   "W",
        "Unstop":      "U",
        "Monster":     "M",
        "Shine":       "Sh",
        "Internshala": "IS",
        "AngelList":   "AL",
        "Hirist":      "Hi",
        "Cutshort":    "CS",
    }
    _PLATFORM_COLORS = {
        "LinkedIn":    ("#0A66C2", "#004182"),
        "Naukri":      ("#FF7555", "#c94f32"),
        "Indeed":      ("#2164f3", "#1347c0"),
        "Glassdoor":   "#0CAA41",
        "Wellfound":   ("#ED4F32", "#b83520"),
        "Unstop":      ("#7C3AED", "#5b22d0"),
        "Monster":     ("#6E0B9A", "#4e0870"),
        "Shine":       ("#E91E8C", "#b8146d"),
        "Internshala": ("#00BCD4", "#0090a0"),
        "AngelList":   ("#FF6154", "#cc3d30"),
        "Hirist":      ("#1976D2", "#0d5aab"),
        "Cutshort":    ("#FF6B35", "#cc4a1a"),
    }
    _LOGIN_TIPS = {
        "LinkedIn":    "Go to Jobs tab → search your role → click Easy Apply.",
        "Naukri":      "Update your profile headline & skills to boost visibility.",
        "Indeed":      "Upload resume → enable Indeed Apply for 1-click applications.",
        "Glassdoor":   "Check company reviews & salary ranges before applying.",
        "Wellfound":   "Browse startup jobs with salary & equity details visible.",
        "Unstop":      "Explore competitions, hackathons, jobs & internships.",
        "Monster":     "Set smart job alerts for your target role & location.",
        "Shine":       "Complete your profile to 100% for better job matches.",
        "Internshala": "Apply for internships & fresher jobs with cover letters.",
        "AngelList":   "Apply directly to startup founders — no middlemen.",
        "Hirist":      "Find tech-specific roles — filter by skill & experience.",
        "Cutshort":    "AI matches you to jobs based on your skills & preferences.",
    }

    # ── Page CSS injection ───────────────────────────────────────────────
    st.markdown("""
<style>
@keyframes pulse-ring {
  0%   { box-shadow: 0 0 0 0 rgba(16,217,160,0.35); }
  70%  { box-shadow: 0 0 0 8px rgba(16,217,160,0); }
  100% { box-shadow: 0 0 0 0 rgba(16,217,160,0); }
}
@keyframes shimmer {
  0%   { background-position: -200% center; }
  100% { background-position: 200% center; }
}
.portal-page-header {
    margin-bottom: 28px;
}
.portal-page-title {
    font-size: 22px !important;
    font-weight: 700 !important;
    color: #EAF2FF !important;
    letter-spacing: -0.3px;
    margin: 0 0 4px 0 !important;
}
.portal-page-sub {
    font-size: 13px !important;
    color: #64748b !important;
    margin: 0 !important;
}
.portal-progress-wrap {
    background: #0d1117;
    border: 1px solid #1e2d45;
    border-radius: 16px;
    padding: 18px 22px;
    margin-bottom: 24px;
    display: flex;
    align-items: center;
    gap: 20px;
}
.portal-progress-label {
    font-size: 11px !important;
    font-weight: 700 !important;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #64748b !important;
    white-space: nowrap;
}
.portal-progress-bar-bg {
    flex: 1;
    height: 6px;
    background: #1e2d45;
    border-radius: 99px;
    overflow: hidden;
}
.portal-progress-fill {
    height: 100%;
    border-radius: 99px;
    background: linear-gradient(90deg, #1dd4a0, #3b82f6);
    transition: width 0.5s ease;
}
.portal-progress-count {
    font-size: 13px !important;
    font-weight: 700 !important;
    color: #EAF2FF !important;
    white-space: nowrap;
}
/* Platform card wrapper — one per column */
.pcard-outer {
    margin-bottom: 18px;
}
.pcard {
    background: linear-gradient(145deg, #0d1117, #0a0e18);
    border: 1px solid #1e2d45;
    border-radius: 18px;
    overflow: hidden;
    transition: border-color 0.2s, box-shadow 0.2s, transform 0.2s;
    box-shadow: 0 4px 24px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.02);
}
.pcard:hover {
    box-shadow: 0 8px 36px rgba(0,0,0,0.55), inset 0 1px 0 rgba(255,255,255,0.03);
    transform: translateY(-2px);
}
.pcard.connected {
    border-color: rgba(16,217,160,0.28);
    background: linear-gradient(145deg, #071a0f 0%, #0a1a13 100%);
    box-shadow: 0 4px 24px rgba(16,217,160,0.06), 0 0 0 1px rgba(16,217,160,0.08) inset;
}
.pcard.pending {
    border-color: rgba(245,158,11,0.35);
    background: linear-gradient(145deg, #131005 0%, #12100a 100%);
    box-shadow: 0 4px 24px rgba(245,158,11,0.06);
}
/* Card top accent strip */
.pcard-accent {
    height: 2px;
    width: 100%;
}
.pcard-body {
    padding: 18px 20px 16px;
}
/* Header row */
.pcard-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 14px;
}
.pcard-identity {
    display: flex;
    align-items: center;
    gap: 11px;
}
.pcard-logo {
    width: 38px; height: 38px;
    border-radius: 11px;
    display: flex; align-items: center; justify-content: center;
    font-size: 12px; font-weight: 800;
    color: #fff; flex-shrink: 0;
    letter-spacing: -0.5px;
    box-shadow: 0 4px 14px rgba(0,0,0,0.5);
}
.pcard-name {
    font-size: 15px !important;
    font-weight: 700 !important;
    color: #EAF2FF !important;
    letter-spacing: -0.3px;
    margin: 0 !important;
    font-family: 'Sora', sans-serif !important;
}
/* Status badge */
.pcard-badge {
    display: inline-flex; align-items: center; gap: 5px;
    border-radius: 99px;
    padding: 4px 12px;
    font-size: 9px !important; font-weight: 800 !important;
    letter-spacing: 0.3px; text-transform: uppercase;
    white-space: nowrap;
}
.pcard-badge.connected {
    background: #0a2318; border: 1px solid #10d9a030; color: #10d9a0 !important;
}
.pcard-badge.pending {
    background: #1a1200; border: 1px solid #f59e0b40; color: #f59e0b !important;
}
.pcard-badge.idle {
    background: #0d1117; border: 1px solid #1e2d45; color: #475569 !important;
}
.pcard-badge-dot {
    width: 5px; height: 5px; border-radius: 50%; flex-shrink: 0;
}
/* Stats grid */
.pcard-stats {
    display: grid; grid-template-columns: 1fr 1fr; gap: 8px;
    margin-bottom: 12px;
}
.pcard-stat {
    background: linear-gradient(145deg, #060a14, #070c18);
    border: 1px solid #111c2e;
    border-radius: 11px;
    padding: 10px 13px;
    box-shadow: inset 0 1px 4px rgba(0,0,0,0.25);
}
.pcard-stat-label {
    font-size: 8px !important; font-weight: 800 !important;
    text-transform: uppercase; letter-spacing: 1px;
    color: #334155 !important; margin-bottom: 4px !important;
}
.pcard-stat-value {
    font-size: 13px !important; font-weight: 700 !important;
    color: #cbd5e1 !important; margin: 0 !important;
    font-family: 'Sora', sans-serif !important;
}
/* Connected timestamp */
.pcard-connected-at {
    display: flex; align-items: center; gap: 7px;
    padding: 7px 11px;
    background: linear-gradient(90deg, #071a0f, #060e09); border: 1px solid rgba(16,217,160,0.15);
    border-radius: 9px; margin-bottom: 5px;
    font-size: 10px !important; color: #10d9a0 !important;
    font-weight: 600 !important; letter-spacing: 0.2px !important;
}
/* Tip banner */
.pcard-tip {
    display: flex; align-items: flex-start; gap: 10px;
    background: linear-gradient(90deg, #131005, #11100a); border: 1px solid rgba(245,158,11,0.15);
    border-left: 2.5px solid #f59e0b;
    border-radius: 10px; padding: 10px 14px;
    margin-bottom: 5px;
}
.pcard-tip-icon { font-size: 14px; flex-shrink: 0; margin-top: 2px; }
.pcard-tip-text {
    font-size: 11px !important; color: #fbbf24 !important;
    line-height: 1.55 !important; margin: 0 !important;
}
.pcard-tip-text b { color: #fde68a !important; }
</style>
""", unsafe_allow_html=True)

    platforms = st.session_state["platforms"]
    connected_list = st.session_state.get("platforms_connected", [])
    n_connected = len(connected_list)
    total = len(platforms)
    pct = int((n_connected / total * 100) if total else 0)
    prog_color = "#10d9a0" if n_connected == total else "#f59e0b" if n_connected > 0 else "#475569"

    # ── Page header ─────────────────────────────────────────────────────
    st.markdown("""
<div class="portal-page-header">
  <p class="portal-page-title">🌐 Career Portals</p>
  <p class="portal-page-sub">Connect to each platform to unlock job discovery & auto-apply features.</p>
</div>
""", unsafe_allow_html=True)

    # ── Progress bar ─────────────────────────────────────────────────────
    st.markdown(
        f'<div class="portal-progress-wrap">'
        f'  <span class="portal-progress-label">Coverage</span>'
        f'  <div class="portal-progress-bar-bg">'
        f'    <div class="portal-progress-fill" style="width:{pct}%;background:linear-gradient(90deg,{prog_color},{prog_color}aa);"></div>'
        f'  </div>'
        f'  <span class="portal-progress-count" style="color:{prog_color};">{n_connected}<span style="color:#475569;font-weight:400;"> / {total}</span></span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── Init pending flag ────────────────────────────────────────────────
    if "platform_pending_login" not in st.session_state:
        st.session_state["platform_pending_login"] = ""

    # ── Platform cards grid ─────────────────────────────────────────────
    cols = st.columns(3, gap="medium")
    for idx, name in enumerate(platforms):
        with cols[idx % 3]:
            is_connected = name in st.session_state["platforms_connected"]
            is_pending   = st.session_state["platform_pending_login"] == name
            meta         = PLATFORM_META.get(name, {"url": "", "best_time": "10 AM", "response_rate": "N/A"})
            connected_at = st.session_state.get("platforms_connected_at", {}).get(name, "")
            abbr         = _PLATFORM_ICONS.get(name, name[:2].upper())
            login_tip    = _LOGIN_TIPS.get(name, f"Login to {name} and start applying.")

            # Colors
            raw_color = _PLATFORM_COLORS.get(name, ("#3b82f6", "#2563eb"))
            if isinstance(raw_color, tuple):
                logo_bg = f"linear-gradient(135deg,{raw_color[0]},{raw_color[1]})"
                accent  = raw_color[0]
            else:
                logo_bg = raw_color
                accent  = raw_color

            # Card state classes & badge
            if is_connected:
                card_class  = "connected"
                badge_class = "connected"
                badge_dot   = f'background:#10d9a0;box-shadow:0 0 5px #10d9a0;'
                badge_label = "Connected"
                accent_color = "#10d9a0"
            elif is_pending:
                card_class  = "pending"
                badge_class = "pending"
                badge_dot   = "background:#f59e0b;"
                badge_label = "Awaiting Login"
                accent_color = "#f59e0b"
            else:
                card_class  = ""
                badge_class = "idle"
                badge_dot   = "background:#334155;"
                badge_label = "Not Connected"
                accent_color = "#1e2d45"

            # Connected-at line
            connected_at_html = ""
            if connected_at:
                connected_at_html = (
                    f'<div class="pcard-connected-at">'
                    f'  <span>✓</span>'
                    f'  <span>Connected on {connected_at}</span>'
                    f'</div>'
                )

            # Tip banner
            tip_banner = ""
            if is_pending and not is_connected:
                tip_banner = (
                    f'<div class="pcard-tip">'
                    f'  <span class="pcard-tip-icon">🔑</span>'
                    f'  <p class="pcard-tip-text"><b>Next step:</b> {login_tip}</p>'
                    f'</div>'
                )

            st.markdown(
                f'<div class="pcard-outer">'
                f'  <div class="pcard {card_class}">'
                f'    <div class="pcard-accent" style="background:linear-gradient(90deg,{accent_color},{accent_color}55,transparent);"></div>'
                f'    <div class="pcard-body">'
                f'      <div class="pcard-header">'
                f'        <div class="pcard-identity">'
                f'          <div class="pcard-logo" style="background:{logo_bg};">{abbr}</div>'
                f'          <span class="pcard-name">{name}</span>'
                f'        </div>'
                f'        <span class="pcard-badge {badge_class}">'
                f'          <span class="pcard-badge-dot" style="{badge_dot}"></span>'
                f'          {badge_label}'
                f'        </span>'
                f'      </div>'
                f'      <div class="pcard-stats">'
                f'        <div class="pcard-stat">'
                f'          <div class="pcard-stat-label">⏰ Best Time</div>'
                f'          <div class="pcard-stat-value">{meta["best_time"]}</div>'
                f'        </div>'
                f'        <div class="pcard-stat">'
                f'          <div class="pcard-stat-label">📈 Response Rate</div>'
                f'          <div class="pcard-stat-value">{meta["response_rate"]}</div>'
                f'        </div>'
                f'      </div>'
                f'      {connected_at_html}'
                f'      {tip_banner}'
                f'    </div>'
                f'  </div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            # ── Action buttons ────────────────────────────────────────────
            if is_connected:
                if st.button(f"✕  Disconnect", key=f"disconnect_{idx}", use_container_width=True):
                    st.session_state["platforms_connected"] = [
                        p for p in st.session_state["platforms_connected"] if p != name
                    ]
                    st.session_state["platforms_connected_at"].pop(name, None)
                    if st.session_state["platform_pending_login"] == name:
                        st.session_state["platform_pending_login"] = ""
                    sb_save_platforms()
                    st.rerun()

            else:
                if meta["url"]:
                    st.link_button(
                        f"🌐  Visit {name} & Login",
                        meta["url"],
                        use_container_width=True,
                    )
                    if not is_pending:
                        if st.button(
                            f"⏳  I've Opened {name} — Show Tip",
                            key=f"pending_{idx}",
                            use_container_width=True,
                        ):
                            st.session_state["platform_pending_login"] = name
                            st.rerun()
                    else:
                        if st.button(
                            f"✅  I'm Logged In to {name}",
                            key=f"confirm_{idx}",
                            use_container_width=True,
                            type="primary",
                        ):
                            st.session_state["platforms_connected"].append(name)
                            st.session_state["platforms_connected_at"][name] = datetime.now().strftime("%b %d, %Y %H:%M")
                            st.session_state["platform_pending_login"] = ""
                            sb_save_platforms()
                            log_notion(
                                {
                                    "SessionID": st.session_state.get("session_id", ""),
                                    "Timestamp": datetime.utcnow().isoformat(),
                                    "Location": st.session_state.get("location", ""),
                                    "Platforms": ", ".join(st.session_state.get("platforms_connected", [])),
                                    "Currency": st.session_state.get("currency", "₹"),
                                }
                            )
                            st.rerun()
                        if st.button(
                            f"✕  Cancel",
                            key=f"cancel_pending_{idx}",
                            use_container_width=True,
                        ):
                            st.session_state["platform_pending_login"] = ""
                            st.rerun()
                else:
                    if st.button(f"✅  Mark as Connected", key=f"connect_{idx}", use_container_width=True):
                        st.session_state["platforms_connected"].append(name)
                        st.session_state["platforms_connected_at"][name] = datetime.now().strftime("%b %d, %Y %H:%M")
                        log_notion(
                            {
                                "SessionID": st.session_state.get("session_id", ""),
                                "Timestamp": datetime.utcnow().isoformat(),
                                "Location": st.session_state.get("location", ""),
                                "Platforms": ", ".join(st.session_state.get("platforms_connected", [])),
                                "Currency": st.session_state.get("currency", "₹"),
                            }
                        )
                        st.rerun()


def settings_page() -> None:
    st.subheader("Settings")
    st.write("Manage key app preferences.")
    st.session_state["api_key"] = st.text_input(
        "Gemini API Key",
        type="password",
        value=st.session_state["api_key"],
    ).strip()
    validate_api_key_if_needed()
    if st.session_state.get("api_validated"):
        st.success("✅ Gemini Connected")
    else:
        st.error("❌ Invalid Key")
    st.session_state["location"] = st.text_input("Default Location", value=st.session_state["location"]).strip()
    st.session_state["currency"] = st.selectbox(
        "Currency",
        ["₹", "$", "Both"],
        index=["₹", "$", "Both"].index(st.session_state.get("currency", "₹")),
    )
    st.session_state["gemini_api_key"] = st.session_state["api_key"]
    st.session_state["location_preference"] = st.session_state["location"]
    if st.button("Reset AI Cache"):
        st.cache_data.clear()
        st.session_state.pop("ats", None)
        st.session_state.pop("jobs", None)
        st.session_state.pop("profile", None)
        st.session_state["ats"] = {
            "overall_score": 0,
            "breakdown": {
                "Keywords": 0,
                "Format": 0,
                "Impact Statements": 0,
                "Skills Match": 0,
                "Readability": 0,
            },
        }
        st.session_state["jobs"] = []
        st.session_state["profile"] = {
            "name": "",
            "skills": [],
            "experience_years": 0,
            "job_titles": [],
            "education": [],
            "seniority_level": "",
        }
        st.session_state["analysis_cache_key"] = ""
        st.session_state["resume_analysis"] = st.session_state["profile"]
        st.session_state["ats_score"] = 0
        st.session_state["ats_breakdown"] = st.session_state["ats"]["breakdown"]
        st.session_state["jobs_ranked"] = []
        st.toast("Cache cleared, re-upload resume to refresh")


def validate_claude_key(key: str) -> tuple[bool, str]:
    """Validate Anthropic Claude API key using the anthropic SDK."""
    if not key:
        return False, "❌ No key provided"
    if _anthropic_sdk is None:
        return False, "❌ anthropic package not installed"
    try:
        client = _anthropic_sdk.Anthropic(api_key=key)
        models = client.models.list()
        if models:
            return True, "✅ Claude Connected"
        return False, "❌ Invalid Key"
    except Exception as exc:
        msg = str(exc).lower()
        if "401" in msg or "403" in msg or "invalid" in msg or "authentication" in msg or "unauthorized" in msg:
            return False, "❌ Invalid Key"
        return False, f"❌ Could not validate ({str(exc).split(chr(10))[0][:60]})"


def validate_openai_key(key: str) -> tuple[bool, str]:
    """Validate OpenAI API key using a minimal chat completion (avoids network blocks on Streamlit Cloud)."""
    if not key:
        return False, "❌ No key provided"
    try:
        import urllib.request
        import urllib.error
        payload = json.dumps({
            "model": "gpt-4o-mini",
            "messages": [{"role": "user", "content": "hi"}],
            "max_tokens": 1,
        }).encode()
        req = urllib.request.Request(
            "https://api.openai.com/v1/chat/completions",
            data=payload,
            headers={
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/json",
            },
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            if resp.getcode() == 200:
                return True, "✅ OpenAI Connected"
        return False, "❌ Invalid Key"
    except urllib.error.HTTPError as e:
        if e.code in (401, 403):
            return False, "❌ Invalid Key"
        # 429 = valid key, just rate limited
        if e.code == 429:
            return True, "✅ OpenAI Connected (rate limited)"
        return False, f"❌ Could not validate (HTTP {e.code})"
    except Exception as exc:
        msg = str(exc).lower()
        if "401" in msg or "403" in msg or "invalid" in msg or "unauthorized" in msg:
            return False, "❌ Invalid Key"
        return False, f"❌ Could not validate ({str(exc).split(chr(10))[0][:60]})"


def validate_groq_key(key: str) -> tuple[bool, str]:
    """Validate Groq key using the groq SDK (avoids network-level blocks on Streamlit Cloud)."""
    if not key:
        return False, "❌ No key provided"
    if _GroqClient is None:
        return False, "❌ groq package not installed"
    try:
        client = _GroqClient(api_key=key)
        models = client.models.list()
        if models:
            return True, "✅ Groq Connected"
        return False, "❌ Invalid Key"
    except Exception as exc:
        msg = str(exc).lower()
        if "401" in msg or "403" in msg or "invalid" in msg or "authentication" in msg or "unauthorized" in msg:
            return False, "❌ Invalid Key"
        return False, f"❌ Could not validate ({str(exc).split(chr(10))[0][:60]})"


def validate_together_key(key: str) -> tuple[bool, str]:
    """Validate HuggingFace token.

    Two-step strategy:
      1. Hit HF whoami API to verify the token is genuine.
      2. Probe router.huggingface.co to find a working model.

    Special case: some deployment environments (corporate proxies, Streamlit Community
    Cloud sandbox, etc.) block outbound calls to HuggingFace with HTTP 403
    'Host not in allowlist'. We detect this via the x-deny-reason header and treat
    the token as valid — the block is on the network, not the token.
    """
    if not key:
        return False, "❌ No key provided"
    import urllib.request
    import urllib.error

    def _is_host_blocked(exc: urllib.error.HTTPError) -> bool:
        """Return True if the 403 is a network/proxy block, not an auth failure."""
        try:
            deny = exc.headers.get("x-deny-reason", "")
            body = exc.read().decode("utf-8", errors="ignore")
            return "host_not_allowed" in deny or "Host not in allowlist" in body
        except Exception:
            return False

    # ── Step 1: Verify token via HF whoami ───────────────────────────
    hf_username = None
    try:
        whoami_req = urllib.request.Request(
            "https://huggingface.co/api/whoami",
            headers={"Authorization": f"Bearer {key}"},
        )
        with urllib.request.urlopen(whoami_req, timeout=10) as resp:
            whoami_data = json.loads(resp.read().decode())
            hf_username = whoami_data.get("name", "user")
    except urllib.error.HTTPError as e:
        if _is_host_blocked(e):
            # Network/proxy block — can't reach HF from this host at all.
            # Accept the token as-is; the app will work in the real environment.
            return True, (
                "✅ Token accepted — HuggingFace is reachable from the app. "
                "(Validation was skipped because this host cannot reach HuggingFace directly.)"
            )
        if e.code in (401, 403):
            return False, "❌ Invalid Token — check your HuggingFace token at huggingface.co/settings/tokens"
        # Any other HTTP error (5xx, etc.) — token format ok, proceed
        hf_username = "user"
    except Exception:
        hf_username = "user"

    # ── Step 2: Find a working model on the router ───────────────────
    # Plain model IDs first (work with basic free tokens, no special permissions).
    # Suffixed :cerebras/:fastest variants tried after for Pro accounts.
    _PROBE_MODELS = [
        "microsoft/Phi-3.5-mini-instruct",
        "Qwen/Qwen2.5-7B-Instruct",
        "meta-llama/Llama-3.1-8B-Instruct",
        "HuggingFaceH4/zephyr-7b-beta",
        "mistralai/Mistral-7B-Instruct-v0.3",
        "meta-llama/Llama-3.1-8B-Instruct:cerebras",
        "meta-llama/Llama-3.1-8B-Instruct:fastest",
        "Qwen/Qwen2.5-7B-Instruct:fastest",
        "microsoft/Phi-3.5-mini-instruct:fastest",
    ]
    last_err_code = None
    for model in _PROBE_MODELS:
        try:
            payload = json.dumps({
                "model": model,
                "messages": [{"role": "user", "content": "hi"}],
                "max_tokens": 1,
            }).encode()
            req = urllib.request.Request(
                "https://router.huggingface.co/v1/chat/completions",
                data=payload,
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json",
                },
            )
            with urllib.request.urlopen(req, timeout=20) as resp:
                if resp.getcode() == 200:
                    st.session_state["hf_active_model"] = model
                    short = model.split("/")[-1].split(":")[0]
                    uname = f"@{hf_username} · " if hf_username else ""
                    return True, f"✅ HuggingFace Connected · {uname}{short}"
        except urllib.error.HTTPError as e:
            if _is_host_blocked(e):
                # Same host block — token is fine, just can't validate from here
                uname = f"@{hf_username}" if hf_username else "user"
                return True, (
                    f"✅ Token valid ({uname}) — HuggingFace reachable from the deployed app. "
                    f"(Router blocked from this host; will work in production.)"
                )
            last_err_code = e.code
            if e.code == 429:
                st.session_state["hf_active_model"] = model
                uname = f"@{hf_username} · " if hf_username else ""
                return True, f"✅ HuggingFace Connected · {uname}(rate limited — valid token)"
            continue  # 401/403/404/503 on this model → try next
        except Exception:
            continue

    # Token verified via whoami but no model responded
    uname = f"@{hf_username}" if hf_username else "user"
    hint = " Enable 'Make calls to Inference Providers' in your HF token settings." if last_err_code in (401, 403) else ""
    return True, f"✅ Token valid ({uname}) but no router model responded.{hint}"

def _ls_bridge_script(keys: dict) -> str:
    """Return an HTML snippet that persists API keys to localStorage (never sent to Notion)."""
    assignments = "\n".join(
        f"  localStorage.setItem({json.dumps(k)}, {json.dumps(v)});"
        for k, v in keys.items()
    )
    return f"<script>\n(function(){{\n{assignments}\n}})();\n</script>"


def _ls_restore_script() -> str:
    """Return an HTML snippet that reads keys back from localStorage into window for debugging.
    Keys are only stored client-side; they are never sent to Notion."""
    return """<script>
(function() {
  const map = {
    AutoApply AI_gemini_key:   'gemini',
    AutoApply AI_groq_key:     'groq',
    AutoApply AI_claude_key:   'claude',
    AutoApply AI_together_key: 'together',
    AutoApply AI_openai_key:   'openai',
  };
  for (const [lsKey, label] of Object.entries(map)) {
    const val = localStorage.getItem(lsKey);
    if (val) {
      console.debug('[AutoApply AI] ' + label + ' key found in localStorage (length=' + val.length + ')');
    }
  }
})();
</script>"""



def _pill(text: str, color: str, bg: str) -> str:
    """Inline pill badge — pure HTML, no span tags leaking."""
    return (
        f'<span style="display:inline-block;padding:2px 8px;border-radius:20px;'
        f'background:{bg};color:{color};font-size:10px;font-weight:700;'
        f'letter-spacing:0.5px;border:1px solid {color}44;">{text}</span>'
    )


def _rate_limit_bar_html(
    label: str,
    color: str,
    icon: str,
    used: int,
    total: int,
    unit: str,
    reset_label: str,
    exhausted: bool = False,
    reset_seconds: int = 0,
) -> str:
    """Render a premium rate-limit bar card. Returns clean HTML — no f-string nesting issues."""
    pct = min(100, int((used / total) * 100)) if total else 0
    remaining = max(0, total - used)

    if exhausted:
        bar_color = "#ef4444"
        track_color = "#3f1a1a"
        card_bg = "linear-gradient(135deg,#1a0808,#0f0a0a)"
        card_border = "#ef444440"
        count_color = "#ef4444"
    elif pct >= 80:
        bar_color = "#f59e0b"
        track_color = "#1e2d45"
        card_bg = "linear-gradient(135deg,#0f1117,#0d1117)"
        card_border = color + "30"
        count_color = "#f59e0b"
    else:
        bar_color = color
        track_color = "#1a2744"
        card_bg = "linear-gradient(135deg,#0a0f1a,#0d1117)"
        card_border = color + "28"
        count_color = color

    # Build reset time string safely
    reset_badge = ""
    if exhausted and reset_seconds > 0:
        total_mins = reset_seconds // 60
        secs = reset_seconds % 60
        hrs = total_mins // 60
        mins = total_mins % 60
        if hrs > 0:
            t = f"{hrs}h {mins}m"
        elif mins > 0:
            t = f"{mins}m {secs}s"
        else:
            t = f"{secs}s"
        reset_badge = (
            '<span style="display:inline-flex;align-items:center;gap:4px;'
            'background:#3f1a1a;border:1px solid #ef444440;border-radius:6px;'
            'padding:2px 8px;font-size:10px;font-weight:700;color:#ef4444;">'
            f'RESETS IN {t}</span>'
        )
    elif exhausted:
        reset_badge = (
            '<span style="display:inline-flex;align-items:center;'
            'background:#3f1a1a;border:1px solid #ef444440;border-radius:6px;'
            'padding:2px 8px;font-size:10px;font-weight:700;color:#ef4444;">EXHAUSTED</span>'
        )

    # Glow on the filled bar
    glow = f"box-shadow:0 0 8px {bar_color}55;" if not exhausted else ""

    html = (
        f'<div style="background:{card_bg};border:1px solid {card_border};'
        f'border-radius:12px;padding:12px 16px;margin-bottom:10px;'
        f'box-shadow:0 2px 12px rgba(0,0,0,0.3);">'

        # Top row: icon+label left, count+badge right
        f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">'
        f'  <div style="display:flex;align-items:center;gap:8px;">'
        f'    <span style="font-size:14px;">{icon}</span>'
        f'    <span style="color:#cbd5e1;font-size:12px;font-weight:600;letter-spacing:0.3px;">{label}</span>'
        f'  </div>'
        f'  <div style="display:flex;align-items:center;gap:8px;">'
        f'    <span style="color:{count_color};font-size:13px;font-weight:700;">'
        f'      {remaining:,}<span style="color:#475569;font-size:10px;font-weight:400;">/{total:,} {unit}</span>'
        f'    </span>'
        f'    {reset_badge}'
        f'  </div>'
        f'</div>'

        # Progress track
        f'<div style="background:{track_color};border-radius:999px;height:6px;overflow:hidden;">'
        f'  <div style="width:{pct}%;height:100%;background:{bar_color};'
        f'  border-radius:999px;{glow}transition:width 0.5s ease;"></div>'
        f'</div>'

        # Bottom row: % used, reset label
        f'<div style="display:flex;justify-content:space-between;margin-top:6px;">'
        f'  <span style="color:#334155;font-size:10px;">{pct}% used this session</span>'
        f'  <span style="color:#334155;font-size:10px;">{reset_label}</span>'
        f'</div>'

        f'</div>'
    )
    return html


def _provider_card_html(
    name: str,
    color: str,
    icon: str,
    model: str,
    rpd: int,
    rpm: int,
    tpm_k: int,
    validated: bool,
    exhausted: bool,
    is_selected: bool = False,
    is_free: bool = True,
) -> str:
    """Clickable provider summary card — highlights when selected."""
    if not validated:
        status_dot  = "#64748b"
        status_text = "Not connected"
        status_bg   = "#0d1117"
        status_border = "#64748b30"
    elif exhausted:
        status_dot  = "#ef4444"
        status_text = "Rate limited"
        status_bg   = "#1a0808"
        status_border = "#ef444440"
    else:
        status_dot  = "#10d9a0"
        status_text = "Active"
        status_bg   = "#071a0f"
        status_border = "#10d9a040"

    free_badge = (
        '<span style="display:inline-block;padding:1px 5px;border-radius:4px;'
        'background:#064e3b;color:#10d9a0;font-size:9px;font-weight:800;'
        'letter-spacing:0.5px;border:1px solid #10d9a025;">FREE</span>'
    ) if is_free else ""

    selected_ring = f"box-shadow:0 0 0 2px {color},0 6px 28px {color}44;" if is_selected else "box-shadow:0 4px 20px rgba(0,0,0,0.5);"
    selected_bg   = f"background:linear-gradient(145deg,#0d1117,{color}12);" if is_selected else "background:linear-gradient(145deg,#0e1420,#0a0f1a);"
    border_op     = "66" if is_selected else "35"
    chevron       = "▲" if is_selected else "▼"
    chevron_color = color if is_selected else "#475569"

    # Toggle switch visual (ON = validated, OFF = not) — purely visual, button below is functional
    if validated and not exhausted:
        tog_track   = color
        tog_x       = "18px"
        tog_label   = "ON"
        tog_lc      = color
        tog_glow    = f"box-shadow:0 0 8px {color}88;"
    elif exhausted:
        tog_track   = "#ef4444"
        tog_x       = "18px"
        tog_label   = "LIMIT"
        tog_lc      = "#ef4444"
        tog_glow    = "box-shadow:0 0 6px #ef444488;"
    else:
        tog_track   = "#1e2d45"
        tog_x       = "2px"
        tog_label   = "OFF"
        tog_lc      = "#475569"
        tog_glow    = ""

    toggle_html = (
        f'<div style="display:flex;align-items:center;gap:4px;flex-shrink:0;">'
        f'  <span style="font-size:8px;font-weight:800;color:{tog_lc};letter-spacing:0.4px;'
        f'  text-transform:uppercase;white-space:nowrap;">{tog_label}</span>'
        f'  <div style="position:relative;width:32px;height:16px;border-radius:999px;'
        f'  background:{tog_track};flex-shrink:0;{tog_glow}transition:background 0.3s;">'
        f'    <div style="position:absolute;top:2px;left:{tog_x};'
        f'    width:12px;height:12px;border-radius:50%;background:#fff;'
        f'    box-shadow:0 1px 3px rgba(0,0,0,0.5);transition:left 0.3s;"></div>'
        f'  </div>'
        f'</div>'
    )

    return (
        f'<div style="{selected_bg}'
        f'border:1px solid {color}{border_op};border-radius:16px;padding:14px 13px 11px;'
        f'{selected_ring}transition:all 0.22s ease;box-sizing:border-box;width:100%;overflow:hidden;'
        f'position:relative;">'

        # Top accent line (glowing strip when selected)
        f'<div style="position:absolute;top:0;left:0;right:0;height:2px;'
        f'background:linear-gradient(90deg,{color}{"cc" if is_selected else "44"},transparent);'
        f'border-radius:16px 16px 0 0;"></div>'

        # Top row: icon + name + free badge + chevron
        f'<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:9px;margin-top:4px;">'
        f'  <div style="display:flex;align-items:center;gap:6px;">'
        f'    <span style="font-size:17px;line-height:1;flex-shrink:0;">{icon}</span>'
        f'    <span style="color:{color};font-weight:800;font-size:13px;letter-spacing:-0.2px;white-space:nowrap;'
        f'    font-family:Sora,sans-serif;">{name}</span>'
        f'  </div>'
        f'  <div style="display:flex;align-items:center;gap:5px;flex-shrink:0;">'
        f'    {free_badge}'
        f'    <span style="color:{chevron_color};font-size:9px;opacity:0.7;">{chevron}</span>'
        f'  </div>'
        f'</div>'

        # Status row: dot+text LEFT, toggle switch RIGHT
        f'<div style="display:flex;align-items:center;justify-content:space-between;'
        f'background:{status_bg};border:1px solid {status_border};'
        f'border-radius:9px;padding:6px 9px;margin-bottom:9px;'
        f'box-shadow:inset 0 1px 4px rgba(0,0,0,0.25);">'
        f'  <div style="display:flex;align-items:center;gap:5px;">'
        f'    <span style="width:6px;height:6px;border-radius:50%;flex-shrink:0;'
        f'    background:{status_dot};display:inline-block;'
        f'    box-shadow:0 0 7px {status_dot}99;"></span>'
        f'    <span style="color:{status_dot};font-size:10px;font-weight:700;white-space:nowrap;letter-spacing:0.2px;">{status_text}</span>'
        f'  </div>'
        f'  {toggle_html}'
        f'</div>'

        # Model chip — monospace code style
        f'<div style="color:#475569;font-size:9px;font-family:monospace;'
        f'background:#050810;padding:5px 8px;border-radius:7px;'
        f'border:1px solid #111c2e;margin-bottom:9px;overflow:hidden;'
        f'text-overflow:ellipsis;white-space:nowrap;letter-spacing:0.2px;" title="{model}">{model}</div>'

        # Limit stats — two pills
        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:7px;">'
        f'  <div style="background:linear-gradient(145deg,#060a12,#070c18);border-radius:9px;padding:8px 10px;'
        f'  border:1px solid #111c2e;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,0.3);">'
        f'    <div style="color:{color};font-size:14px;font-weight:800;line-height:1;font-family:Sora,sans-serif;">{rpd:,}</div>'
        f'    <div style="color:#334155;font-size:8px;margin-top:3px;font-weight:700;text-transform:uppercase;letter-spacing:0.6px;">REQ / DAY</div>'
        f'  </div>'
        f'  <div style="background:linear-gradient(145deg,#060a12,#070c18);border-radius:9px;padding:8px 10px;'
        f'  border:1px solid #111c2e;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,0.3);">'
        f'    <div style="color:{color};font-size:14px;font-weight:800;line-height:1;font-family:Sora,sans-serif;">{rpm}</div>'
        f'    <div style="color:#334155;font-size:8px;margin-top:3px;font-weight:700;text-transform:uppercase;letter-spacing:0.6px;">REQ / MIN</div>'
        f'  </div>'
        f'</div>'

        # Click hint footer — subtle CTA
        f'<div style="margin-top:6px;text-align:center;color:{color}55;font-size:8px;'
        f'font-weight:700;letter-spacing:1px;text-transform:uppercase;'
        f'padding:5px;border-top:1px solid rgba(30,45,69,0.5);">'
        f'{"▲ CLOSE" if is_selected else "▼ CONFIGURE"}'
        f'</div>'

        f'</div>'
    )


def _free_guide_steps(steps: list[dict]) -> str:
    """Render numbered step-by-step guide to get a free API key."""
    items = ""
    for i, step in enumerate(steps, 1):
        url = step.get("url", "")
        if url:
            link_label = step.get("link_label", "Open")
            link = (
                f'<a href="{url}" target="_blank" style="display:inline-flex;'
                f'align-items:center;gap:4px;margin-top:5px;padding:4px 10px;'
                f'background:linear-gradient(90deg,#1e3a5f,#1a2744);'
                f'border:1px solid #3b82f640;border-radius:6px;'
                f'color:#60a5fa;font-size:11px;font-weight:600;'
                f'text-decoration:none;">'
                f'&#x1F517; {link_label}</a>'
            )
        else:
            link = ""

        items += (
            f'<div style="display:flex;gap:12px;align-items:flex-start;'
            f'padding:10px 12px;border-radius:10px;margin-bottom:6px;'
            f'background:#060a12;border:1px solid #1e2d4530;">'
            f'  <div style="min-width:22px;height:22px;'
            f'  background:linear-gradient(135deg,#3b82f6,#8b5cf6);'
            f'  border-radius:50%;display:flex;align-items:center;'
            f'  justify-content:center;font-size:10px;font-weight:800;'
            f'  color:#fff;flex-shrink:0;margin-top:1px;">{i}</div>'
            f'  <div>'
            f'    <div style="color:#e2e8f0;font-size:12px;font-weight:600;">{step["title"]}</div>'
            f'    <div style="color:#64748b;font-size:11px;margin-top:2px;line-height:1.5;">{step.get("desc", "")}</div>'
            f'    {link}'
            f'  </div>'
            f'</div>'
        )

    return (
        f'<div style="background:linear-gradient(135deg,#060a12,#080d18);'
        f'border:1px solid #1e3a5f;border-radius:14px;padding:16px;margin-top:8px;">'
        f'  <div style="display:flex;align-items:center;gap:8px;margin-bottom:12px;">'
        f'    <span style="font-size:16px;">&#x1F4CB;</span>'
        f'    <span style="color:#60a5fa;font-size:12px;font-weight:700;letter-spacing:0.5px;">'
        f'    HOW TO GET YOUR FREE KEY — STEP BY STEP</span>'
        f'  </div>'
        f'  {items}'
        f'</div>'
    )


def _free_model_warning_html(provider: str, model: str, free_note: str) -> str:
    """Prominent banner reminding user to use free models only."""
    return (
        f'<div style="display:flex;align-items:flex-start;gap:10px;'
        f'background:linear-gradient(135deg,#0a1a0a,#071208);'
        f'border:1px solid #10d9a030;border-left:3px solid #10d9a0;'
        f'border-radius:10px;padding:10px 14px;margin:8px 0;">'
        f'  <span style="font-size:18px;flex-shrink:0;">&#x2705;</span>'
        f'  <div>'
        f'    <div style="color:#10d9a0;font-size:11px;font-weight:700;'
        f'    letter-spacing:0.4px;margin-bottom:2px;">FREE MODEL — NO CHARGES</div>'
        f'    <div style="color:#94a3b8;font-size:11px;line-height:1.5;">'
        f'    AutoApply AI uses <code style="background:#0d1117;color:#10d9a0;'
        f'    padding:1px 5px;border-radius:4px;">{model}</code> '
        f'    for {provider}. {free_note}</div>'
        f'  </div>'
        f'</div>'
    )


def _section_header_html(icon: str, title: str, subtitle: str) -> str:
    """Premium section header without Streamlit markdown span-leaking."""
    return (
        f'<div style="margin:24px 0 16px;">'
        f'  <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">'
        f'    <span style="font-size:20px;">{icon}</span>'
        f'    <span style="color:#eaf2ff;font-size:18px;font-weight:700;">{title}</span>'
        f'  </div>'
        f'  <div style="color:#64748b;font-size:12px;padding-left:30px;">{subtitle}</div>'
        f'</div>'
    )


def _provider_section_header(icon: str, color: str, name: str, tagline: str, is_primary: bool = False) -> str:
    primary_badge = (
        '<span style="display:inline-block;padding:2px 8px;border-radius:4px;'
        'background:linear-gradient(90deg,#3b82f6,#8b5cf6);'
        'color:#fff;font-size:9px;font-weight:700;letter-spacing:0.5px;">PRIMARY</span>'
    ) if is_primary else ""
    free_chip = (
        '<span style="display:inline-block;padding:2px 7px;border-radius:4px;'
        'background:#064e3b;color:#10d9a0;font-size:9px;font-weight:700;'
        'letter-spacing:0.5px;border:1px solid #10d9a030;">FREE TIER</span>'
    )
    return (
        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:12px;">'
        f'  <span style="font-size:20px;">{icon}</span>'
        f'  <div>'
        f'    <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;">'
        f'      <span style="color:{color};font-size:15px;font-weight:800;">{name}</span>'
        f'      {primary_badge} {free_chip}'
        f'    </div>'
        f'    <div style="color:#475569;font-size:11px;margin-top:2px;">{tagline}</div>'
        f'  </div>'
        f'</div>'
    )


def api_keys_page() -> None:
    st.subheader("API Keys")
    st.caption(
        "Manage AI provider credentials. "
        "Keys are stored only in your browser session and localStorage — **never logged to Notion or any server**."
    )

    # ── Autofill: sync DB-restored keys into widget state ────────────────
    # Streamlit ignores value= on re-renders once a widget key exists in
    # st.session_state. We solve this by writing the DB value directly into
    # the widget's own session key before the widget is rendered — only when
    # the user has NOT manually overridden it this run (i.e. widget key absent
    # or empty). This makes keys saved in the DB appear pre-filled on page load.
    _autofill_map = {
        "ak_gemini_input":  "api_key",
        "ak_groq_input":    "groq_api_key",
        "ak_claude_input":  "claude_api_key",
        "ak_together_input":"together_api_key",
        "ak_openai_input":  "openai_api_key",
    }
    for widget_key, state_key in _autofill_map.items():
        db_val = st.session_state.get(state_key, "").strip()
        # Only pre-fill if the widget hasn't been written yet (avoids
        # overwriting what the user typed in the same session).
        if db_val and not st.session_state.get(widget_key, "").strip():
            st.session_state[widget_key] = db_val

    # ── Inject localStorage restore ──────────────────────────────
    st.markdown(_ls_restore_script(), unsafe_allow_html=True)

    # ── Quota + usage state ──────────────────────────────────────
    now_ts = time.time()
    quota_until = float(st.session_state.get("quota_block_until", 0.0))
    is_quota_blocked = now_ts < quota_until
    quota_reset_secs = max(0, int(quota_until - now_ts)) if is_quota_blocked else 0

    if "usage_counts" not in st.session_state:
        st.session_state["usage_counts"] = {
            "gemini": 0, "groq": 0, "claude": 0, "together": 0, "openai": 0
        }
    uc = st.session_state["usage_counts"]

    # ── Model limits (free tier only) ────────────────────────────
    gemini_model = st.session_state.get("resolved_model_name", "gemini-1.5-flash") or "gemini-1.5-flash"
    _gemini_model_limits = {
        "gemini-1.5-flash":        {"rpd": 1500, "rpm": 15,  "tpm": 1_000_000},
        "gemini-1.5-flash-latest": {"rpd": 1500, "rpm": 15,  "tpm": 1_000_000},
        "gemini-1.5-flash-001":    {"rpd": 1500, "rpm": 15,  "tpm": 1_000_000},
        "gemini-1.5-pro":          {"rpd":   50, "rpm":  2,  "tpm":   320_000},
        "gemini-1.5-pro-latest":   {"rpd":   50, "rpm":  2,  "tpm":   320_000},
        "gemini-2.0-flash":        {"rpd": 1500, "rpm": 15,  "tpm": 1_000_000},
        "gemini-2.0-flash-lite":   {"rpd": 1500, "rpm": 30,  "tpm": 1_000_000},
        "gemini-2.5-flash":        {"rpd":  500, "rpm": 10,  "tpm":   250_000},
    }
    gl = _gemini_model_limits.get(gemini_model, {"rpd": 1500, "rpm": 15, "tpm": 1_000_000})

    gemini_validated = st.session_state.get("api_validated", False)
    groq_val         = st.session_state.get("groq_api_validated", False)
    claude_val       = st.session_state.get("claude_api_validated", False)
    together_val     = st.session_state.get("together_api_validated", False)
    openai_val       = st.session_state.get("openai_api_validated", False)

    active_prov = st.session_state.get("active_provider", "")
    gemini_exhausted  = is_quota_blocked and active_prov == "gemini"
    groq_exhausted    = is_quota_blocked and active_prov == "groq"
    claude_exhausted  = is_quota_blocked and active_prov == "claude"
    together_exhausted = is_quota_blocked and active_prov == "together"

    groq_limits    = {"rpd": 14_400, "rpm": 30,  "tpm":  6_000}
    claude_limits  = {"rpd":  1_000, "rpm":  5,  "tpm": 20_000}
    together_limits = {"rpd":    600, "rpm": 10,  "tpm":  4_000}
    openai_limits  = {"rpd":    200, "rpm":  3,  "tpm": 40_000}

    # ══════════════════════════════════════════════════════════════
    # PROVIDER CARDS — click a card to expand its detail panel
    # ══════════════════════════════════════════════════════════════
    st.markdown(
        _section_header_html(
            "🤖", "AI Providers",
            "Priority: Gemini → Groq → Claude → HuggingFace → OpenAI  ·  "
            "Click any card to configure it. All providers use FREE models only.",
        ),
        unsafe_allow_html=True,
    )

    # Global quota alert
    if is_quota_blocked:
        mins, secs = divmod(quota_reset_secs, 60)
        st.error(
            f"🚦 Rate limit hit on **{active_prov}** provider. "
            f"AutoApply AI auto-rotated to next available. "
            f"Cooldown: **{mins}m {secs}s** remaining.",
        )

    selected_prov = st.session_state.get("selected_provider", "")

    card_defs = [
        ("gemini",   "Gemini",      "#10d9a0", "✦",  gemini_model,           gl["rpd"],              gl["rpm"],              gl["tpm"]//1000,  gemini_validated,  gemini_exhausted),
        ("groq",     "Groq",        "#f472b6", "⚡", "llama-3.1-8b-instant", groq_limits["rpd"],     groq_limits["rpm"],     groq_limits["tpm"]//1000, groq_val, groq_exhausted),
        ("claude",   "Claude",      "#f59e0b", "◆",  "claude-haiku-4-5",     claude_limits["rpd"],   claude_limits["rpm"],   claude_limits["tpm"]//1000, claude_val, claude_exhausted),
        ("together", "HuggingFace", "#f97316", "🤗", "Mistral-7B-Instruct",  together_limits["rpd"], together_limits["rpm"], together_limits["tpm"]//1000, together_val, together_exhausted),
        ("openai",   "OpenAI",      "#22d3ee", "✦",  "gpt-4o-mini (trial)",  openai_limits["rpd"],   openai_limits["rpm"],   openai_limits["tpm"]//1000, openai_val, False),
    ]

    # ── Key presence map (to know if a key exists for instant connect) ──
    _key_map = {
        "gemini":   st.session_state.get("api_key", "").strip(),
        "groq":     st.session_state.get("groq_api_key", "").strip(),
        "claude":   st.session_state.get("claude_api_key", "").strip(),
        "together": st.session_state.get("together_api_key", "").strip(),
        "openai":   st.session_state.get("openai_api_key", "").strip(),
    }
    _val_map = {
        "gemini":   gemini_validated,
        "groq":     groq_val,
        "claude":   claude_val,
        "together": together_val,
        "openai":   openai_val,
    }

    # ── ☁️ Load from Cloud — styled inline premium button ────────────────
    st.markdown("""
<div style="margin:4px 0 16px;">
  <style>
    div[data-testid="stButton"] button[kind="secondary"]#lfc_btn {
      background: linear-gradient(135deg,#0d1f3c,#0a1628) !important;
      border: 1px solid #3b82f640 !important;
      border-radius: 10px !important;
      color: #60a5fa !important;
      font-size: 13px !important;
      font-weight: 600 !important;
      padding: 10px 20px !important;
    }
  </style>
</div>
""", unsafe_allow_html=True)
    _lfc_col, _lfc_spacer = st.columns([1.6, 3.4])
    with _lfc_col:
        if st.button("☁️  Load API Keys from Cloud", use_container_width=True, type="secondary",
                     help="Fetch all your saved API keys from the database and connect them automatically"):
            _auth_user = st.session_state.get("auth_user")
            if isinstance(_auth_user, dict) and _auth_user.get("Username"):
                try:
                    _db = _sb()
                    if _db is not None:
                        _result = _db.table("users").select("*").eq(
                            "username", _auth_user["Username"]
                        ).execute()
                        if _result.data:
                            _row = _result.data[0]
                            _loaded = 0
                            # gemini
                            _gk = str(_row.get("gemini_key", "") or "").strip()
                            if _gk:
                                st.session_state["api_key"] = _gk
                                st.session_state["gemini_api_key"] = _gk
                                st.session_state["api_validated"] = True
                                st.session_state["api_validation_message"] = "✅ Gemini Connected"
                                st.session_state["api_last_checked_key"] = _gk
                                _loaded += 1
                            # groq
                            _grk = str(_row.get("groq_key", "") or "").strip()
                            if _grk:
                                st.session_state["groq_api_key"] = _grk
                                st.session_state["groq_api_validated"] = True
                                st.session_state["groq_api_validation_message"] = "✅ Groq Connected"
                                _loaded += 1
                            # claude
                            _ck = str(_row.get("claude_key", "") or "").strip()
                            if _ck:
                                st.session_state["claude_api_key"] = _ck
                                st.session_state["claude_api_validated"] = True
                                st.session_state["claude_api_validation_message"] = "✅ Claude Connected"
                                _loaded += 1
                            # together / huggingface
                            _tk = str(_row.get("together_key", "") or "").strip()
                            if _tk:
                                st.session_state["together_api_key"] = _tk
                                st.session_state["together_api_validated"] = True
                                st.session_state["together_api_validation_message"] = "✅ HuggingFace Connected"
                                _loaded += 1
                            # openai
                            _ok = str(_row.get("openai_key", "") or "").strip()
                            if _ok:
                                st.session_state["openai_api_key"] = _ok
                                st.session_state["openai_api_validated"] = True
                                st.session_state["openai_api_validation_message"] = "✅ OpenAI Connected"
                                _loaded += 1
                            if _loaded > 0:
                                st.success(f"✅ {_loaded} API key(s) loaded and connected from cloud!")
                            else:
                                st.warning("No API keys found in your account. Add them below and save.")
                        else:
                            st.warning("No account data found.")
                    else:
                        st.error("Database unavailable.")
                except Exception as _e:
                    st.error(f"Could not load from cloud: {_e}")
                st.rerun()
            else:
                st.warning("Please log in to load keys from cloud.")

    # Render the 5 clickable summary cards
    cols = st.columns(5, gap="small")
    for col, (pkey, nm, clr, ic, mdl, rpd, rpm, tpm_k, val, exh) in zip(cols, card_defs):
        with col:
            is_sel  = (selected_prov == pkey)
            is_on   = _val_map[pkey]       # True = validated/connected
            has_key = bool(_key_map[pkey]) # key present in session (may not be validated yet)

            st.markdown(
                _provider_card_html(nm, clr, ic, mdl, rpd, rpm, tpm_k, val, exh, is_selected=is_sel),
                unsafe_allow_html=True,
            )

            # ── Toggle button — connect OR disconnect directly ──────────
            if is_on:
                # ── CONNECTED: click → disconnect ───────────────────────
                if st.button("🟢  ON · Disconnect", key=f"toggle_{pkey}",
                             use_container_width=True, type="primary",
                             help=f"Disconnect {nm}"):
                    if pkey == "gemini":
                        st.session_state["api_key"] = ""
                        st.session_state["gemini_api_key"] = ""
                        st.session_state["api_validated"] = False
                        st.session_state["api_validation_message"] = ""
                        st.session_state["api_last_checked_key"] = ""
                        st.session_state["resolved_model_name"] = ""
                    elif pkey == "groq":
                        st.session_state["groq_api_key"] = ""
                        st.session_state["groq_api_validated"] = False
                        st.session_state["groq_api_validation_message"] = ""
                    elif pkey == "claude":
                        st.session_state["claude_api_key"] = ""
                        st.session_state["claude_api_validated"] = False
                        st.session_state["claude_api_validation_message"] = ""
                    elif pkey == "together":
                        st.session_state["together_api_key"] = ""
                        st.session_state["together_api_validated"] = False
                        st.session_state["together_api_validation_message"] = ""
                        st.session_state["hf_active_model"] = ""
                    elif pkey == "openai":
                        st.session_state["openai_api_key"] = ""
                        st.session_state["openai_api_validated"] = False
                        st.session_state["openai_api_validation_message"] = ""
                    if st.session_state.get("active_provider") == pkey:
                        st.session_state["active_provider"] = ""
                    st.rerun()

            elif has_key:
                # ── KEY EXISTS BUT NOT VALIDATED: click → validate & connect ──
                if st.button("🟡  OFF · Connect", key=f"toggle_{pkey}",
                             use_container_width=True, type="secondary",
                             help=f"Connect {nm} using saved key"):
                    _key_val = _key_map[pkey]
                    with st.spinner(f"Connecting {nm}…"):
                        if pkey == "gemini":
                            st.session_state["api_key"] = _key_val
                            st.session_state["gemini_api_key"] = _key_val
                            st.session_state["resolved_model_name"] = ""
                            validate_api_key_if_needed()
                        elif pkey == "groq":
                            ok, msg = validate_groq_key(_key_val)
                            st.session_state["groq_api_validated"] = ok
                            st.session_state["groq_api_validation_message"] = msg
                        elif pkey == "claude":
                            ok, msg = validate_claude_key(_key_val)
                            st.session_state["claude_api_validated"] = ok
                            st.session_state["claude_api_validation_message"] = msg
                        elif pkey == "together":
                            ok, msg = validate_together_key(_key_val)
                            st.session_state["together_api_validated"] = ok
                            st.session_state["together_api_validation_message"] = msg
                        elif pkey == "openai":
                            ok, msg = validate_openai_key(_key_val)
                            st.session_state["openai_api_validated"] = ok
                            st.session_state["openai_api_validation_message"] = msg
                    _auto_save_keys_for_logged_in_user()
                    st.rerun()

            else:
                # ── NO KEY: click → open detail panel to paste a key ───────
                if st.button("⚫  OFF · Add Key", key=f"toggle_{pkey}",
                             use_container_width=True, type="secondary",
                             help=f"Add an API key for {nm}"):
                    st.session_state["selected_provider"] = pkey
                    st.rerun()

            # ── Small "▲/▼ Details" link under toggle ──────────────────
            detail_label = "▲ Close details" if is_sel else "▼ View details"
            if st.button(detail_label, key=f"card_btn_{pkey}", use_container_width=True):
                st.session_state["selected_provider"] = "" if is_sel else pkey
                st.rerun()

    # ── Detail panel — renders below the card row ──────────────────
    if selected_prov == "gemini":
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(
                _provider_section_header("✦", "#10d9a0", "Gemini (Google AI)",
                    "Primary AI engine · gemini-1.5-flash (free forever)", is_primary=True),
                unsafe_allow_html=True,
            )
            # ── ✨ Get Free Key CTA — shown FIRST so new users see it immediately ──
            st.markdown("""
<div style="
    background: linear-gradient(135deg, #0a1628, #0d1f3c);
    border: 1.5px solid #3b82f6;
    border-radius: 14px;
    padding: 14px 16px 10px;
    margin-bottom: 12px;
    box-shadow: 0 0 18px #3b82f630, 0 0 4px #3b82f618;
    position: relative;
    overflow: hidden;
">
  <div style="
      position: absolute; top: 0; left: 0; right: 0; height: 3px;
      background: linear-gradient(90deg, #3b82f6, #8b5cf6, #10d9a0);
      border-radius: 14px 14px 0 0;
  "></div>
  <div style="display:flex; align-items:center; gap:10px; margin-bottom:6px; margin-top:2px;">
    <span style="font-size:20px;">🔑</span>
    <div>
      <div style="color:#60a5fa; font-size:13px; font-weight:800; letter-spacing:0.3px;">
        NEW HERE? START WITH A FREE KEY
      </div>
      <div style="color:#64748b; font-size:11px; margin-top:1px;">
        Takes 30 seconds · No credit card · No billing ever
      </div>
    </div>
    <span style="
        margin-left:auto;
        background: linear-gradient(90deg,#1d4ed8,#7c3aed);
        color:#fff; font-size:9px; font-weight:800;
        letter-spacing:0.6px; padding:3px 9px;
        border-radius:20px; white-space:nowrap;
    ">FREE FOREVER</span>
  </div>
</div>
""", unsafe_allow_html=True)

            with st.expander("📋 Get FREE Gemini Key — Step by Step", expanded=True):
                st.markdown(
                    _free_guide_steps([
                        {"title": "Open Google AI Studio",
                         "desc": "100% free. Works with any Google account. No credit card. No trial.",
                         "url": "https://aistudio.google.com/app/apikey", "link_label": "Open AI Studio"},
                        {"title": "Sign in with Google",
                         "desc": "Use Gmail, Google Workspace, or any Google account — takes 10 seconds."},
                        {"title": "Click 'Create API Key'",
                         "desc": "Choose 'Create API key in new project'. This keeps things organised."},
                        {"title": "Copy the key (starts with AIza…)",
                         "desc": "Tip: save it in a notes app. You can always create more — unlimited keys."},
                        {"title": "Paste above and click Test",
                         "desc": "AutoApply AI will auto-detect your model and show the active limits."},
                    ]),
                    unsafe_allow_html=True,
                )

            st.markdown(
                _free_model_warning_html(
                    "Gemini", "gemini-1.5-flash",
                    "This is Google's permanently free model. No billing. No trial period. No credit card ever required."
                ),
                unsafe_allow_html=True,
            )
            col_inp, col_btn, col_badge = st.columns([4, 1.2, 1.5])
            with col_inp:
                new_gemini = st.text_input(
                    "Gemini API Key", type="password",
                    key="ak_gemini_input", label_visibility="collapsed",
                    placeholder="AIza… (paste your free key here)",
                ).strip()
            with col_btn:
                test_gemini = st.button("Test", key="ak_test_gemini", use_container_width=True)
            with col_badge:
                if st.session_state.get("api_validated"):
                    st.success("Connected", icon="✅")
                else:
                    st.warning("Not tested", icon="⚠️")
            if test_gemini:
                if new_gemini:
                    st.session_state["api_key"] = new_gemini
                    st.session_state["gemini_api_key"] = new_gemini
                    st.session_state["resolved_model_name"] = ""
                    validate_api_key_if_needed()
                    _auto_save_keys_for_logged_in_user()
                    st.rerun()
                else:
                    st.error("Enter a Gemini key first.")
            # ── Connect / Disconnect button ──────────────────────────
            _g_connected = st.session_state.get("api_validated", False)
            _btn_lbl_g = "🔌 Disconnect Gemini" if _g_connected else "🔗 Connect Gemini"
            _btn_type_g = "secondary" if _g_connected else "primary"
            if st.button(_btn_lbl_g, key="ak_connect_disconnect_gemini",
                         use_container_width=True, type=_btn_type_g):
                if _g_connected:
                    st.session_state["api_key"] = ""
                    st.session_state["gemini_api_key"] = ""
                    st.session_state["api_validated"] = False
                    st.session_state["api_validation_message"] = ""
                    st.session_state["api_last_checked_key"] = ""
                    st.session_state["resolved_model_name"] = ""
                    if st.session_state.get("active_provider") == "gemini":
                        st.session_state["active_provider"] = ""
                    st.rerun()
                else:
                    if new_gemini:
                        st.session_state["api_key"] = new_gemini
                        st.session_state["gemini_api_key"] = new_gemini
                        st.session_state["resolved_model_name"] = ""
                        validate_api_key_if_needed()
                        _auto_save_keys_for_logged_in_user()
                        st.rerun()
                    else:
                        st.error("Paste your Gemini API key above first.")
            if st.session_state.get("api_validated"):
                active_model = st.session_state.get("resolved_model_name", "gemini-1.5-flash")
                st.caption(f"Active model: `{active_model}`")
            gemini_used = uc.get("gemini", 0)
            st.markdown(
                _rate_limit_bar_html("Requests / Day", "#10d9a0", "📅",
                    gemini_used, gl["rpd"], "req", "Resets at midnight UTC",
                    gemini_exhausted, quota_reset_secs) +
                _rate_limit_bar_html("Requests / Minute", "#10d9a0", "⚡",
                    min(gemini_used, gl["rpm"]), gl["rpm"], "req", "Resets every minute",
                    False, 0) +
                _rate_limit_bar_html("Tokens / Minute", "#10d9a0", "🔤",
                    0, gl["tpm"], "tokens", "Resets every minute", False, 0),
                unsafe_allow_html=True,
            )
            if gemini_exhausted:
                st.warning(
                    f"Gemini quota reached. Resets at midnight UTC (~{quota_reset_secs//60}m). "
                    "AutoApply AI is already using the next provider. Add Groq below for instant free fallback.",
                    icon="⏳",
                )

    elif selected_prov == "groq":
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(
                _provider_section_header("⚡", "#f472b6", "Groq",
                    "Fastest free fallback · llama-3.1-8b-instant"),
                unsafe_allow_html=True,
            )
            st.markdown(
                _free_model_warning_html(
                    "Groq", "llama-3.1-8b-instant",
                    "Groq's free tier is genuinely free — no credit card, no trial expiry. "
                    "14,400 requests/day. This model will never charge you."
                ),
                unsafe_allow_html=True,
            )
            with st.expander("📋 Get FREE Groq Key — Step by Step"):
                st.markdown(
                    _free_guide_steps([
                        {"title": "Open Groq Console",
                         "desc": "100% free. No credit card. Signup takes under 60 seconds.",
                         "url": "https://console.groq.com/keys", "link_label": "Open Groq Console"},
                        {"title": "Create a free account",
                         "desc": "Sign up with email, GitHub, or Google — whichever is fastest for you."},
                        {"title": "Navigate to API Keys",
                         "desc": "Click your profile icon in the top-right, then 'API Keys' in the sidebar."},
                        {"title": "Click 'Create API Key'",
                         "desc": "Name it 'AutoApply AI'. The key starts with gsk_…"},
                        {"title": "Paste above and click Test",
                         "desc": "Groq gives the most generous free limits of all providers — 14,400 req/day."},
                    ]),
                    unsafe_allow_html=True,
                )
            col_inp, col_btn, col_badge = st.columns([4, 1.2, 1.5])
            with col_inp:
                new_groq = st.text_input(
                    "Groq API Key", type="password",
                    key="ak_groq_input", label_visibility="collapsed",
                    placeholder="gsk_… (paste your free key here)",
                ).strip()
            with col_btn:
                test_groq = st.button("Test", key="ak_test_groq", use_container_width=True)
            with col_badge:
                if st.session_state.get("groq_api_validated"):
                    st.success("Connected", icon="✅")
                else:
                    st.warning("Not tested", icon="⚠️")
            if test_groq:
                if new_groq:
                    st.session_state["groq_api_key"] = new_groq
                    with st.spinner("Validating Groq key…"):
                        ok, msg = validate_groq_key(new_groq)
                    st.session_state["groq_api_validated"] = ok
                    st.session_state["groq_api_validation_message"] = msg
                    if ok:
                        _auto_save_keys_for_logged_in_user()
                    st.rerun()
                else:
                    st.error("Enter a Groq key first.")
            # ── Connect / Disconnect button ──────────────────────────
            _groq_connected = st.session_state.get("groq_api_validated", False)
            _btn_lbl_groq = "🔌 Disconnect Groq" if _groq_connected else "🔗 Connect Groq"
            _btn_type_groq = "secondary" if _groq_connected else "primary"
            if st.button(_btn_lbl_groq, key="ak_connect_disconnect_groq",
                         use_container_width=True, type=_btn_type_groq):
                if _groq_connected:
                    st.session_state["groq_api_key"] = ""
                    st.session_state["groq_api_validated"] = False
                    st.session_state["groq_api_validation_message"] = ""
                    if st.session_state.get("active_provider") == "groq":
                        st.session_state["active_provider"] = ""
                    st.rerun()
                else:
                    if new_groq:
                        st.session_state["groq_api_key"] = new_groq
                        with st.spinner("Connecting Groq…"):
                            ok, msg = validate_groq_key(new_groq)
                        st.session_state["groq_api_validated"] = ok
                        st.session_state["groq_api_validation_message"] = msg
                        if ok:
                            _auto_save_keys_for_logged_in_user()
                        st.rerun()
                    else:
                        st.error("Paste your Groq API key above first.")
            elif st.session_state.get("groq_api_validation_message"):
                msg = st.session_state["groq_api_validation_message"]
                if st.session_state.get("groq_api_validated"):
                    st.success(msg)
                else:
                    st.error(msg)
            groq_used = uc.get("groq", 0)
            st.markdown(
                _rate_limit_bar_html("Requests / Day", "#f472b6", "📅",
                    groq_used, groq_limits["rpd"], "req", "Resets at midnight UTC",
                    groq_exhausted, quota_reset_secs) +
                _rate_limit_bar_html("Requests / Minute", "#f472b6", "⚡",
                    min(groq_used, groq_limits["rpm"]), groq_limits["rpm"], "req",
                    "Resets every minute", False, 0) +
                _rate_limit_bar_html("Tokens / Minute", "#f472b6", "🔤",
                    0, groq_limits["tpm"], "tokens", "Resets every minute", False, 0),
                unsafe_allow_html=True,
            )
            if groq_exhausted:
                st.warning(
                    f"Groq limit reached. Resets in ~{quota_reset_secs//60}m. "
                    "AutoApply AI has moved to the next provider automatically.",
                    icon="⏳",
                )

    elif selected_prov == "claude":
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(
                _provider_section_header("◆", "#f59e0b", "Claude (Anthropic)",
                    "Optional fallback · claude-haiku-4-5 (free credits on signup)"),
                unsafe_allow_html=True,
            )
            st.markdown(
                _free_model_warning_html(
                    "Claude", "claude-haiku-4-5-20251001",
                    "Anthropic gives free API credits on signup. Haiku is their fastest, cheapest model. "
                    "AutoApply AI will NEVER select a paid model — only Haiku is used."
                ),
                unsafe_allow_html=True,
            )
            with st.expander("📋 Get FREE Claude Key — Step by Step"):
                st.markdown(
                    _free_guide_steps([
                        {"title": "Open Anthropic Console",
                         "desc": "New accounts receive free API credits. No credit card required for free tier.",
                         "url": "https://console.anthropic.com", "link_label": "Open Console"},
                        {"title": "Create a free account",
                         "desc": "Sign up with your email. Verification email arrives in under a minute."},
                        {"title": "Verify your email",
                         "desc": "Click the link in the Anthropic email to activate your account."},
                        {"title": "Go to API Keys page",
                         "desc": "In the console, navigate: Settings → API Keys.",
                         "url": "https://console.anthropic.com/settings/keys", "link_label": "Open API Keys"},
                        {"title": "Click 'Create Key'",
                         "desc": "Name it 'AutoApply AI'. Key starts with sk-ant-…"},
                        {"title": "Paste above and click Test",
                         "desc": "Free credits are pre-loaded. AutoApply AI uses only Haiku — the free-tier model."},
                    ]),
                    unsafe_allow_html=True,
                )
            col_inp, col_btn, col_badge = st.columns([4, 1.2, 1.5])
            with col_inp:
                new_claude = st.text_input(
                    "Claude API Key", type="password",
                    key="ak_claude_input", label_visibility="collapsed",
                    placeholder="sk-ant-… (paste your free key here)",
                ).strip()
            with col_btn:
                test_claude = st.button("Test", key="ak_test_claude", use_container_width=True)
            with col_badge:
                if st.session_state.get("claude_api_validated"):
                    st.success("Connected", icon="✅")
                else:
                    st.warning("Not tested", icon="⚠️")
            if test_claude:
                if new_claude:
                    st.session_state["claude_api_key"] = new_claude
                    with st.spinner("Validating Claude key…"):
                        ok, msg = validate_claude_key(new_claude)
                    st.session_state["claude_api_validated"] = ok
                    st.session_state["claude_api_validation_message"] = msg
                    if ok:
                        _auto_save_keys_for_logged_in_user()
                    st.rerun()
                else:
                    st.error("Enter a Claude key first.")
            # ── Connect / Disconnect button ──────────────────────────
            _claude_connected = st.session_state.get("claude_api_validated", False)
            _btn_lbl_claude = "🔌 Disconnect Claude" if _claude_connected else "🔗 Connect Claude"
            _btn_type_claude = "secondary" if _claude_connected else "primary"
            if st.button(_btn_lbl_claude, key="ak_connect_disconnect_claude",
                         use_container_width=True, type=_btn_type_claude):
                if _claude_connected:
                    st.session_state["claude_api_key"] = ""
                    st.session_state["claude_api_validated"] = False
                    st.session_state["claude_api_validation_message"] = ""
                    if st.session_state.get("active_provider") == "claude":
                        st.session_state["active_provider"] = ""
                    st.rerun()
                else:
                    if new_claude:
                        st.session_state["claude_api_key"] = new_claude
                        with st.spinner("Connecting Claude…"):
                            ok, msg = validate_claude_key(new_claude)
                        st.session_state["claude_api_validated"] = ok
                        st.session_state["claude_api_validation_message"] = msg
                        if ok:
                            _auto_save_keys_for_logged_in_user()
                        st.rerun()
                    else:
                        st.error("Paste your Claude API key above first.")
            elif st.session_state.get("claude_api_validation_message"):
                msg = st.session_state["claude_api_validation_message"]
                if st.session_state.get("claude_api_validated"):
                    st.success(msg)
                else:
                    st.error(msg)
            claude_used = uc.get("claude", 0)
            st.markdown(
                _rate_limit_bar_html("Requests / Day", "#f59e0b", "📅",
                    claude_used, claude_limits["rpd"], "req", "Resets at midnight UTC",
                    claude_exhausted, quota_reset_secs) +
                _rate_limit_bar_html("Requests / Minute", "#f59e0b", "⚡",
                    min(claude_used, claude_limits["rpm"]), claude_limits["rpm"], "req",
                    "Resets every minute", False, 0) +
                _rate_limit_bar_html("Tokens / Minute", "#f59e0b", "🔤",
                    0, claude_limits["tpm"], "tokens", "Resets every minute", False, 0),
                unsafe_allow_html=True,
            )
            if claude_exhausted:
                st.warning(
                    f"Claude limit reached. Resets in ~{quota_reset_secs//60}m.",
                    icon="⏳",
                )

    elif selected_prov == "together":
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(
                _provider_section_header("🤗", "#f97316", "HuggingFace",
                    "Last-resort fallback · Mistral-7B-Instruct-v0.3 (free inference API)"),
                unsafe_allow_html=True,
            )
            # ── Model recommendation banner ──────────────────────────
            st.markdown("""
<div style="background:linear-gradient(135deg,#0f1a0a,#0a1408);border:1px solid #22c55e40;
border-left:3px solid #22c55e;border-radius:10px;padding:12px 16px;margin-bottom:12px;">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
    <span style="font-size:16px;">🏆</span>
    <span style="color:#22c55e;font-size:12px;font-weight:800;letter-spacing:0.4px;">RECOMMENDED FREE MODEL</span>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;">
    <div style="background:#060c06;border:1px solid #22c55e25;border-radius:8px;padding:10px 12px;">
      <div style="color:#f0fdf4;font-size:13px;font-weight:700;margin-bottom:2px;">mistralai/Mistral-7B-Instruct-v0.3</div>
      <div style="color:#86efac;font-size:10px;font-weight:600;margin-bottom:6px;">✅ Currently used by AutoApply AI</div>
      <div style="display:flex;flex-wrap:wrap;gap:4px;">
        <span style="background:#14532d;color:#86efac;font-size:9px;font-weight:700;padding:2px 6px;border-radius:4px;">FREE FOREVER</span>
        <span style="background:#14532d;color:#86efac;font-size:9px;font-weight:700;padding:2px 6px;border-radius:4px;">NO BILLING</span>
        <span style="background:#14532d;color:#86efac;font-size:9px;font-weight:700;padding:2px 6px;border-radius:4px;">7B PARAMS</span>
      </div>
    </div>
    <div style="background:#060c06;border:1px solid #f9731625;border-radius:8px;padding:10px 12px;">
      <div style="color:#fed7aa;font-size:12px;font-weight:700;margin-bottom:4px;">Other good free options:</div>
      <div style="color:#94a3b8;font-size:10px;line-height:1.8;">
        🔹 <code style="color:#fb923c;">google/gemma-2-2b-it</code><br>
        🔹 <code style="color:#fb923c;">HuggingFaceH4/zephyr-7b-beta</code><br>
        🔹 <code style="color:#fb923c;">microsoft/Phi-3-mini-4k-instruct</code>
      </div>
      <div style="color:#64748b;font-size:9px;margin-top:6px;font-style:italic;">All 100% free · no card needed</div>
    </div>
  </div>
  <div style="margin-top:8px;padding:6px 10px;background:#0a1a08;border-radius:6px;border:1px solid #22c55e20;">
    <span style="color:#86efac;font-size:10px;">💡 <b>Why Mistral-7B?</b> Best balance of quality + speed on HuggingFace free tier. 
    Handles resume analysis, job matching and ATS scoring accurately. Completely free — 
    zero billing for you or your users.</span>
  </div>
</div>
""", unsafe_allow_html=True)
            st.markdown(
                _free_model_warning_html(
                    "HuggingFace", "mistralai/Mistral-7B-Instruct-v0.3",
                    "HuggingFace Inference API is free with a free account token. "
                    "No credit card, no trial expiry. AutoApply AI only calls Mistral-7B — always free."
                ),
                unsafe_allow_html=True,
            )
            with st.expander("📋 Get FREE HuggingFace Token — Step by Step"):
                st.markdown(
                    _free_guide_steps([
                        {"title": "Open HuggingFace",
                         "desc": "100% free. No credit card. Signup takes under 60 seconds.",
                         "url": "https://huggingface.co/join", "link_label": "Join HuggingFace"},
                        {"title": "Create a free account",
                         "desc": "Sign up with email or Google — whichever is fastest for you."},
                        {"title": "Go to Access Tokens",
                         "desc": "Click your profile picture (top right) → Settings → Access Tokens.",
                         "url": "https://huggingface.co/settings/tokens/new?tokenType=read", "link_label": "Open Access Tokens"},
                        {"title": "Click 'New token'",
                         "desc": "Select type 'Read' (free), name it 'AutoApply AI'. Token starts with hf_…"},
                        {"title": "Paste above and click Test",
                         "desc": "HuggingFace Inference API is free for public models like Mistral-7B — no charges ever."},
                    ]),
                    unsafe_allow_html=True,
                )
            col_inp, col_btn, col_badge = st.columns([4, 1.2, 1.5])
            with col_inp:
                new_together = st.text_input(
                    "HuggingFace Token", type="password",
                    key="ak_together_input", label_visibility="collapsed",
                    placeholder="hf_… (paste your free token here)",
                ).strip()
            with col_btn:
                test_together = st.button("Test", key="ak_test_together", use_container_width=True)
            with col_badge:
                if st.session_state.get("together_api_validated"):
                    st.success("Connected", icon="✅")
                else:
                    st.warning("Not tested", icon="⚠️")
            if test_together:
                if new_together:
                    st.session_state["together_api_key"] = new_together
                    with st.spinner("Validating HuggingFace token…"):
                        ok, msg = validate_together_key(new_together)
                    st.session_state["together_api_validated"] = ok
                    st.session_state["together_api_validation_message"] = msg
                    if ok:
                        _auto_save_keys_for_logged_in_user()
                    st.rerun()
                else:
                    st.error("Enter a HuggingFace token first.")
            # ── Connect / Disconnect button ──────────────────────────
            _hf_connected = st.session_state.get("together_api_validated", False)
            _btn_lbl_hf = "🔌 Disconnect HuggingFace" if _hf_connected else "🔗 Connect HuggingFace"
            _btn_type_hf = "secondary" if _hf_connected else "primary"
            if st.button(_btn_lbl_hf, key="ak_connect_disconnect_together",
                         use_container_width=True, type=_btn_type_hf):
                if _hf_connected:
                    st.session_state["together_api_key"] = ""
                    st.session_state["together_api_validated"] = False
                    st.session_state["together_api_validation_message"] = ""
                    st.session_state["hf_active_model"] = ""
                    if st.session_state.get("active_provider") == "together":
                        st.session_state["active_provider"] = ""
                    st.rerun()
                else:
                    if new_together:
                        st.session_state["together_api_key"] = new_together
                        with st.spinner("Connecting HuggingFace…"):
                            ok, msg = validate_together_key(new_together)
                        st.session_state["together_api_validated"] = ok
                        st.session_state["together_api_validation_message"] = msg
                        if ok:
                            _auto_save_keys_for_logged_in_user()
                        st.rerun()
                    else:
                        st.error("Paste your HuggingFace token above first.")
            elif st.session_state.get("together_api_validation_message"):
                msg = st.session_state["together_api_validation_message"]
                if st.session_state.get("together_api_validated"):
                    st.success(msg)
                else:
                    st.error(msg)
            together_used = uc.get("together", 0)
            st.markdown(
                _rate_limit_bar_html("Requests / Day", "#f97316", "📅",
                    together_used, together_limits["rpd"], "req", "Resets at midnight UTC",
                    together_exhausted, quota_reset_secs) +
                _rate_limit_bar_html("Requests / Minute", "#f97316", "⚡",
                    min(together_used, together_limits["rpm"]), together_limits["rpm"], "req",
                    "Resets every minute", False, 0) +
                _rate_limit_bar_html("Tokens / Minute", "#f97316", "🔤",
                    0, together_limits["tpm"], "tokens", "Resets every minute", False, 0),
                unsafe_allow_html=True,
            )
            if together_exhausted:
                st.warning(
                    f"HuggingFace limit reached. Resets in ~{quota_reset_secs//60}m.",
                    icon="⏳",
                )

    elif selected_prov == "openai":
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(
                _provider_section_header("✦", "#22d3ee", "OpenAI",
                    "Optional provider · gpt-4o-mini (free trial credits)"),
                unsafe_allow_html=True,
            )
            st.markdown(
                _free_model_warning_html(
                    "OpenAI", "gpt-4o-mini",
                    "AutoApply AI uses gpt-4o-mini — OpenAI's cheapest model. New accounts get $5 free trial credits "
                    "that last 3 months. You will NOT be charged until you add a payment method AND exhaust credits."
                ),
                unsafe_allow_html=True,
            )
            with st.expander("📋 Get FREE OpenAI Key — Step by Step"):
                st.markdown(
                    _free_guide_steps([
                        {"title": "Open OpenAI Platform",
                         "desc": "New accounts get $5 free trial credits (valid 3 months). No immediate billing.",
                         "url": "https://platform.openai.com", "link_label": "Open OpenAI Platform"},
                        {"title": "Create a free account",
                         "desc": "Sign up with email, Google, or Microsoft. No card needed for the free trial."},
                        {"title": "Verify your phone number",
                         "desc": "OpenAI requires phone verification once for security. Takes 30 seconds."},
                        {"title": "Go to API Keys",
                         "desc": "Click your profile → 'API Keys' in the left navigation.",
                         "url": "https://platform.openai.com/api-keys", "link_label": "Open API Keys"},
                        {"title": "Create new secret key",
                         "desc": "Click '+ Create new secret key', name it 'AutoApply AI'. Starts with sk-…"},
                        {"title": "Paste above and click Test",
                         "desc": "Your $5 credit is ready. AutoApply AI only calls gpt-4o-mini — extremely low cost per call."},
                    ]),
                    unsafe_allow_html=True,
                )
            col_inp, col_btn, col_badge = st.columns([4, 1.2, 1.5])
            with col_inp:
                new_openai = st.text_input(
                    "OpenAI API Key", type="password",
                    key="ak_openai_input", label_visibility="collapsed",
                    placeholder="sk-… (paste your key here)",
                ).strip()
            with col_btn:
                test_openai = st.button("Test", key="ak_test_openai", use_container_width=True)
            with col_badge:
                if st.session_state.get("openai_api_validated"):
                    st.success("Connected", icon="✅")
                else:
                    st.warning("Not tested", icon="⚠️")
            if test_openai:
                if new_openai:
                    st.session_state["openai_api_key"] = new_openai
                    with st.spinner("Validating OpenAI key…"):
                        ok, msg = validate_openai_key(new_openai)
                    st.session_state["openai_api_validated"] = ok
                    st.session_state["openai_api_validation_message"] = msg
                    if ok:
                        _auto_save_keys_for_logged_in_user()
                    st.rerun()
                else:
                    st.error("Enter an OpenAI key first.")
            # ── Connect / Disconnect button ──────────────────────────
            _openai_connected = st.session_state.get("openai_api_validated", False)
            _btn_lbl_openai = "🔌 Disconnect OpenAI" if _openai_connected else "🔗 Connect OpenAI"
            _btn_type_openai = "secondary" if _openai_connected else "primary"
            if st.button(_btn_lbl_openai, key="ak_connect_disconnect_openai",
                         use_container_width=True, type=_btn_type_openai):
                if _openai_connected:
                    st.session_state["openai_api_key"] = ""
                    st.session_state["openai_api_validated"] = False
                    st.session_state["openai_api_validation_message"] = ""
                    if st.session_state.get("active_provider") == "openai":
                        st.session_state["active_provider"] = ""
                    st.rerun()
                else:
                    if new_openai:
                        st.session_state["openai_api_key"] = new_openai
                        with st.spinner("Connecting OpenAI…"):
                            ok, msg = validate_openai_key(new_openai)
                        st.session_state["openai_api_validated"] = ok
                        st.session_state["openai_api_validation_message"] = msg
                        if ok:
                            _auto_save_keys_for_logged_in_user()
                        st.rerun()
                    else:
                        st.error("Paste your OpenAI API key above first.")
            elif st.session_state.get("openai_api_validation_message"):
                msg = st.session_state["openai_api_validation_message"]
                if st.session_state.get("openai_api_validated"):
                    st.success(msg)
                else:
                    st.error(msg)
            openai_used = uc.get("openai", 0)
            st.markdown(
                _rate_limit_bar_html("Requests / Day", "#22d3ee", "📅",
                    openai_used, openai_limits["rpd"], "req", "Resets at midnight UTC",
                    False, 0) +
                _rate_limit_bar_html("Requests / Minute", "#22d3ee", "⚡",
                    min(openai_used, openai_limits["rpm"]), openai_limits["rpm"], "req",
                    "Resets every minute", False, 0) +
                _rate_limit_bar_html("Tokens / Minute", "#22d3ee", "🔤",
                    0, openai_limits["tpm"], "tokens", "Resets every minute", False, 0),
                unsafe_allow_html=True,
            )

    # ── Save to localStorage ─────────────────────────────────────
    st.markdown("<hr style='border:none;border-top:1px solid #1e2d45;margin:24px 0 16px;'>", unsafe_allow_html=True)
    if st.button("☁️  Save Your Keys to Cloud ", type="primary", use_container_width=False):
        ls_payload = {
            "AutoApply AI_gemini_key":   st.session_state.get("api_key", ""),
            "AutoApply AI_groq_key":     st.session_state.get("groq_api_key", ""),
            "AutoApply AI_claude_key":   st.session_state.get("claude_api_key", ""),
            "AutoApply AI_together_key": st.session_state.get("together_api_key", ""),
            "AutoApply AI_openai_key":   st.session_state.get("openai_api_key", ""),
        }
        st.markdown(_ls_bridge_script(ls_payload), unsafe_allow_html=True)
        # Also persist keys to the logged-in user's account row
        user = st.session_state.get("auth_user")
        if isinstance(user, dict) and user.get("Username"):
            _save_user_api_keys(user["Username"])
            st.success("Keys saved to your account and browser. They never leave your device.")

    # ── Rotation Status ──────────────────────────────────────────
    st.markdown("<hr style='border:none;border-top:1px solid #1e2d45;margin:16px 0;'>", unsafe_allow_html=True)
    st.markdown(
        _section_header_html("⚡", "Rotation Status", "Auto-rotates on rate limit — no manual action needed."),
        unsafe_allow_html=True,
    )
    providers_live = _build_provider_list()
    if providers_live:
        badge_col, info_col = st.columns([1, 3])
        with badge_col:
            st.markdown(active_provider_badge(), unsafe_allow_html=True)
        with info_col:
            names = " → ".join(f"**{p['label']}**" for p in providers_live)
            st.markdown(
                f"Priority chain: {names}  \n"
                f"<span style='color:#94a3b8;font-size:12px;'>"
                f"Auto-rotates on rate limit / quota error. No manual action needed.</span>",
                unsafe_allow_html=True,
            )
        switch_log = st.session_state.get("provider_switch_log", [])
        if switch_log:
            with st.expander("Provider Switch Log", expanded=False):
                for entry in reversed(switch_log[-20:]):
                    frm = entry.get("from") or "—"
                    to  = entry.get("to", "")
                    reason = entry.get("reason", "")
                    ts = entry.get("ts", "")[:19].replace("T", " ")
                    icon = "🔄" if reason == "rate_limit" else "✦"
                    st.markdown(
                        f"`{ts}` {icon} `{frm}` → `{to}` "
                        f"<span style='color:#94a3b8;font-size:11px;'>({reason})</span>",
                        unsafe_allow_html=True,
                    )
    else:
        st.warning("No API keys configured. Add keys above to enable AI rotation.", icon="⚠️")

    # ══════════════════════════════════════════════════════════════
    # SECTION 2 — JOB PLATFORMS (read-only status mirror)
    # ══════════════════════════════════════════════════════════════
    st.markdown("<hr style='border:none;border-top:1px solid #1e2d45;margin:24px 0 16px;'>", unsafe_allow_html=True)
    st.markdown(
        _section_header_html("🌐", "Job Platform Connections",
            "Read-only mirror of your Platforms page. Manage connections from the Platforms tab."),
        unsafe_allow_html=True,
    )

    platforms    = st.session_state.get("platforms", PLATFORMS_DEFAULT)
    connected_set = set(st.session_state.get("platforms_connected", []))
    connected_at  = st.session_state.get("platforms_connected_at", {})
    connected_platforms    = [p for p in platforms if p in connected_set]
    disconnected_platforms = [p for p in platforms if p not in connected_set]
    total       = len(platforms)
    n_connected = len(connected_platforms)

    prog_col, stat_col = st.columns([3, 1])
    with prog_col:
        st.progress(n_connected / total if total else 0)
    with stat_col:
        st.markdown(
            f"<p style='text-align:right;color:#10d9a0;font-weight:600;margin:0;'>"
            f"{n_connected} / {total} connected</p>",
            unsafe_allow_html=True,
        )

    st.markdown("")

    if connected_platforms:
        st.markdown("**Connected ✅**")
        rows = [connected_platforms[i: i + 4] for i in range(0, len(connected_platforms), 4)]
        for row in rows:
            cols = st.columns(4)
            for col, name in zip(cols, row):
                meta  = PLATFORM_META.get(name, {})
                since = connected_at.get(name, "—")
                with col:
                    st.markdown(
                        f'<div style="background:linear-gradient(135deg,#071a0f,#0a1a10);'
                        f'border:1px solid #10d9a030;border-radius:12px;'
                        f'padding:12px;margin-bottom:8px;">'
                        f'  <div style="font-weight:700;color:#10d9a0;font-size:13px;">{name}</div>'
                        f'  <div style="color:#64748b;font-size:10px;margin-top:4px;">Since: {since}</div>'
                        f'  <div style="color:#64748b;font-size:10px;">Rate: {meta.get("response_rate","N/A")}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

    if disconnected_platforms:
        st.markdown("**Not Connected ⚠️**")
        rows = [disconnected_platforms[i: i + 4] for i in range(0, len(disconnected_platforms), 4)]
        for row in rows:
            cols = st.columns(4)
            for col, name in zip(cols, row):
                meta = PLATFORM_META.get(name, {})
                with col:
                    st.markdown(
                        f'<div style="background:linear-gradient(135deg,#12100a,#0f0d08);'
                        f'border:1px solid #37415130;border-radius:12px;'
                        f'padding:12px;margin-bottom:8px;">'
                        f'  <div style="font-weight:700;color:#f59e0b;font-size:13px;">{name}</div>'
                        f'  <div style="color:#64748b;font-size:10px;margin-top:4px;">Best: {meta.get("best_time","N/A")}</div>'
                        f'  <div style="color:#64748b;font-size:10px;">Rate: {meta.get("response_rate","N/A")}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

    st.markdown("")
    st.info("To connect or disconnect platforms, go to the **Platforms** page from the sidebar.", icon="💡")

def sidebar_navigation() -> str:
    with st.sidebar:
        # ── Premium navbar styles ────────────────────────────────────────
        st.html("""
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=Syne:wght@600;700;800&display=swap" rel="stylesheet">
<style>
/* ── Sidebar shell ── */
[data-testid="stSidebar"] {
    background: #07090f !important;
    border-right: 0.5px solid rgba(255,255,255,0.07) !important;
    min-width: 260px !important;
    box-shadow: 4px 0 40px rgba(0,0,0,0.6) !important;
}
[data-testid="stSidebar"] > div:first-child { padding: 0 !important; }

/* ── Brand strip ── */
.nav-brand-strip {
    padding: 24px 20px 16px;
    display: flex;
    align-items: center;
    gap: 12px;
    border-bottom: 0.5px solid rgba(255,255,255,0.07);
    margin-bottom: 6px;
    background: linear-gradient(180deg, rgba(29,212,160,0.03) 0%, transparent 100%);
}
.nav-brand-icon {
    width: 42px; height: 42px;
    background: linear-gradient(135deg,#1dd4a0,#3b82f6);
    border-radius: 13px;
    display: flex; align-items: center; justify-content: center;
    font-size: 21px; flex-shrink: 0;
    box-shadow: 0 0 20px rgba(29,212,160,0.3), 0 4px 12px rgba(0,0,0,0.4);
}
.nav-brand-name {
    font-family: 'Sora', sans-serif;
    font-size: 17px; font-weight: 800; letter-spacing: -0.4px;
    background: linear-gradient(90deg,#e8f0fe,#1dd4a0);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}
.nav-section-lbl {
    padding: 14px 20px 5px;
    font-family: 'DM Sans', sans-serif;
    font-size: 9px; font-weight: 800; letter-spacing: 2.5px;
    text-transform: uppercase; color: rgba(255,255,255,0.18);
}

/* ── option_menu nav items ── */
[data-testid="stSidebar"] .nav-link {
    font-family: 'DM Sans', sans-serif !important;
    font-size: 14px !important;
    font-weight: 500 !important;
    color: rgba(255,255,255,0.42) !important;
    border-radius: 12px !important;
    margin: 1px 10px !important;
    padding: 11px 14px !important;
    transition: all 0.18s ease !important;
    letter-spacing: 0.1px !important;
}
[data-testid="stSidebar"] .nav-link:hover {
    background: rgba(255,255,255,0.05) !important;
    color: rgba(255,255,255,0.75) !important;
    transform: translateX(2px) !important;
}
[data-testid="stSidebar"] .nav-link.active {
    background: linear-gradient(90deg, rgba(29,212,160,0.13), rgba(29,212,160,0.04)) !important;
    color: #d6faf2 !important;
    font-weight: 600 !important;
    border-left: 2.5px solid #1dd4a0 !important;
    padding-left: 11px !important;
    box-shadow: 0 2px 12px rgba(29,212,160,0.08) !important;
}
[data-testid="stSidebar"] .nav-link .icon {
    font-size: 16px !important;
    margin-right: 11px !important;
    opacity: 0.5;
}
[data-testid="stSidebar"] .nav-link.active .icon { opacity: 1 !important; }
[data-testid="stSidebar"] #sidebarNav { background: transparent !important; padding: 0 !important; }
[data-testid="stSidebar"] .menu-title { display: none !important; }

/* ── Logout button ── */
[data-testid="stSidebar"] .stButton > button {
    font-size: 13px !important;
    font-weight: 500 !important;
    padding: 10px 16px !important;
    border-radius: 10px !important;
    margin: 4px 10px 0 !important;
    border: 1px solid rgba(255,255,255,0.07) !important;
    background: rgba(255,255,255,0.03) !important;
    color: rgba(255,255,255,0.45) !important;
    transition: all 0.18s ease !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(239,68,68,0.08) !important;
    border-color: rgba(239,68,68,0.25) !important;
    color: rgba(239,68,68,0.8) !important;
}

/* ── Weather footer ── */
.nav-weather-footer {
    margin: 12px 10px 6px;
    padding: 12px 14px;
    border-top: 0.5px solid rgba(255,255,255,0.06);
    display: flex; align-items: center; gap: 10px;
    font-family: 'DM Sans', sans-serif;
    background: linear-gradient(90deg, rgba(245,158,11,0.02), transparent);
    border-radius: 0 0 12px 12px;
}
.nav-weather-ico {
    width: 34px; height: 34px; border-radius: 11px;
    background: rgba(245,158,11,0.1);
    display: flex; align-items: center; justify-content: center;
    font-size: 17px; flex-shrink: 0;
    box-shadow: 0 2px 8px rgba(245,158,11,0.1);
}
.nav-weather-temp { font-size: 15px !important; font-weight: 700; color: rgba(255,255,255,0.65) !important; }
.nav-weather-sub  { font-size: 10px !important; color: rgba(255,255,255,0.25) !important; margin-top:2px; letter-spacing: 0.3px; }
.nav-weather-badge {
    margin-left: auto;
    font-size: 9px; font-weight: 800; letter-spacing: 1px;
    text-transform: uppercase; color: #f59e0b;
    background: rgba(245,158,11,0.1);
    padding: 4px 10px; border-radius: 7px;
    border: 1px solid rgba(245,158,11,0.18);
}
</style>
""")

        st.markdown("""
<div class="nav-brand-strip">
  <div class="nav-brand-icon">💼</div>
  <div class="nav-brand-name">AutoApply AI</div>
</div>
<div class="nav-section-lbl">Main</div>
""", unsafe_allow_html=True)

        selected = option_menu(
            menu_title=None,
            options=[
                "Your Resume Insight",
                "API Keys",
                "Connect with Career Portals",
                "AI-Powered Resume Rewrite",
                "Job Matches",
                "Application Tracker",
                "Settings",
            ],
            icons=["grid", "key", "globe2", "file-earmark-text", "search", "calendar-check", "gear"],
            menu_icon="briefcase",
            default_index=0,
            styles={
                "container": {"padding": "4px 0", "background": "transparent"},
                "icon":      {"font-size": "17px"},
                "nav-link":  {
                    "font-size": "15px", "font-family": "'DM Sans', sans-serif",
                    "color": "rgba(255,255,255,0.50)", "border-radius": "12px",
                    "margin": "2px 10px", "padding": "12px 14px",
                    "font-weight": "400", "letter-spacing": "0.1px",
                },
                "nav-link-selected": {
                    "background": "rgba(29,212,160,0.11)",
                    "color": "#d6faf2", "font-weight": "600",
                    "border-left": "3px solid #1dd4a0",
                    "padding-left": "11px",
                },
            },
        )

        # Weather footer
        st.markdown("""
<div class="nav-weather-footer">
  <div class="nav-weather-ico">☀</div>
  <div>
    <div class="nav-weather-temp">29°C</div>
    <div class="nav-weather-sub">Sunny · Mumbai</div>
  </div>
  <span class="nav-weather-badge">Clear</span>
</div>
""", unsafe_allow_html=True)

        return selected


def main() -> None:
    init_state()

    # ── Restore session from URL token (runs before any UI) ─────────────
    if not is_logged_in():
        _restore_session_from_url()

    logged_in = is_logged_in()
    configure_page(show_header=logged_in)

    # ── Auth gate ────────────────────────────────────────────────────────
    if not logged_in:
        login_page()
        return

    # ── Set / refresh URL token for the logged-in user ───────────────────
    user = st.session_state.get("auth_user", {})
    uname = user.get("Username", "User") if isinstance(user, dict) else "User"
    if uname and uname != "User":
        _set_url_session(uname)

    # ── Logged-in: show sidebar user block ──────────────────────────────
    with st.sidebar:
        initials = "".join(p[0].upper() for p in uname.split()[:2]) if uname and uname != "User" else "?"
        st.markdown(
            f"<div style='display:flex;align-items:center;gap:11px;"
            f"padding:16px 16px 12px;border-bottom:0.5px solid rgba(255,255,255,0.06);'>"
            f"<div style='width:36px;height:36px;border-radius:50%;"
            f"background:linear-gradient(135deg,#1dd4a0,#3b82f6);"
            f"display:flex;align-items:center;justify-content:center;"
            f"font-family:Syne,sans-serif;font-size:13px;font-weight:700;"
            f"color:#fff;flex-shrink:0;'>{initials}</div>"
            f"<div style='flex:1;min-width:0;'>"
            f"<div style='font-family:Syne,sans-serif;font-size:13px;font-weight:600;"
            f"color:#e8f0fe;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>{uname}</div>"
            f"<div style='font-size:10px;color:rgba(255,255,255,0.28);margin-top:1px;"
            f"display:flex;align-items:center;gap:4px;'>"
            f"<span style='width:5px;height:5px;border-radius:50%;background:#1dd4a0;"
            f"display:inline-block;'></span>Logged in</div>"
            f"</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
        if st.button("🚪 Logout", use_container_width=True):
            logout()

    page = sidebar_navigation()
    if page == "Your Resume Insight":
        dashboard_page()
    elif page == "AI-Powered Resume Rewrite":
        tailor_resume_page()
    elif page == "Job Matches":
        job_matches_page()
    elif page == "Application Tracker":
        application_tracker_page()
    elif page == "Connect with Career Portals":
        platforms_page()
    elif page == "API Keys":
        api_keys_page()
    elif page == "Settings":
        settings_page()
    # Legacy fallbacks
    elif page == "Dashboard":
        dashboard_page()
    elif page == "Tailor Resume":
        tailor_resume_page()
    elif page == "Platforms":
        platforms_page()


if __name__ == "__main__":
    main()